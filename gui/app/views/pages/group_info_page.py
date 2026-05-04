"""
Group Info page -- detailed view of a group conversation showing all members,
their roles, message contribution stats, and activity timelines.
Opened when a group conversation name/header is clicked.
"""

from __future__ import annotations

import base64
from datetime import datetime

from PySide6.QtCore import QRect, QSize, Qt, Signal
from PySide6.QtGui import QBrush, QColor, QFont, QPainter, QPainterPath, QPixmap
from PySide6.QtWidgets import (
    QAbstractItemView, QDialog, QFrame, QHBoxLayout, QHeaderView, QLabel,
    QLineEdit, QPushButton, QScrollArea, QSizePolicy, QStyledItemDelegate,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)

from app.config import format_timestamp as _fmt_ts
from app.services.database import Database
from app.services.theme_manager import ThemeManager

# Avatar background colors (same palette as conversations page)
AVATAR_COLORS = [
    "#00897b", "#6a1b9a", "#c62828", "#1565c0",
    "#ef6c00", "#2e7d32", "#ad1457", "#4527a0",
    "#00838f", "#827717", "#4e342e", "#37474f",
]

# Role display configuration
ROLE_CONFIG = {
    "superadmin": {"label": "Super Admin", "bg": "#f4511e", "fg": "#ffffff"},
    "admin":      {"label": "Admin",       "bg": "#ffb300", "fg": "#1a1a1a"},
    "member":     {"label": "Member",      "bg": "transparent", "fg": "#90a4ae"},
}

# Column indices
COL_AVATAR = 0
COL_NAME = 1
COL_PHONE = 2
COL_ROLE = 3
COL_LABEL = 4
COL_TOTAL = 5
COL_TEXT = 6
COL_IMAGE = 7
COL_VIDEO = 8
COL_AUDIO = 9
COL_LINKS = 10
COL_FIRST_MSG = 11
COL_LAST_MSG = 12
COL_JOIN_TS = 13
COL_JOIN_METHOD = 14

COLUMN_HEADERS = [
    "", "Name", "Phone", "Role", "Label", "Messages",
    "Text", "Images", "Videos", "Voice", "Links",
    "First Message", "Last Message", "Joined", "Join Method",
]

COLUMN_WIDTHS = [
    50,   # avatar
    220,  # name (wider for readability)
    140,  # phone
    100,  # role
    160,  # label (admin-assigned tag)
    80,   # messages
    60,   # text
    60,   # images
    60,   # videos
    60,   # voice
    60,   # links
    130,  # first msg
    130,  # last msg
    130,  # joined
    180,  # join method (wider for "Added by Name")
]


AVATAR_BLOB_ROLE = Qt.UserRole + 10


class _SortableItem(QTableWidgetItem):
    """QTableWidgetItem that sorts numerically by Qt.UserRole value."""

    def __lt__(self, other: QTableWidgetItem) -> bool:
        v1 = self.data(Qt.UserRole)
        v2 = other.data(Qt.UserRole) if other else 0
        if v1 is None:
            v1 = 0
        if v2 is None:
            v2 = 0
        try:
            return float(v1) < float(v2)
        except (TypeError, ValueError):
            return str(v1) < str(v2)


class AvatarDelegate(QStyledItemDelegate):
    """Draws a circular avatar with photo or initials in the first column."""

    CIRCLE_SIZE = 34

    def __init__(self, parent=None):
        super().__init__(parent)
        self._cache: dict[int, QPixmap | None] = {}

    def paint(self, painter: QPainter, option, index):
        if index.column() != COL_AVATAR:
            super().paint(painter, option, index)
            return

        painter.save()
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setRenderHint(QPainter.SmoothPixmapTransform)

        initials = index.data(Qt.DisplayRole) or "?"
        color_idx = index.data(Qt.UserRole) or 0
        avatar_blob = index.data(AVATAR_BLOB_ROLE)
        bg_color = QColor(AVATAR_COLORS[color_idx % len(AVATAR_COLORS)])

        # Center the circle in the cell
        cx = option.rect.x() + (option.rect.width() - self.CIRCLE_SIZE) // 2
        cy = option.rect.y() + (option.rect.height() - self.CIRCLE_SIZE) // 2

        # Try to draw actual photo
        photo_drawn = False
        if avatar_blob and len(avatar_blob) > 100:
            pxm = self._get_avatar(color_idx, avatar_blob)
            if pxm and not pxm.isNull():
                clip = QPainterPath()
                clip.addEllipse(float(cx), float(cy),
                                float(self.CIRCLE_SIZE), float(self.CIRCLE_SIZE))
                painter.setClipPath(clip)
                scaled = pxm.scaled(self.CIRCLE_SIZE, self.CIRCLE_SIZE,
                                    Qt.KeepAspectRatioByExpanding,
                                    Qt.SmoothTransformation)
                dx = (scaled.width() - self.CIRCLE_SIZE) // 2
                dy = (scaled.height() - self.CIRCLE_SIZE) // 2
                painter.drawPixmap(cx - dx, cy - dy, scaled)
                painter.setClipping(False)
                photo_drawn = True

        if not photo_drawn:
            path = QPainterPath()
            path.addEllipse(float(cx), float(cy),
                            float(self.CIRCLE_SIZE), float(self.CIRCLE_SIZE))
            painter.fillPath(path, bg_color)

            painter.setFont(QFont("Segoe UI", 10, QFont.Bold))
            painter.setPen(QColor(255, 255, 255))
            painter.drawText(
                QRect(cx, cy, self.CIRCLE_SIZE, self.CIRCLE_SIZE),
                Qt.AlignCenter, initials[:2].upper(),
            )

        painter.restore()

    def _get_avatar(self, contact_id: int, blob: bytes) -> QPixmap | None:
        if contact_id in self._cache:
            return self._cache[contact_id]
        pxm = QPixmap()
        pxm.loadFromData(blob)
        if pxm.isNull():
            self._cache[contact_id] = None
            return None
        self._cache[contact_id] = pxm
        if len(self._cache) > 500:
            oldest = next(iter(self._cache))
            del self._cache[oldest]
        return pxm

    def sizeHint(self, option, index):
        if index.column() == COL_AVATAR:
            from PySide6.QtCore import QSize
            return QSize(50, 44)
        return super().sizeHint(option, index)


class GroupInfoPage(QWidget):
    """Detailed group information page with member list and contribution stats."""

    back_requested = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._tm = ThemeManager.get()
        _lt = self._tm.is_light
        # Theme-aware color palette
        self._c_text = "#111b21" if _lt else "#e9edef"
        self._c_text2 = "#667781" if _lt else "#78909c"
        self._c_text3 = "#8696a0" if _lt else "#90a4ae"
        self._c_text_dim = "#a0aab0" if _lt else "#90a4ae"
        self._c_bg = "#fafafa" if _lt else "#0b141a"
        self._c_bg_header = "#f0f2f5" if _lt else "rgba(128,128,128,0.06)"
        self._c_bg_card = ("#ffffff" if _lt else "rgba(128,128,128,0.06)")
        self._c_border = "#e0e3e7" if _lt else "rgba(128,128,128,0.12)"
        self._c_border2 = "#f0f2f5" if _lt else "rgba(128,128,128,0.08)"
        self._c_accent = "#00897b" if _lt else "#00bcd4"
        self._c_accent_bg = "rgba(0,137,123,0.08)" if _lt else "rgba(0,188,212,0.08)"
        self._c_accent_border = "rgba(0,137,123,0.2)" if _lt else "rgba(0,188,212,0.2)"
        self._c_scroll = "rgba(0,0,0,0.12)" if _lt else "rgba(128,128,128,0.18)"
        self._c_hover = "#f5f6f6" if _lt else "rgba(128,128,128,0.08)"
        self._c_sel = "rgba(0,137,123,0.12)" if _lt else "rgba(0,188,212,0.12)"
        self._c_hdr_bg = "rgba(0,0,0,0.04)" if _lt else "rgba(128,128,128,0.08)"
        self._c_hdr_fg = "#667781" if _lt else "#607d8b"
        self._c_sep = "#e8eaed" if _lt else "rgba(128,128,128,0.12)"

        self._conv_id: int | None = None
        self._group_name: str = ""

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ---- Header bar ----
        self._build_header(main_layout)

        # ---- Scrollable content area ----
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet(f"""
            QScrollArea {{ background-color: {self._c_bg}; border: none; }}
            QScrollBar:vertical {{
                background: transparent; width: 8px; border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {self._c_scroll}; border-radius: 4px;
                min-height: 30px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)

        content = QWidget()
        content.setStyleSheet(f"background-color: {self._c_bg};")
        self._content_layout = QVBoxLayout(content)
        self._content_layout.setContentsMargins(24, 20, 24, 24)
        self._content_layout.setSpacing(16)

        # Stats bar
        self._build_stats_bar()

        # Group creation info
        self._build_creation_info()

        # Analytics section (top senders, activity)
        self._build_analytics_section()

        # Member table
        self._build_member_table()

        self._content_layout.addStretch()
        scroll.setWidget(content)
        main_layout.addWidget(scroll, 1)

    # ------------------------------------------------------------------ #
    # Header
    # ------------------------------------------------------------------ #

    def _build_header(self, parent_layout: QVBoxLayout) -> None:
        header = QFrame()
        header.setObjectName("groupInfoHeader")
        header.setFixedHeight(60)
        header.setStyleSheet(f"""
            QFrame#groupInfoHeader {{
                background-color: {self._c_bg_header};
                border-bottom: 1px solid {self._c_border};
            }}
        """)
        hl = QHBoxLayout(header)
        hl.setContentsMargins(12, 0, 12, 0)
        hl.setSpacing(12)

        # Back button
        self._back_btn = QPushButton("\u25C0")  # ◀ solid left triangle
        self._back_btn.setFixedSize(36, 36)
        self._back_btn.setCursor(Qt.PointingHandCursor)
        self._back_btn.setToolTip("Back")
        self._back_btn.setStyleSheet(self._tm.chat_back_btn_style())
        self._back_btn.clicked.connect(self.back_requested.emit)
        hl.addWidget(self._back_btn)

        # Avatar
        self._avatar = QLabel("?")
        self._avatar.setFixedSize(42, 42)
        self._avatar.setAlignment(Qt.AlignCenter)
        self._avatar.setStyleSheet("""
            QLabel { background: #00897b; border-radius: 21px;
                     color: white; font-size: 16px; font-weight: bold; }
        """)
        hl.addWidget(self._avatar)

        # Title column
        title_col = QVBoxLayout()
        title_col.setSpacing(0)
        self._title_label = QLabel("Group Info")
        f = QFont()
        f.setPointSize(14)
        f.setBold(True)
        self._title_label.setFont(f)
        self._title_label.setStyleSheet(f"color: {self._c_text};")
        title_col.addWidget(self._title_label)

        self._subtitle_label = QLabel("")
        self._subtitle_label.setStyleSheet(
            f"color: {self._c_text2}; font-size: 11px;"
        )
        title_col.addWidget(self._subtitle_label)
        hl.addLayout(title_col, 1)

        # Group Edit History button
        self._edit_history_btn = QPushButton("\u270E Group Edit History")
        self._edit_history_btn.setCursor(Qt.PointingHandCursor)
        self._edit_history_btn.setToolTip(
            "View complete timeline of group metadata changes:\n"
            "• Subject (name) changes with before/after\n"
            "• Description changes\n"
            "• Profile picture (DP) changes with thumbnails\n"
            "• Settings changes (admin-only, disappearing, etc.)"
        )
        self._edit_history_btn.setFixedHeight(32)
        _lt = self._tm.is_light
        self._edit_history_btn.setStyleSheet(f"""
            QPushButton {{
                background: {self._c_accent_bg};
                border: 1px solid {self._c_accent_border};
                border-radius: 6px;
                color: {self._c_accent};
                font-size: 11px; font-weight: bold;
                padding: 4px 14px;
            }}
            QPushButton:hover {{
                background: {"rgba(0,137,123,0.15)" if _lt else "rgba(0,188,212,0.15)"};
                border-color: {self._c_accent};
            }}
        """)
        self._edit_history_btn.clicked.connect(self._show_edit_history)
        hl.addWidget(self._edit_history_btn)

        # Generate Report button
        self._report_btn = QPushButton("\U0001F4CB Report")
        self._report_btn.setCursor(Qt.PointingHandCursor)
        self._report_btn.setToolTip(
            "Generate a comprehensive forensic HTML report for this group:\n"
            "• Edit history, member roster, mention network\n"
            "• Activity patterns, admin audit trail, media stats"
        )
        self._report_btn.setFixedHeight(32)
        self._report_btn.setStyleSheet(f"""
            QPushButton {{
                background: {self._c_accent_bg};
                border: 1px solid {self._c_accent_border};
                border-radius: 6px;
                color: {self._c_accent};
                font-size: 11px; font-weight: bold;
                padding: 4px 14px;
            }}
            QPushButton:hover {{
                background: {"rgba(0,137,123,0.15)" if _lt else "rgba(0,188,212,0.15)"};
                border-color: {self._c_accent};
            }}
        """)
        self._report_btn.clicked.connect(self._generate_group_report)
        hl.addWidget(self._report_btn)

        # Export chat HTML button — exports THIS group as a standalone V2 viewer bundle
        # Export participant roster XLSX
        self._participants_xlsx_btn = QPushButton("\u21E9 Participants XLSX")
        self._participants_xlsx_btn.setCursor(Qt.PointingHandCursor)
        self._participants_xlsx_btn.setToolTip(
            "Export group participants as XLSX:\n"
            "\u2022 Current members, DB past participants, and message-only participants\n"
            "\u2022 Contact DPs embedded in cells, plus device owner flag when known\n"
            "\u2022 First/last message details and full message-count breakdown"
        )
        self._participants_xlsx_btn.setFixedHeight(32)
        self._participants_xlsx_btn.setStyleSheet(self._report_btn.styleSheet())
        self._participants_xlsx_btn.clicked.connect(self._export_participants_xlsx)
        hl.addWidget(self._participants_xlsx_btn)

        self._export_btn = QPushButton("\u21E9 Export Chat HTML")
        self._export_btn.setCursor(Qt.PointingHandCursor)
        self._export_btn.setToolTip(
            "Export this group as a portable HTML bundle:\n"
            "\u2022 Works offline (double-click index.html)\n"
            "\u2022 Includes every message, reaction, poll, location, voice note\n"
            "\u2022 Media files copied with original WhatsApp filenames\n"
            "\u2022 Ctrl+K search, \u2139 forensic provenance on every bubble"
        )
        self._export_btn.setFixedHeight(32)
        self._export_btn.setStyleSheet(self._report_btn.styleSheet())
        self._export_btn.clicked.connect(self._export_chat_html)
        hl.addWidget(self._export_btn)

        parent_layout.addWidget(header)

    # ------------------------------------------------------------------ #
    # Stats bar
    # ------------------------------------------------------------------ #

    def _build_stats_bar(self) -> None:
        stats_frame = QFrame()
        stats_frame.setObjectName("groupStatsBar")
        stats_frame.setStyleSheet(f"""
            QFrame#groupStatsBar {{
                background: {self._c_accent_bg};
                border: 1px solid {self._c_accent_border};
                border-radius: 8px;
            }}
        """)
        stats_layout = QHBoxLayout(stats_frame)
        stats_layout.setContentsMargins(20, 14, 20, 14)
        stats_layout.setSpacing(0)

        self._stat_participants = self._make_stat_widget("Participants", "0")
        self._stat_admins = self._make_stat_widget("Admins", "0")
        self._stat_messages = self._make_stat_widget("Total Messages", "0")
        self._stat_media = self._make_stat_widget("Media Shared", "0")
        self._stat_links = self._make_stat_widget("Links Shared", "0")

        for i, widget in enumerate([
            self._stat_participants, self._stat_admins,
            self._stat_messages, self._stat_media, self._stat_links,
        ]):
            if i > 0:
                sep = QFrame()
                sep.setFixedWidth(1)
                sep.setStyleSheet(f"background: {self._c_sep};")
                sep.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
                stats_layout.addWidget(sep)
            stats_layout.addWidget(widget, 1)

        self._content_layout.addWidget(stats_frame)

    def _make_stat_widget(self, label: str, value: str) -> QWidget:
        w = QWidget()
        vl = QVBoxLayout(w)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(2)
        vl.setAlignment(Qt.AlignCenter)

        val_lbl = QLabel(value)
        val_lbl.setAlignment(Qt.AlignCenter)
        val_font = QFont()
        val_font.setPointSize(18)
        val_font.setBold(True)
        val_lbl.setFont(val_font)
        val_lbl.setStyleSheet(f"color: {self._c_accent};")
        val_lbl.setObjectName("statValue")
        vl.addWidget(val_lbl)

        lbl = QLabel(label)
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet(f"color: {self._c_text2}; font-size: 10px;")
        vl.addWidget(lbl)

        return w

    def _update_stat(self, widget: QWidget, value: str) -> None:
        val_label = widget.findChild(QLabel, "statValue")
        if val_label:
            val_label.setText(value)

    # ------------------------------------------------------------------ #
    # Creation info
    # ------------------------------------------------------------------ #

    def _build_creation_info(self) -> None:
        self._creation_frame = QFrame()
        self._creation_frame.setStyleSheet(f"""
            QFrame {{
                background: {self._c_bg_card};
                border: 1px solid {self._c_border};
                border-radius: 8px;
            }}
        """)
        cl = QVBoxLayout(self._creation_frame)
        cl.setContentsMargins(16, 12, 16, 12)
        cl.setSpacing(8)
        self._creation_label = QLabel("")
        self._creation_label.setStyleSheet(
            f"color: {self._c_text2}; font-size: 11px;"
        )
        self._creation_label.setWordWrap(True)
        cl.addWidget(self._creation_label)

        # Group description (hidden by default, shown if available)
        self._description_label = QLabel("")
        self._description_label.setStyleSheet(
            f"color: {self._c_text}; font-size: 12px; "
            f"padding: 8px 0 0 0; "
            f"border-top: 1px solid {self._c_border};"
        )
        self._description_label.setWordWrap(True)
        self._description_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self._description_label.setVisible(False)
        cl.addWidget(self._description_label)

        # Group settings (admin-only messaging, ephemeral, addressing mode, etc.)
        self._settings_label = QLabel("")
        self._settings_label.setStyleSheet(
            f"color: {self._c_text}; font-size: 11px; "
            f"padding: 8px 0 0 0; "
            f"border-top: 1px solid {self._c_border};"
        )
        self._settings_label.setWordWrap(True)
        self._settings_label.setVisible(False)
        cl.addWidget(self._settings_label)

        self._content_layout.addWidget(self._creation_frame)

    # ------------------------------------------------------------------ #
    # Analytics Section
    # ------------------------------------------------------------------ #

    def _build_analytics_section(self) -> None:
        self._analytics_frame = QFrame()
        self._analytics_frame.setObjectName("groupAnalytics")
        self._analytics_frame.setStyleSheet(f"""
            QFrame#groupAnalytics {{
                background: {self._c_bg_card};
                border: 1px solid {self._c_border};
                border-radius: 8px;
            }}
        """)
        al = QVBoxLayout(self._analytics_frame)
        al.setContentsMargins(20, 16, 20, 16)
        al.setSpacing(12)

        hdr = QLabel("Group Analytics")
        f = QFont()
        f.setPointSize(13)
        f.setBold(True)
        hdr.setFont(f)
        hdr.setStyleSheet(f"color: {self._c_text};")
        al.addWidget(hdr)

        # Top senders row
        self._top_senders_layout = QHBoxLayout()
        self._top_senders_layout.setSpacing(8)
        al.addLayout(self._top_senders_layout)

        # Activity summary row
        self._activity_layout = QHBoxLayout()
        self._activity_layout.setSpacing(16)
        al.addLayout(self._activity_layout)

        self._content_layout.addWidget(self._analytics_frame)

    # ------------------------------------------------------------------ #
    # Member table
    # ------------------------------------------------------------------ #

    def _build_member_table(self) -> None:
        # Owner-status banner — populated by _update_owner_banner() once
        # the conversation is loaded. Empty by default so it takes no
        # vertical space if the detector has nothing to say.
        self._owner_banner = QLabel("")
        self._owner_banner.setWordWrap(True)
        self._owner_banner.setVisible(False)
        self._owner_banner.setStyleSheet(
            "padding: 10px 14px; border-radius: 6px; "
            "font-size: 12px; font-weight: 600; margin-bottom: 8px;"
        )
        self._content_layout.addWidget(self._owner_banner)

        # Section header + filter row
        member_header_row = QHBoxLayout()
        member_header_row.setSpacing(12)
        section_header = QLabel("Group Members")
        f = QFont()
        f.setPointSize(13)
        f.setBold(True)
        section_header.setFont(f)
        section_header.setStyleSheet(f"color: {self._c_text};")
        member_header_row.addWidget(section_header)

        self._member_count_label = QLabel("")
        self._member_count_label.setStyleSheet(f"color: {self._c_text2}; font-size: 11px;")
        member_header_row.addWidget(self._member_count_label)
        member_header_row.addStretch()

        # Filter input
        self._filter_input = QLineEdit()
        self._filter_input.setPlaceholderText("\u2315 Filter members by name, phone, or role...")
        self._filter_input.setFixedWidth(300)
        self._filter_input.setFixedHeight(30)
        self._filter_input.setClearButtonEnabled(True)
        _lt = self._tm.is_light
        self._filter_input.setStyleSheet(f"""
            QLineEdit {{
                background: {"#ffffff" if _lt else "rgba(128,128,128,0.08)"};
                border: 1px solid {self._c_border};
                border-radius: 6px;
                color: {self._c_text};
                padding: 0 10px;
                font-size: 11px;
            }}
            QLineEdit:focus {{
                border: 1px solid {self._c_accent};
            }}
        """)
        self._filter_input.textChanged.connect(self._filter_members)
        member_header_row.addWidget(self._filter_input)

        # Resolution filter: jump straight to "members whose
        # actual phone we resolved" vs "LID-only members".
        # Critical in large announcement-style communities where
        # the device only ever saw a LID identity for most
        # participants — this lets the investigator instantly
        # see which numbers are forensically actionable versus
        # WhatsApp privacy-shielded.
        from PySide6.QtWidgets import QComboBox
        self._resolution_filter = QComboBox()
        self._resolution_filter.addItem("All members", "all")
        self._resolution_filter.addItem("✓ Phone resolved only", "resolved")
        self._resolution_filter.addItem("⚠ LID-only (unresolved)", "lid_only")
        self._resolution_filter.addItem("📇 Saved (in address book)", "saved")
        self._resolution_filter.addItem("~ Unsaved (WhatsApp-seen only)", "unsaved")
        self._resolution_filter.setFixedWidth(260)
        self._resolution_filter.setFixedHeight(30)
        self._resolution_filter.setStyleSheet(f"""
            QComboBox {{
                background: {"#ffffff" if _lt else "rgba(128,128,128,0.08)"};
                border: 1px solid {self._c_border};
                border-radius: 6px;
                color: {self._c_text};
                padding: 0 10px;
                font-size: 11px;
            }}
            QComboBox:focus {{
                border: 1px solid {self._c_accent};
            }}
            QComboBox::drop-down {{ border: 0; }}
        """)
        self._resolution_filter.currentIndexChanged.connect(self._filter_members)
        member_header_row.addWidget(self._resolution_filter)

        self._content_layout.addLayout(member_header_row)

        self._table = QTableWidget()
        self._table.setColumnCount(len(COLUMN_HEADERS))
        self._table.setHorizontalHeaderLabels(COLUMN_HEADERS)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.setShowGrid(False)
        self._table.setSortingEnabled(True)
        self._table.verticalHeader().setDefaultSectionSize(52)
        self._table.setMinimumHeight(400)
        self._table.setWordWrap(True)

        # Delegate for avatar column
        self._avatar_delegate = AvatarDelegate(self._table)
        self._table.setItemDelegateForColumn(COL_AVATAR, self._avatar_delegate)

        # Column widths
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(COL_NAME, QHeaderView.Stretch)
        hdr.setMinimumSectionSize(150)  # prevent name column from getting too narrow
        for col, w in enumerate(COLUMN_WIDTHS):
            if col != COL_NAME:
                hdr.setSectionResizeMode(col, QHeaderView.Fixed)
                hdr.resizeSection(col, w)

        _alt_bg = "#f5f6f6" if _lt else "rgba(128,128,128,0.04)"
        self._table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {"#ffffff" if _lt else "transparent"};
                border: 1px solid {self._c_border};
                border-radius: 8px;
                color: {self._c_text};
                gridline-color: transparent;
                font-size: 11px;
                alternate-background-color: {_alt_bg};
            }}
            QTableWidget::item {{
                padding: 6px 8px;
                border-bottom: 1px solid {self._c_border2};
            }}
            QTableWidget::item:selected {{
                background: {self._c_sel};
            }}
            QTableWidget::item:hover {{
                background: {self._c_hover};
            }}
            QHeaderView::section {{
                background-color: {self._c_hdr_bg};
                color: {self._c_hdr_fg};
                border: none;
                border-bottom: 1px solid {self._c_sep};
                padding: 8px 8px;
                font-size: 10px;
                font-weight: bold;
                text-transform: uppercase;
            }}
            QHeaderView::section:hover {{
                background-color: {self._c_hover};
            }}
            QScrollBar:vertical {{
                background: transparent; width: 8px; border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {self._c_scroll}; border-radius: 4px;
                min-height: 30px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)

        # Context menu + double-click on member rows
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._member_context_menu)
        self._table.doubleClicked.connect(self._on_member_double_click)

        self._content_layout.addWidget(self._table, 1)

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def load_group(self, conv_id: int, display_name: str) -> None:
        """Load and display group information for the given conversation."""
        self._conv_id = conv_id
        self._group_name = display_name or f"Group #{conv_id}"

        db = Database.get()

        # Update header — try to load group avatar
        self._title_label.setText(self._group_name)

        avatar_blob = db.scalar(
            "SELECT avatar_blob FROM conversation WHERE id = ?", (conv_id,)
        )
        if avatar_blob and len(avatar_blob) > 100:
            pxm = QPixmap()
            pxm.loadFromData(avatar_blob)
            if not pxm.isNull():
                scaled = pxm.scaled(42, 42, Qt.KeepAspectRatioByExpanding,
                                    Qt.SmoothTransformation)
                self._avatar.setPixmap(scaled)
                self._avatar.setText("")
                self._avatar.setStyleSheet(
                    "QLabel { border-radius: 21px; }"
                )
            else:
                avatar_blob = None  # fallback to initials

        if not avatar_blob or len(avatar_blob) <= 100:
            initials = "".join(
                w[0] for w in self._group_name.split()[:2]
                if w and w[0].isalpha()
            )
            if not initials:
                initials = "#"
            avatar_bg = AVATAR_COLORS[conv_id % len(AVATAR_COLORS)]
            self._avatar.setPixmap(QPixmap())
            self._avatar.setText(initials[:2].upper())
            self._avatar.setStyleSheet(
                f"QLabel {{ background: {avatar_bg}; border-radius: 21px; "
                f"color: white; font-size: 16px; font-weight: bold; }}"
            )

        # Load conversation metadata
        conv = db.fetchone(
            "SELECT created_timestamp, participant_count, message_count, "
            "group_type, addressing_mode, ephemeral_duration, is_locked, "
            "description, is_archived, is_muted, jid_raw_string "
            "FROM conversation WHERE id = ?", (conv_id,)
        )

        created_ts = conv["created_timestamp"] if conv else None
        participant_count = (conv["participant_count"] if conv else 0) or 0
        total_messages = (conv["message_count"] if conv else 0) or 0
        group_type = conv["group_type"] if conv else None
        addressing_mode = conv["addressing_mode"] if conv else None
        ephemeral_dur = conv["ephemeral_duration"] if conv else None
        is_locked = conv["is_locked"] if conv else False
        description = conv["description"] if conv else None
        group_jid = conv["jid_raw_string"] if conv else None

        # Find creator
        creator_name = ""
        creator_row = db.fetchone(
            """
            SELECT COALESCE(c.resolved_name, c.wa_name, c.phone_number, 'Unknown') AS name,
                   c.phone_number, c.phone_jid
            FROM group_member gm
            JOIN contact c ON c.id = gm.contact_id
            WHERE gm.conversation_id = ? AND gm.role = 'superadmin'
            LIMIT 1
            """,
            (conv_id,),
        )
        if creator_row:
            creator_name = creator_row["name"]
            creator_phone = creator_row["phone_number"] or ""
            if not creator_phone and creator_row["phone_jid"]:
                jid = creator_row["phone_jid"]
                if "@" in jid:
                    creator_phone = jid.split("@")[0]
            if creator_phone and creator_phone != creator_name:
                creator_name = f"{creator_name} (+{creator_phone})" if not creator_phone.startswith("+") else f"{creator_name} ({creator_phone})"
        else:
            # Fallback: extract creator phone from group JID for older groups
            # Format: 15551234567-1419532254@g.us — number before '-' is creator
            group_jid = db.scalar(
                "SELECT jid_raw_string FROM conversation WHERE id = ?", (conv_id,)
            ) or ""
            if "@g.us" in group_jid:
                prefix = group_jid.split("@")[0]
                if "-" in prefix:
                    creator_phone = prefix.split("-")[0]
                    cr = db.fetchone("""
                        SELECT COALESCE(c.resolved_name, c.wa_name, c.phone_number) AS name,
                               c.phone_number, c.phone_jid
                        FROM contact c
                        WHERE c.phone_number = ? OR c.phone_jid = ?
                        LIMIT 1
                    """, (creator_phone, f"{creator_phone}@s.whatsapp.net"))
                    if cr:
                        creator_name = cr["name"] or f"+{creator_phone}"
                        if cr["phone_number"] and cr["phone_number"] != cr["name"]:
                            creator_name = f"{creator_name} (+{cr['phone_number']})"
                    else:
                        creator_name = f"+{creator_phone} (from JID)"

        # Creation info
        creation_parts = []
        if created_ts:
            try:
                creation_parts.append(
                    f"Created on {_fmt_ts(created_ts, '%B %d, %Y at %H:%M')}"
                )
            except (ValueError, OSError):
                pass
        if creator_name:
            creation_parts.append(f"Created by {creator_name}")
        self._creation_label.setText("  \u2022  ".join(creation_parts) if creation_parts else "")

        # Group description
        desc_row = db.fetchone(
            "SELECT description FROM conversation WHERE id = ?", (conv_id,)
        )
        description = desc_row["description"] if desc_row and desc_row["description"] else None
        if description:
            self._description_label.setText(description)
            self._description_label.setVisible(True)
        else:
            self._description_label.setVisible(False)

        # Group settings display
        settings_parts = []
        _group_type_names = {
            0: "Regular Group", 1: "Community (Announce)", 2: "Community Sub-group",
            3: "Community Default", 4: "Newsletter", 5: "Announcement Group", 6: "Community Sub-group",
        }
        if group_type is not None:
            settings_parts.append(f"Type: {_group_type_names.get(group_type, f'Type {group_type}')}")
        if addressing_mode:
            mode_label = "LID (Privacy-restricted)" if addressing_mode == "lid" else "Phone Number"
            settings_parts.append(f"Addressing: {mode_label}")
        if group_type == 1 or group_type == 5:
            settings_parts.append("\u26A0 Only admins can send messages")
        if ephemeral_dur and ephemeral_dur > 0:
            if ephemeral_dur >= 86400:
                dur_str = f"{ephemeral_dur // 86400} days"
            elif ephemeral_dur >= 3600:
                dur_str = f"{ephemeral_dur // 3600} hours"
            else:
                dur_str = f"{ephemeral_dur // 60} min"
            settings_parts.append(f"Disappearing messages: {dur_str}")
        if is_locked:
            settings_parts.append("\U0001F512 Chat locked")
        if settings_parts:
            self._settings_label.setText("  |  ".join(settings_parts))
            self._settings_label.setVisible(True)
        else:
            self._settings_label.setVisible(False)

        subtitle_parts = [f"{participant_count} participants", f"{total_messages:,} messages"]
        if creator_name:
            subtitle_parts.append(f"by {creator_name}")
        # Always show the group JID — it's the primary forensic identifier
        # for the conversation (matches msgstore.chat.jid_row_id and is what
        # WhatsApp uses internally to address this group).
        if group_jid:
            subtitle_parts.append(f"JID: {group_jid}")
        self._subtitle_label.setText("  |  ".join(subtitle_parts))

        # Load members.  Also pull lid_masked_phone (so LID-only members
        # show "+91∙∙∙∙∙∙∙∙58" instead of "+105261209485559") AND
        # is_saved so the GUI can distinguish "in device owner's address
        # book" (📇) from "just-seen in this group" (~tilde).
        members = db.fetchall(
            """
            SELECT gm.id, gm.contact_id, gm.role, gm.label,
                   gm.join_timestamp, gm.join_method,
                   c.resolved_name, c.phone_number, c.phone_jid, c.lid_jid,
                   c.wa_name, c.display_name, c.avatar_blob,
                   c.lid_masked_phone, c.is_saved
            FROM group_member gm
            JOIN contact c ON c.id = gm.contact_id
            WHERE gm.conversation_id = ?
            ORDER BY gm.role DESC, c.resolved_name
            """,
            (conv_id,),
        )

        # Enrich join info from system events (who added whom)
        join_events = db.fetchall(
            """
            SELECT se.target_id,
                   COALESCE(ac.resolved_name, ac.wa_name, ac.phone_number) AS actor_name,
                   COALESCE(tc.resolved_name, tc.wa_name, tc.phone_number) AS target_name
            FROM system_event se
            JOIN message m ON m.id = se.message_id
            LEFT JOIN contact ac ON ac.id = se.actor_id
            LEFT JOIN contact tc ON tc.id = se.target_id
            WHERE m.conversation_id = ?
              AND se.event_label = 'participant_joined'
              AND se.actor_id IS NOT NULL AND se.target_id IS NOT NULL
              AND se.actor_id != se.target_id
            """,
            (conv_id,),
        )
        added_by_map: dict[int, str] = {}
        for ev in join_events:
            if ev["target_id"] and ev["actor_name"]:
                added_by_map[ev["target_id"]] = ev["actor_name"]

        # Load contribution stats per sender
        contrib_rows = db.fetchall(
            """
            SELECT m.sender_id,
                   COUNT(*) as total,
                   SUM(CASE WHEN m.message_type = 0 THEN 1 ELSE 0 END) as text_count,
                   SUM(CASE WHEN m.message_type = 1 THEN 1 ELSE 0 END) as image_count,
                   SUM(CASE WHEN m.message_type = 3 THEN 1 ELSE 0 END) as video_count,
                   SUM(CASE WHEN m.message_type = 2 THEN 1 ELSE 0 END) as audio_count,
                   SUM(CASE WHEN m.message_type = 9 THEN 1 ELSE 0 END) as doc_count,
                   SUM(CASE WHEN m.is_forwarded = 1 THEN 1 ELSE 0 END) as fwd_count,
                   MIN(m.timestamp) as first_msg,
                   MAX(m.timestamp) as last_msg
            FROM message m
            WHERE m.conversation_id = ?
              AND m.sender_id IS NOT NULL
              AND m.message_type != 7
            GROUP BY m.sender_id
            """,
            (conv_id,),
        )
        contrib_map: dict[int, dict] = {}
        for row in contrib_rows:
            contrib_map[row["sender_id"]] = {
                "total": row["total"] or 0,
                "text": row["text_count"] or 0,
                "image": row["image_count"] or 0,
                "video": row["video_count"] or 0,
                "audio": row["audio_count"] or 0,
                "doc": row["doc_count"] or 0,
                "fwd": row["fwd_count"] or 0,
                "first_msg": row["first_msg"],
                "last_msg": row["last_msg"],
            }

        # Load link counts per sender
        link_rows = db.fetchall(
            """
            SELECT m.sender_id, COUNT(*) as link_count
            FROM message_link_detail mld
            JOIN message m ON m.id = mld.message_id
            WHERE m.conversation_id = ?
              AND m.sender_id IS NOT NULL
            GROUP BY m.sender_id
            """,
            (conv_id,),
        )
        link_map: dict[int, int] = {}
        for row in link_rows:
            link_map[row["sender_id"]] = row["link_count"] or 0

        # Compute totals for stats bar
        admin_count = sum(
            1 for m in members
            if (m["role"] or "").lower() in ("admin", "superadmin")
        )
        total_msg_sum = sum(c["total"] for c in contrib_map.values())
        total_media = sum(
            c["image"] + c["video"] + c["audio"]
            for c in contrib_map.values()
        )
        total_links = sum(link_map.values())

        # Update stats bar
        self._update_stat(self._stat_participants, f"{len(members):,}")
        self._update_stat(self._stat_admins, str(admin_count))
        self._update_stat(self._stat_messages, f"{total_msg_sum:,}")
        self._update_stat(self._stat_media, f"{total_media:,}")
        self._update_stat(self._stat_links, f"{total_links:,}")

        # Populate analytics section
        self._populate_analytics(members, contrib_map, link_map, db, conv_id)

        # Populate table
        self._table.setSortingEnabled(False)
        self._table.setRowCount(len(members))

        for row_idx, member in enumerate(members):
            contact_id = member["contact_id"]
            name = (
                member["resolved_name"]
                or member["display_name"]
                or member["wa_name"]
                or member["phone_number"]
                or member["phone_jid"]
                or "Unknown"
            )
            phone = member["phone_number"] or member["phone_jid"] or ""
            role = (member["role"] or "member").lower()
            join_ts = member["join_timestamp"]
            join_method = member["join_method"] or ""

            stats = contrib_map.get(contact_id, {})
            links = link_map.get(contact_id, 0)

            # -- Avatar column (with photo support) --
            initials_text = "".join(
                w[0] for w in name.split()[:2]
                if w and w[0].isalpha()
            )
            if not initials_text:
                initials_text = "#"

            avatar_item = QTableWidgetItem(initials_text[:2].upper())
            avatar_item.setData(Qt.UserRole, contact_id)
            avatar_item.setData(AVATAR_BLOB_ROLE, member["avatar_blob"])
            avatar_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            avatar_item.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(row_idx, COL_AVATAR, avatar_item)

            # -- Name column.  The "~" prefix follows WhatsApp's own
            # UI rule: it means "this is a NAME the contact chose for
            # their own WhatsApp profile, NOT a name from your phone
            # book" - so you should treat it with appropriate
            # skepticism (they could change it tomorrow).
            #
            # We only add ~ when there's an actual contact-controlled
            # name to mark.  A phone number or a fallback placeholder
            # is just shown as-is - no tilde, because there's no name
            # for the tilde to qualify.
            #
            #   • saved              ->  "Jane Doe"          (trusted)
            #   • unsaved + wa_name  ->  "~Jane"              (their WA name)
            #   • unsaved phone-only ->  "+1 555 555 5555"    (no name -> no tilde)
            #   • unsaved no info    ->  "+105261209485559"   (or LID; no tilde)
            try:
                is_saved = bool(member["is_saved"])
            except (IndexError, KeyError):
                is_saved = False
            display_name_member = (member["display_name"] or "").strip()
            wa_name = (member["wa_name"] or "").strip()
            phone_for_name = (member["phone_number"] or "").strip()
            if is_saved and display_name_member:
                name_for_display = display_name_member
            elif wa_name:
                # Unsaved but the contact's WhatsApp profile name is set -
                # the tilde marks this is their self-chosen name, not
                # one you saved.
                name_for_display = f"~{wa_name}"
            elif phone_for_name:
                name_for_display = f"+{phone_for_name}" if not phone_for_name.startswith("+") else phone_for_name
            else:
                # No name and no phone - just show the existing fallback
                # (LID / phone JID / "Unknown") with NO tilde.  Adding
                # "~" to a placeholder would be misleading because
                # someone could legitimately set their WA name to
                # "Unknown" and we'd have no way to tell apart.
                name_for_display = name.lstrip("~ ").strip() or "Unknown"
            name_item = QTableWidgetItem(name_for_display)
            if role in ("superadmin", "admin"):
                name_item.setForeground(QBrush(QColor(self._c_text)))
            else:
                name_item.setForeground(QBrush(QColor("#3b4a54" if self._tm.is_light else "#dce3e8")))
            name_font = QFont("Segoe UI", 11)
            name_font.setBold(role in ("admin", "superadmin") or is_saved)
            name_item.setFont(name_font)
            name_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            # Sortable saved-flag for filter logic
            name_item.setData(Qt.UserRole + 2, 1 if is_saved else 0)
            # Full details in tooltip
            tooltip_parts = [
                name_for_display,
                "Saved in address book" if is_saved else "Not in address book (WhatsApp-seen only)",
            ]
            if phone:
                tooltip_parts.append(f"Phone: {phone}")
            if wa_name and wa_name != name_for_display:
                tooltip_parts.append(f"WhatsApp: {wa_name}")
            lid_jid = member["lid_jid"] or ""
            if lid_jid:
                tooltip_parts.append(f"LID: {lid_jid}")
            name_item.setToolTip("\n".join(tooltip_parts))
            self._table.setItem(row_idx, COL_NAME, name_item)

            # -- Phone column.  Three-tier fallback so LIDs are
            # never displayed as a fake "+1052612..." phone:
            #   1. real resolved phone -> "+1 555 555 5555"
            #   2. masked phone hint   -> "+91∙∙∙∙∙∙∙∙58 (LID-only)"
            #   3. truly nothing       -> "—" with full LID in tooltip
            real_phone = (member["phone_number"] or "").strip()
            try:
                masked_phone = (member["lid_masked_phone"] or "").strip()
            except (IndexError, KeyError):
                masked_phone = ""
            lid_jid = (member["lid_jid"] or "").strip()
            phone_jid = (member["phone_jid"] or "").strip()

            phone_resolved = False
            if real_phone:
                # Tier 1: a known phone number
                phone_display = real_phone if real_phone.startswith("+") else f"+{real_phone}"
                phone_resolved = True
                tooltip = real_phone
            elif phone_jid and "@" in phone_jid:
                # Tier 1b: phone-JID is set even if phone_number is null
                # (some legacy rows).  Strip the suffix, "+digits".
                raw = phone_jid.split("@")[0]
                if raw.isdigit():
                    phone_display = f"+{raw}"
                    phone_resolved = True
                    tooltip = phone_jid
                else:
                    phone_display = "—"
                    tooltip = phone_jid
            elif masked_phone:
                # Tier 2: WhatsApp's privacy-masked hint - the contact's
                # phone never reached this device (Phone Number Privacy
                # was on, or LID-only group join), but WhatsApp gave us
                # last-2-digit visibility.
                phone_display = f"{masked_phone}  (LID-only)"
                tooltip = (
                    f"LID-only contact - phone never written to msgstore.\n"
                    f"Mask: {masked_phone}\n"
                    f"LID:  {lid_jid}"
                )
            else:
                # Tier 3: no phone, no mask, just LID.  Show em-dash so
                # the column never lies; LID lives in the tooltip.
                phone_display = "—"
                tooltip = f"LID: {lid_jid}" if lid_jid else "no identity"

            phone_item = QTableWidgetItem(phone_display)
            phone_item.setForeground(QBrush(QColor(
                self._c_text2 if phone_resolved else (
                    "#9b6e00" if self._tm.is_light else "#d4a800"
                )
            )))
            phone_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            phone_item.setToolTip(tooltip)
            # Stash a sortable resolution flag (1 = real phone, 0 = LID-only)
            phone_item.setData(Qt.UserRole + 1, 1 if phone_resolved else 0)
            self._table.setItem(row_idx, COL_PHONE, phone_item)

            # -- Role (color-coded) --
            role_cfg = ROLE_CONFIG.get(role, ROLE_CONFIG["member"])
            role_item = QTableWidgetItem(role_cfg["label"])
            role_item.setTextAlignment(Qt.AlignCenter)
            role_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            if role_cfg["bg"] != "transparent":
                role_item.setBackground(QBrush(QColor(role_cfg["bg"])))
            _role_fg = role_cfg["fg"]
            if role == "member":
                _role_fg = self._c_text2  # theme-aware for members
            role_item.setForeground(QBrush(QColor(_role_fg)))
            role_font = QFont()
            role_font.setPointSize(9)
            role_font.setBold(True)
            role_item.setFont(role_font)
            # Store raw role for sorting (superadmin=0, admin=1, member=2)
            sort_val = {"superadmin": 0, "admin": 1}.get(role, 2)
            role_item.setData(Qt.UserRole, sort_val)
            self._table.setItem(row_idx, COL_ROLE, role_item)

            # -- Label / admin-assigned tag (apartment number, etc.) --
            label_text = member["label"] or ""
            label_item = QTableWidgetItem(label_text)
            label_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            label_item.setForeground(QBrush(QColor(self._c_text2)))
            if label_text:
                label_item.setToolTip(label_text)
            self._table.setItem(row_idx, COL_LABEL, label_item)

            # -- Numeric stats columns --
            total = stats.get("total", 0)
            for col, val in [
                (COL_TOTAL, total),
                (COL_TEXT,  stats.get("text", 0)),
                (COL_IMAGE, stats.get("image", 0)),
                (COL_VIDEO, stats.get("video", 0)),
                (COL_AUDIO, stats.get("audio", 0)),
                (COL_LINKS, links),
            ]:
                item = _SortableItem()
                item.setData(Qt.DisplayRole, f"{val:,}" if val else "0")
                item.setData(Qt.UserRole, val)
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)

                # Color-code message count by volume
                if col == COL_TOTAL:
                    if val > 5000:
                        item.setForeground(QBrush(QColor("#c62828" if self._tm.is_light else "#ef5350")))
                        bold_f = QFont()
                        bold_f.setBold(True)
                        item.setFont(bold_f)
                    elif val > 1000:
                        item.setForeground(QBrush(QColor("#e65100" if self._tm.is_light else "#ffa726")))
                    elif val > 0:
                        item.setForeground(QBrush(QColor(self._c_text)))
                    else:
                        item.setForeground(QBrush(QColor(self._c_text_dim)))
                else:
                    if val > 0:
                        item.setForeground(QBrush(QColor(self._c_text2)))
                    else:
                        item.setForeground(QBrush(QColor(self._c_text_dim)))

                self._table.setItem(row_idx, col, item)

            # -- Timestamp columns --
            for col, ts_val in [
                (COL_FIRST_MSG, stats.get("first_msg")),
                (COL_LAST_MSG, stats.get("last_msg")),
                (COL_JOIN_TS, join_ts),
            ]:
                display = ""
                sort_key = 0
                if ts_val:
                    try:
                        display = _fmt_ts(ts_val, "minute")
                        sort_key = ts_val
                    except (ValueError, OSError):
                        display = str(ts_val)
                        sort_key = ts_val if isinstance(ts_val, (int, float)) else 0

                ts_item = _SortableItem(display)
                ts_item.setData(Qt.UserRole, sort_key)
                ts_item.setTextAlignment(Qt.AlignCenter)
                ts_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                if display:
                    ts_item.setForeground(QBrush(QColor(self._c_text2)))
                else:
                    ts_item.setForeground(QBrush(QColor(self._c_text_dim)))
                self._table.setItem(row_idx, col, ts_item)

            # -- Join method (enriched with system event data) --
            added_by = added_by_map.get(contact_id, "")
            if added_by:
                method_display = f"Added by {added_by}"
            elif join_method == 3:
                method_display = "\u26D3 Invite Link"
            elif join_method == 1:
                method_display = "Added by Admin"
            elif join_method == 5:
                method_display = "Via Community"
            elif role == "superadmin":
                method_display = "\u2606 Group Creator"
            elif join_method == 0:
                method_display = "Member"
            else:
                method_display = self._format_join_method(join_method)
            method_item = QTableWidgetItem(method_display)
            method_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            method_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            method_item.setToolTip(method_display)
            if "Invite Link" in method_display:
                method_item.setForeground(QBrush(QColor("#29b6f6")))
            elif "Added by" in method_display:
                method_item.setForeground(QBrush(QColor("#66bb6a")))
            elif "Creator" in method_display:
                method_item.setForeground(QBrush(QColor("#ffb300")))
            else:
                method_item.setForeground(QBrush(QColor(self._c_text2)))
            self._table.setItem(row_idx, COL_JOIN_METHOD, method_item)

        self._table.setSortingEnabled(True)
        self._filter_input.clear()
        # Reset resolution dropdown to "All" + drive count refresh so the
        # "resolved / LID-only" split shows immediately.
        try:
            self._resolution_filter.setCurrentIndex(0)
        except Exception:
            pass
        self._filter_members()

        # ---- Owner membership status banner ----
        self._update_owner_banner(db, conv_id)

        # ---- Past / Former Participants ----
        self._load_past_participants(db, conv_id)

    def _update_owner_banner(self, db, conv_id: int) -> None:
        """Render the device-owner membership banner for this group."""
        if not hasattr(self, "_owner_banner"):
            return
        owner_cid = self._resolve_device_owner_cid(db)
        info = self._detect_owner_membership(db, conv_id, owner_cid)
        status = info["status"]
        source = info.get("source", "none")

        def _fmt_ts(ts):
            if not ts:
                return ""
            try:
                from datetime import datetime as _dt
                return _dt.fromtimestamp(ts / 1000).strftime("%Y-%m-%d %H:%M")
            except (ValueError, OSError):
                return ""

        join_str  = _fmt_ts(info["join_ts"])
        leave_str = _fmt_ts(info["leave_ts"])

        # Pull admin settings + admins-only-send timeline
        admin_row = db.fetchone(
            "SELECT announcement_group, restrict_mode, "
            "       require_membership_approval, member_add_mode "
            "FROM conversation WHERE id = ?",
            (conv_id,),
        )
        admin_notes = []
        if admin_row:
            if admin_row["announcement_group"]:
                admin_notes.append("\U0001F4E2 only admins can send messages")
            if admin_row["restrict_mode"]:
                admin_notes.append("\U0001F512 only admins can edit group info")
            if admin_row["require_membership_approval"]:
                admin_notes.append("\U0001F6AA admin approval required to join")
            if admin_row["member_add_mode"]:
                admin_notes.append("\U0001F464 only admins can add members")

        # If admins-only-send is ON, derive and surface a clear
        # "owner cannot send" line plus the timeline when we have it.
        owner_send_note: str | None = None
        timeline_note: str | None = None
        if admin_row and admin_row["announcement_group"]:
            # Owner can send only if they are admin/creator (status in {3,4})
            # — otherwise blocked. If status is unknown fall back to the
            # system-event detector's result.
            owner_can_send = status in ("admin", "creator")
            if not owner_can_send and status in ("member", "current"):
                owner_send_note = (
                    "\u26D4 Device owner does NOT have permission to send "
                    "messages in this group (only admins can send)."
                )
            elif owner_can_send:
                owner_send_note = (
                    "\u2705 Device owner is an admin and CAN send messages "
                    "in this group."
                )
            # Look up the LATEST admin_only_send_on event timestamp
            tl = db.fetchone(
                "SELECT gmc.timestamp, "
                "       COALESCE(c.resolved_name, c.wa_name, c.phone_number) AS actor "
                "FROM group_metadata_change gmc "
                "LEFT JOIN contact c ON c.id = gmc.changed_by_id "
                "WHERE gmc.conversation_id = ? AND gmc.change_type = 'admin_only_send_on' "
                "ORDER BY gmc.timestamp DESC LIMIT 1",
                (conv_id,),
            )
            if tl and tl["timestamp"]:
                timeline_note = (
                    f"\U0001F553 Admins-only-send enabled on "
                    f"{_fmt_ts(tl['timestamp'])}"
                    + (f" (by {tl['actor']})" if tl['actor'] else "")
                )

        # Status → visual style + title + verb mapping
        if status == "creator":
            bg, border, icon = "#1b5e20", "#2e7d32", "\u2B50"
            title = "Device owner is the CREATOR of this group"
        elif status == "admin":
            bg, border, icon = "#1565c0", "#1976d2", "\U0001F6E1" # 🛡
            title = "Device owner is an ADMIN of this group"
        elif status == "member":
            bg, border, icon = "#1b5e20", "#2e7d32", "\u2705"
            title = "Device owner is a MEMBER of this group"
        elif status == "current":
            bg, border, icon = "#1b5e20", "#2e7d32", "\u2705"
            title = "Device owner is currently a MEMBER of this group"
        elif status == "former":
            bg, border, icon = "#b71c1c", "#c62828", "\u274C"
            title = "Device owner is NO LONGER a member of this group"
        elif status == "former_left":
            bg, border, icon = "#b71c1c", "#c62828", "\u274C"
            title = "Device owner is NO LONGER in this group \u2014 left voluntarily"
        elif status == "former_removed":
            bg, border, icon = "#b71c1c", "#c62828", "\u274C"
            title = "Device owner is NO LONGER in this group \u2014 removed by admin"
        else:
            bg, border, icon = "#455a64", "#546e7a", "\u2753"
            title = "Device owner membership: unknown"

        sub = info["detail"]
        if join_str:  sub += f"  \u2022  joined {join_str}"
        if leave_str: sub += f"  \u2022  left {leave_str}"

        # Source attribution — show the user where this status came from
        SRC_LABELS = {
            "participation_status": "chat.participation_status (msgstore.db)",
            "system_event":         "system events (msgstore.db message_system)",
            "inferred":             "inferred from sent messages (no explicit record)",
            "none":                 "no data",
        }
        src_text = SRC_LABELS.get(source, source)

        html_parts = [
            f"<div style='font-size:14px;'>{icon} <b>{title}</b></div>",
            f"<div style='font-size:11px; opacity:0.85; margin-top:4px;'>{sub}</div>",
        ]
        if admin_notes:
            html_parts.append(
                f"<div style='font-size:11px; opacity:0.80; margin-top:4px;'>"
                f"<b>Group rules:</b> {' \u2022 '.join(admin_notes)}</div>"
            )
        if owner_send_note:
            html_parts.append(
                f"<div style='font-size:12px; font-weight:600; margin-top:6px;"
                f" padding:6px 8px; background: rgba(0,0,0,0.20);"
                f" border-radius:4px;'>{owner_send_note}</div>"
            )
        if timeline_note:
            html_parts.append(
                f"<div style='font-size:11px; opacity:0.90; margin-top:4px;'>"
                f"{timeline_note}</div>"
            )
        html_parts.append(
            f"<div style='font-size:10px; opacity:0.65; margin-top:4px;'>"
            f"Source: {src_text}</div>"
        )

        self._owner_banner.setText("".join(html_parts))
        self._owner_banner.setStyleSheet(
            f"QLabel {{ background: {bg}; color: #ffffff; "
            f"border-left: 4px solid {border}; "
            f"padding: 10px 14px; border-radius: 6px; "
            f"margin-bottom: 8px; }}"
        )
        self._owner_banner.setVisible(True)

    def _load_past_participants(self, db, conv_id: int) -> None:
        """Load and display former group members.

        Rules:
          * Authoritative "current member" = row in
            ``group_participant_user`` (mirrored to our
            ``group_member`` with ``is_current = 1``).
          * Authoritative "past event" = row in
            ``group_past_participant_user`` (mirrored to our
            ``group_past_participant``).
          * A contact CAN have both rows — they left the group
            once and rejoined.  In that case they are currently a
            member and must NOT be shown as former.
          * The device owner is *implicit* — msgstore.db does NOT
            list them in ``group_participant_user``, so we cannot
            show them as "former" just because they sent a
            message.  Only include the owner as former if there
            is explicit evidence (a past_participant row or an
            ``is_current = 0`` group_member row).
        """
        # Remove old widget if exists
        if hasattr(self, "_past_frame"):
            self._past_frame.setVisible(False)
            self._past_frame.deleteLater()
            del self._past_frame

        owner_cid = self._resolve_device_owner_cid(db)

        # Current members — authoritative: anyone with is_current=1 is NOT
        # a former member, regardless of any past_participant entries they
        # may also have (left-and-rejoined case).
        _current_rows = db.fetchall(
            "SELECT contact_id FROM group_member "
            "WHERE conversation_id = ? AND is_current = 1",
            (conv_id,),
        )
        current_member_ids = {r["contact_id"] for r in _current_rows}

        # Past participants from group_past_participant_user, BUT filtered
        # to exclude anyone who is currently a member (rejoined case).
        past = db.fetchall("""
            SELECT DISTINCT c.id, c.resolved_name, c.phone_number, c.phone_jid,
                   c.wa_name, gpp.state, gpp.last_seen_ts
            FROM group_past_participant gpp
            JOIN contact c ON c.id = gpp.contact_id
            WHERE gpp.conversation_id = ?
              AND gpp.contact_id NOT IN (
                  SELECT gm.contact_id FROM group_member gm
                  WHERE gm.conversation_id = ? AND gm.is_current = 1
              )
            ORDER BY gpp.last_seen_ts DESC
        """, (conv_id, conv_id))

        # group_member with is_current=0 — they were added and have left
        left_members = db.fetchall("""
            SELECT DISTINCT c.id, c.resolved_name, c.phone_number, c.phone_jid,
                   c.wa_name, -1 AS state, gm.left_timestamp AS last_seen_ts
            FROM group_member gm
            JOIN contact c ON c.id = gm.contact_id
            WHERE gm.conversation_id = ? AND gm.is_current = 0
              AND gm.contact_id NOT IN (
                  SELECT gpp.contact_id FROM group_past_participant gpp
                  WHERE gpp.conversation_id = ?
              )
            ORDER BY gm.left_timestamp DESC
        """, (conv_id, conv_id))

        # Contacts who sent messages but aren't in any member table.
        # EXCLUDE the device owner here — they are implicit in msgstore.db
        # and sending messages in a group does NOT mean they left.
        known_ids = {r["id"] for r in past} | {r["id"] for r in left_members}
        params = [conv_id, conv_id, conv_id]
        owner_exclusion_sql = ""
        if owner_cid:
            owner_exclusion_sql = " AND m.sender_id != ?"
            params.append(owner_cid)
        msg_only = db.fetchall(f"""
            SELECT DISTINCT c.id, c.resolved_name, c.phone_number, c.phone_jid,
                   c.wa_name, -2 AS state, MAX(m.timestamp) AS last_seen_ts
            FROM message m
            JOIN contact c ON c.id = m.sender_id
            WHERE m.conversation_id = ? AND m.sender_id IS NOT NULL
              AND m.sender_id NOT IN (
                  SELECT gm.contact_id FROM group_member gm WHERE gm.conversation_id = ?
              )
              AND m.sender_id NOT IN (
                  SELECT gpp.contact_id FROM group_past_participant gpp WHERE gpp.conversation_id = ?
              )
              {owner_exclusion_sql}
            GROUP BY m.sender_id
            ORDER BY MAX(m.timestamp) DESC
        """, tuple(params))
        # Filter out anyone already known (left_members) and anyone who is
        # currently a member (defensive — the SQL above should have handled
        # it, but this catches edge cases where group_member.is_current=1
        # but the sender_id lookup went through a different contact_id).
        msg_only = [
            r for r in msg_only
            if r["id"] not in known_ids and r["id"] not in current_member_ids
        ]

        all_past = list(past) + list(left_members) + list(msg_only)

        if not all_past:
            return

        from PySide6.QtWidgets import QFrame, QVBoxLayout, QLabel
        frame = QFrame()
        frame.setObjectName("pastParticipantsFrame")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(12, 8, 12, 8)

        header = QLabel(f"Former Members ({len(all_past)})")
        header.setStyleSheet("font-weight: bold; font-size: 13px; color: #90a4ae;")
        lay.addWidget(header)

        # Check if group creator (superadmin) is among former members
        creator_ids = set()
        try:
            _cr = db.fetchall(
                "SELECT contact_id FROM group_member WHERE conversation_id = ? AND role = 'superadmin'",
                (conv_id,))
            creator_ids = {r[0] for r in _cr if r[0]}
        except Exception:
            pass

        for p in all_past:
            name = p["resolved_name"] or p["wa_name"] or p["phone_number"] or p["phone_jid"] or "Unknown"
            phone = p["phone_number"] or ""
            if not phone and p["phone_jid"]:
                phone = p["phone_jid"].split("@")[0] if "@" in p["phone_jid"] else p["phone_jid"]
            last_seen = p["last_seen_ts"]
            last_str = ""
            if last_seen:
                try:
                    from datetime import datetime as _dt
                    last_str = _dt.fromtimestamp(last_seen / 1000).strftime("%Y-%m-%d %H:%M")
                except (ValueError, OSError):
                    pass

            # Card row with avatar
            from PySide6.QtWidgets import QHBoxLayout, QPushButton
            from PySide6.QtGui import QPixmap, QColor, QPainter, QFont, QPainterPath
            row_w = QFrame()
            row_w.setStyleSheet(
                "QFrame { border-bottom: 1px solid rgba(128,128,128,0.1); padding: 4px 0; }"
            )
            row_l = QHBoxLayout(row_w)
            row_l.setContentsMargins(4, 4, 4, 4)
            row_l.setSpacing(10)

            # Avatar
            avatar_lbl = QLabel()
            avatar_lbl.setFixedSize(36, 36)
            contact_id = p["id"]
            # Try to get avatar from contact
            avatar_blob = db.fetchone(
                "SELECT avatar_blob FROM contact WHERE id = ?", (contact_id,)
            )
            if avatar_blob and avatar_blob[0] and len(avatar_blob[0]) > 100:
                pxm = QPixmap()
                pxm.loadFromData(avatar_blob[0])
                if not pxm.isNull():
                    scaled = pxm.scaled(36, 36, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
                    avatar_lbl.setPixmap(scaled)
                    avatar_lbl.setStyleSheet("border-radius: 18px;")
                else:
                    avatar_lbl.setText(name[0].upper() if name else "?")
                    avatar_lbl.setAlignment(Qt.AlignCenter)
                    avatar_lbl.setStyleSheet(
                        "background: #607d8b; color: white; border-radius: 18px;"
                        " font-weight: bold; font-size: 14px;"
                    )
            else:
                avatar_lbl.setText(name[0].upper() if name else "?")
                avatar_lbl.setAlignment(Qt.AlignCenter)
                avatar_lbl.setStyleSheet(
                    "background: #607d8b; color: white; border-radius: 18px;"
                    " font-weight: bold; font-size: 14px;"
                )
            row_l.addWidget(avatar_lbl)

            # Name + phone + last seen + badges
            contact_id = p["id"]
            info_html = f"<b style='font-size:12px;'>{name}</b>"
            if owner_cid and contact_id == owner_cid:
                info_html += " <span style='color:#e65100;font-size:9px;font-weight:bold;background:rgba(230,81,0,0.1);padding:1px 4px;border-radius:3px;'>Device Owner</span>"
            if contact_id in creator_ids:
                info_html += " <span style='color:#ff6f00;font-size:9px;font-weight:bold;background:rgba(255,111,0,0.1);padding:1px 4px;border-radius:3px;'>Group Creator</span>"
            if phone:
                info_html += f"<br><span style='color:#667781;font-size:10px;'>+{phone}</span>"
            if last_str:
                info_html += f"<br><span style='color:#90a4ae;font-size:10px;'>Last active: {last_str}</span>"
            state = p["state"] if "state" in p.keys() else -2
            if state == -2:
                info_html += " <span style='color:#ff9800;font-size:9px;'>(found via messages)</span>"
            info_lbl = QLabel(info_html)
            info_lbl.setTextFormat(Qt.RichText)
            row_l.addWidget(info_lbl, 1)

            # "View Profile" button for former members
            from PySide6.QtWidgets import QPushButton
            _cid = contact_id
            _cname = name
            profile_btn = QPushButton("\U0001F464")
            profile_btn.setFixedSize(28, 28)
            profile_btn.setCursor(Qt.PointingHandCursor)
            profile_btn.setToolTip(f"View {name}'s profile in this group")
            profile_btn.setStyleSheet(
                "QPushButton { background: rgba(0,137,123,0.08); border: 1px solid rgba(0,137,123,0.2);"
                " border-radius: 14px; font-size: 12px; }"
                "QPushButton:hover { background: rgba(0,137,123,0.2); }"
            )
            profile_btn.clicked.connect(
                lambda checked=False, c=_cid, n=_cname: self._show_member_profile(c, n)
            )
            row_l.addWidget(profile_btn)

            lay.addWidget(row_w)

        # No limit — show all former members

        # Insert past participants frame after the member table
        # Find the main layout and add it
        parent_layout = self._table.parent().layout() if self._table.parent() else None
        if parent_layout:
            parent_layout.addWidget(frame)
        self._past_frame = frame

    # ------------------------------------------------------------------ #
    # Helpers
    # ------------------------------------------------------------------ #

    def _filter_members(self, *_args) -> None:
        """Filter members by free-text AND a single resolution-mode
        dropdown that combines two orthogonal flags:
          • phone-resolved (real phone we know) vs LID-only
          • saved (in device address book) vs unsaved
        All filters AND together; the count label always shows the
        full split so the investigator sees both axes at a glance."""
        text = self._filter_input.text().strip().lower() if hasattr(self, "_filter_input") else ""
        try:
            mode = self._resolution_filter.currentData() or "all"
        except Exception:
            mode = "all"
        visible = 0
        resolved_count = 0
        lid_only_count = 0
        saved_count = 0
        unsaved_count = 0
        total = self._table.rowCount()
        for row in range(total):
            phone_item = self._table.item(row, COL_PHONE)
            name_item = self._table.item(row, COL_NAME)
            is_resolved = bool(phone_item and phone_item.data(Qt.UserRole + 1))
            is_saved = bool(name_item and name_item.data(Qt.UserRole + 2))
            if is_resolved:
                resolved_count += 1
            else:
                lid_only_count += 1
            if is_saved:
                saved_count += 1
            else:
                unsaved_count += 1

            show = True
            if text:
                role_item = self._table.item(row, COL_ROLE)
                method_item = self._table.item(row, COL_JOIN_METHOD)
                haystack = " ".join(filter(None, [
                    name_item.text() if name_item else "",
                    phone_item.text() if phone_item else "",
                    role_item.text() if role_item else "",
                    method_item.text() if method_item else "",
                    name_item.toolTip() if name_item else "",
                    phone_item.toolTip() if phone_item else "",
                ])).lower()
                show = text in haystack
            if mode == "resolved" and not is_resolved:
                show = False
            elif mode == "lid_only" and is_resolved:
                show = False
            elif mode == "saved" and not is_saved:
                show = False
            elif mode == "unsaved" and is_saved:
                show = False

            self._table.setRowHidden(row, not show)
            if show:
                visible += 1

        suffix = (f"  ·  {resolved_count} resolved · {lid_only_count} LID-only"
                  f"  ·  {saved_count} saved · {unsaved_count} unsaved")
        if text or mode != "all":
            self._member_count_label.setText(f"({visible}/{total}){suffix}")
        else:
            self._member_count_label.setText(f"({total}){suffix}")

    def _get_contact_id_for_row(self, row: int) -> int | None:
        """Get the contact_id stored in the avatar item's UserRole for a table row."""
        avatar_item = self._table.item(row, COL_AVATAR)
        if avatar_item:
            return avatar_item.data(Qt.UserRole)
        return None

    def _get_member_name_for_row(self, row: int) -> str:
        """Get the display name for a table row."""
        name_item = self._table.item(row, COL_NAME)
        return name_item.text() if name_item else "Unknown"

    def _member_context_menu(self, pos) -> None:
        """Show context menu on right-click of a member row."""
        index = self._table.indexAt(pos)
        if not index.isValid():
            return
        row = index.row()
        contact_id = self._get_contact_id_for_row(row)
        if not contact_id:
            return
        member_name = self._get_member_name_for_row(row)

        from PySide6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background: {"#ffffff" if self._tm.is_light else "#1a2630"};
                border: 1px solid {self._c_border};
                border-radius: 6px;
                padding: 4px;
            }}
            QMenu::item {{
                padding: 6px 20px;
                font-size: 12px;
                color: {self._c_text};
            }}
            QMenu::item:selected {{
                background: {self._c_accent_bg};
                color: {self._c_accent};
            }}
        """)

        act_profile = menu.addAction("\U0001F464 View Profile in This Group")
        act_report = menu.addAction("\U0001F4CB Generate Contact Report (Group-Scoped)")
        act_report_full = menu.addAction("\U0001F4C4 Generate Full Contact Report")
        menu.addSeparator()
        act_go_to_chat = menu.addAction("\U0001F4AC Go to Direct Chat")

        action = menu.exec(self._table.viewport().mapToGlobal(pos))
        if action == act_profile:
            self._show_member_profile(contact_id, member_name)
        elif action == act_report:
            self._generate_member_report(contact_id, member_name, group_scoped=True)
        elif action == act_report_full:
            self._generate_member_report(contact_id, member_name, group_scoped=False)
        elif action == act_go_to_chat:
            self._go_to_direct_chat(contact_id)

    def _on_member_double_click(self, index) -> None:
        """Open member profile dialog on double-click."""
        if not index.isValid():
            return
        row = index.row()
        contact_id = self._get_contact_id_for_row(row)
        if contact_id:
            member_name = self._get_member_name_for_row(row)
            self._show_member_profile(contact_id, member_name)

    def _show_member_profile(self, contact_id: int, member_name: str) -> None:
        """Open the member group profile dialog."""
        if not self._conv_id:
            return
        dlg = MemberGroupProfileDialog(
            contact_id=contact_id,
            conversation_id=self._conv_id,
            group_name=self._group_name,
            member_name=member_name,
            parent=self,
        )
        dlg.go_to_message.connect(self._on_go_to_message)
        dlg.exec()

    def _generate_member_report(self, contact_id: int, member_name: str,
                                group_scoped: bool = True) -> None:
        """Generate a contact report, optionally scoped to this group."""
        import webbrowser
        from app.views.pages._report_loader import load_contact_report
        generate_contact_report = load_contact_report()

        db = Database.get()
        db_path = db.path

        safe_name = "".join(
            c if c.isalnum() or c in " _-" else "_"
            for c in member_name
        )[:40]
        safe_group = "".join(
            c if c.isalnum() or c in " _-" else "_"
            for c in (self._group_name or "")
        )[:30]

        output_dir = db_path.parent / "reports"
        output_dir.mkdir(exist_ok=True)

        if group_scoped:
            output_file = output_dir / f"contact_report_{safe_name}_in_{safe_group}_{contact_id}.html"
        else:
            output_file = output_dir / f"contact_report_{safe_name}_{contact_id}.html"

        try:
            from PySide6.QtWidgets import QApplication
            QApplication.processEvents()

            result_path = generate_contact_report(
                analysis_db_path=str(db_path),
                contact_id=contact_id,
                output_path=str(output_file),
                group_conversation_id=self._conv_id if group_scoped else None,
            )
            webbrowser.open(result_path.as_uri())

        except Exception as e:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(
                self, "Report Generation Failed",
                f"Failed to generate contact report:\n\n{e}",
            )

    def _go_to_direct_chat(self, contact_id: int) -> None:
        """Navigate to the direct chat with this contact."""
        db = Database.get()
        conv = db.fetchone(
            "SELECT id FROM conversation WHERE contact_id = ? AND chat_type = 'individual' LIMIT 1",
            (contact_id,),
        )
        if conv:
            from app.views.main_window import MainWindow
            main = self.window()
            if isinstance(main, MainWindow):
                main._switch_to_conversation(conv["id"])
        else:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.information(
                self, "No Direct Chat",
                "No direct conversation found with this contact.",
            )

    def _populate_analytics(self, members, contrib_map, link_map, db, conv_id):
        """Populate the analytics section with top senders and activity info."""
        # Clear existing widgets
        while self._top_senders_layout.count():
            item = self._top_senders_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        while self._activity_layout.count():
            item = self._activity_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Build name map from current members
        name_map: dict[int, str] = {}
        for m in members:
            cid = m["contact_id"]
            name_map[cid] = (
                m["resolved_name"] or m["display_name"] or m["wa_name"]
                or m["phone_number"] or "Unknown"
            )

        # For any contributor NOT in the current-members list
        # (former members, non-member participants), resolve their
        # name directly from the contact table — otherwise top
        # senders would show "Unknown".
        unmapped = [cid for cid in contrib_map.keys() if cid not in name_map]
        if unmapped:
            try:
                from app.services.database import Database
                _db = Database.get()
                placeholders = ",".join("?" * len(unmapped))
                rows = _db.fetchall(
                    f"SELECT id, resolved_name, display_name, wa_name, "
                    f"       phone_number, phone_jid "
                    f"FROM contact WHERE id IN ({placeholders})",
                    tuple(unmapped),
                )
                for r in rows:
                    resolved = (
                        r["resolved_name"] or r["display_name"] or r["wa_name"]
                        or (f"+{r['phone_number']}" if r["phone_number"] else None)
                        or (r["phone_jid"].split("@", 1)[0] if r["phone_jid"] else None)
                        or f"cid#{r['id']}"
                    )
                    name_map[r["id"]] = resolved
            except Exception:
                pass

        # Top 5 senders by message count
        top_senders = sorted(
            [(cid, stats["total"], name_map.get(cid, f"cid#{cid}"))
             for cid, stats in contrib_map.items() if stats["total"] > 0],
            key=lambda x: -x[1]
        )[:5]

        if top_senders:
            max_msgs = top_senders[0][1] if top_senders else 1
            colors = (["#00796b", "#00897b", "#009688", "#00a89a", "#26a69a"]
                      if self._tm.is_light else
                      ["#00bcd4", "#26c6da", "#4dd0e1", "#80deea", "#b2ebf2"])

            title = QLabel("Top Senders")
            title.setStyleSheet(f"color: {self._c_text2}; font-size: 10px;")
            self._top_senders_layout.addWidget(title)

            for i, (cid, count, name) in enumerate(top_senders):
                card = QFrame()
                card.setFixedHeight(48)
                card.setStyleSheet(f"""
                    QFrame {{
                        background: rgba({int(colors[i][1:3], 16)},{int(colors[i][3:5], 16)},{int(colors[i][5:7], 16)},0.08);
                        border: 1px solid rgba({int(colors[i][1:3], 16)},{int(colors[i][3:5], 16)},{int(colors[i][5:7], 16)},0.2);
                        border-radius: 6px;
                    }}
                """)
                cl = QVBoxLayout(card)
                cl.setContentsMargins(10, 4, 10, 4)
                cl.setSpacing(1)

                name_lbl = QLabel(f"#{i+1} {name}")
                name_lbl.setStyleSheet(f"color: {colors[i]}; font-size: 11px; font-weight: bold;")
                name_lbl.setToolTip(name)
                cl.addWidget(name_lbl)

                pct = count / max_msgs * 100
                count_lbl = QLabel(f"{count:,} messages")
                count_lbl.setStyleSheet(f"color: {self._c_text2}; font-size: 9px;")
                cl.addWidget(count_lbl)

                self._top_senders_layout.addWidget(card, 1)

            self._top_senders_layout.addStretch()

        # Activity summary cards
        # Most active hour
        hour_rows = db.fetchall(
            """
            SELECT CAST(strftime('%H', datetime(m.timestamp/1000, 'unixepoch', 'localtime')) AS INTEGER) AS hour,
                   COUNT(*) as cnt
            FROM message m
            WHERE m.conversation_id = ? AND m.message_type != 7
            GROUP BY hour ORDER BY cnt DESC LIMIT 3
            """,
            (conv_id,),
        )
        if hour_rows:
            peak_hour = hour_rows[0]["hour"]
            peak_label = f"{peak_hour:02d}:00 - {(peak_hour+1) % 24:02d}:00"
            self._add_activity_card("Peak Hour", peak_label, "#ffb300")

        # Most active day of week
        dow_rows = db.fetchall(
            """
            SELECT CASE CAST(strftime('%w', datetime(m.timestamp/1000, 'unixepoch', 'localtime')) AS INTEGER)
                     WHEN 0 THEN 'Sunday' WHEN 1 THEN 'Monday' WHEN 2 THEN 'Tuesday'
                     WHEN 3 THEN 'Wednesday' WHEN 4 THEN 'Thursday'
                     WHEN 5 THEN 'Friday' WHEN 6 THEN 'Saturday' END AS day_name,
                   COUNT(*) as cnt
            FROM message m
            WHERE m.conversation_id = ? AND m.message_type != 7
            GROUP BY day_name ORDER BY cnt DESC LIMIT 1
            """,
            (conv_id,),
        )
        if dow_rows:
            self._add_activity_card("Most Active Day", dow_rows[0]["day_name"], "#66bb6a")

        # Average messages per day
        span = db.fetchone(
            """
            SELECT MIN(m.timestamp) as first_ts, MAX(m.timestamp) as last_ts,
                   COUNT(*) as total
            FROM message m
            WHERE m.conversation_id = ? AND m.message_type != 7
            """,
            (conv_id,),
        )
        if span and span["first_ts"] and span["last_ts"]:
            days = max(1, (span["last_ts"] - span["first_ts"]) / (1000 * 86400))
            avg = span["total"] / days
            self._add_activity_card("Avg/Day", f"{avg:.1f} msgs", "#29b6f6")

        # Media ratio
        total_msg_count = sum(c["total"] for c in contrib_map.values())
        total_media_count = sum(
            c["image"] + c["video"] + c["audio"]
            for c in contrib_map.values()
        )
        if total_msg_count > 0:
            media_pct = total_media_count / total_msg_count * 100
            self._add_activity_card("Media %", f"{media_pct:.1f}%", "#ab47bc")

        self._activity_layout.addStretch()

    def _add_activity_card(self, label: str, value: str, color: str):
        card = QFrame()
        card.setFixedSize(120, 52)
        card.setStyleSheet(f"""
            QFrame {{
                background: {self._c_bg_card};
                border: 1px solid {self._c_border};
                border-radius: 6px;
            }}
        """)
        cl = QVBoxLayout(card)
        cl.setContentsMargins(10, 6, 10, 6)
        cl.setSpacing(2)

        val_lbl = QLabel(value)
        val_lbl.setStyleSheet(f"color: {color}; font-size: 13px; font-weight: bold;")
        val_lbl.setAlignment(Qt.AlignCenter)
        cl.addWidget(val_lbl)

        lbl = QLabel(label)
        lbl.setStyleSheet(f"color: {self._c_text2}; font-size: 9px;")
        lbl.setAlignment(Qt.AlignCenter)
        cl.addWidget(lbl)

        self._activity_layout.addWidget(card)

    # ------------------------------------------------------------------ #
    # Generate Group Report
    # ------------------------------------------------------------------ #

    def _export_participants_xlsx(self) -> None:
        """Export current, former, and message-derived group participants."""
        if not self._conv_id:
            return

        from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox

        db = Database.get()
        safe_group = "".join(
            c if c.isalnum() or c in " _-" else "_"
            for c in (self._group_name or f"Group_{self._conv_id}")
        )[:50].strip() or f"Group_{self._conv_id}"
        default_dir = db.path.parent / "exports"
        default_dir.mkdir(parents=True, exist_ok=True)
        default_path = default_dir / f"group_participants_{safe_group}_{self._conv_id}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Group Participants XLSX",
            str(default_path),
            "Excel Workbook (*.xlsx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            self._participants_xlsx_btn.setText("Exporting...")
            self._participants_xlsx_btn.setEnabled(False)
            QApplication.processEvents()

            headers, rows = self._build_participants_export_rows(db, self._conv_id)
            metadata_rows = self._build_group_metadata_export_rows(db, self._conv_id, rows)
            preferred = [
                ("dp", "DP"),
                ("display_name", "Name"),
                ("group_specific_name", "Group Specific Name"),
                ("phone_number", "Phone"),
                ("role", "Role"),
                ("label", "Label"),
                ("total_messages", "Messages"),
                ("text_count", "Text"),
                ("image_count", "Images"),
                ("video_count", "Videos"),
                ("audio_count", "Voice"),
                ("link_count", "Links"),
                ("first_message_time", "First Message"),
                ("last_message_time", "Last Message"),
                ("join_time", "Joined"),
                ("join_method_label", "Join Method"),
                ("participant_origin_category", "Origin"),
                ("participant_source", "Source Detail"),
                ("is_device_owner", "Device Owner"),
                ("is_non_roster_sender", "Non-roster Sender"),
                ("non_roster_sender_type", "Non-roster Type"),
            ]
            used = {key for key, _ in preferred if key != "dp"}
            columns = preferred + [
                (h, h.replace("_", " ").title()) for h in headers if h not in used
            ]
            self._write_participants_xlsx(file_path, columns, rows, metadata_rows)

            current_count = sum(1 for r in rows if r["is_current_member"] == "Yes")
            past_db_count = sum(
                1 for r in rows
                if r["has_past_participant_db_record"] == "Yes"
                and r["is_current_member"] != "Yes"
            )
            message_only_count = sum(1 for r in rows if r["message_only_participant"] == "Yes")
            QMessageBox.information(
                self,
                "Participants XLSX Exported",
                f"Exported {len(rows):,} participant rows to:\n\n{file_path}\n\n"
                f"Current: {current_count:,}\n"
                f"Past by DB: {past_db_count:,}\n"
                f"Message-only: {message_only_count:,}",
            )
        except Exception as e:
            QMessageBox.warning(
                self,
                "XLSX Export Failed",
                f"Failed to export group participants:\n\n{e}",
            )
        finally:
            self._participants_xlsx_btn.setText("\u21E9 Participants XLSX")
            self._participants_xlsx_btn.setEnabled(True)

    def _build_group_metadata_export_rows(self, db, conv_id: int,
                                          participant_rows: list[dict]) -> list[tuple[str, str]]:
        """Collect group-level context for the XLSX metadata sheet."""
        conv_cols = set()
        try:
            conv_cols = {r[1] for r in db.fetchall("PRAGMA table_info(conversation)")}
        except Exception:
            pass

        wanted = [
            "id", "source_chat_id", "jid_raw_string", "chat_type",
            "display_name", "subject", "description", "created_timestamp",
            "participant_count", "message_count", "media_count",
            "first_message_ts", "last_message_ts", "group_type",
            "community_parent_id", "addressing_mode", "ephemeral_duration",
            "is_archived", "is_pinned", "is_muted", "is_locked",
            "participation_status", "announcement_group", "restrict_mode",
            "require_membership_approval", "member_add_mode", "creator_jid_raw",
        ]
        cols = [c for c in wanted if c == "id" or c in conv_cols]
        conv = {}
        try:
            row = db.fetchone(
                f"SELECT {', '.join(cols)} FROM conversation WHERE id = ?",
                (conv_id,),
            )
            conv = dict(row) if row else {}
        except Exception:
            conv = {"id": conv_id}

        owner_cid = self._resolve_device_owner_cid(db)
        owner_info = self._detect_owner_membership(db, conv_id, owner_cid)

        creator_name = ""
        try:
            creator = db.fetchone("""
                SELECT COALESCE(c.resolved_name, c.display_name, c.wa_name,
                                c.phone_number, c.phone_jid, c.lid_jid) AS name,
                       c.phone_number, c.phone_jid, c.lid_jid
                FROM group_member gm
                JOIN contact c ON c.id = gm.contact_id
                WHERE gm.conversation_id = ? AND gm.role = 'superadmin'
                LIMIT 1
            """, (conv_id,))
            if creator:
                creator_name = creator["name"] or ""
                if creator["phone_number"]:
                    creator_name += f" ({creator['phone_number']})"
        except Exception:
            pass

        current_count = sum(1 for r in participant_rows if r.get("is_current_member") == "Yes")
        past_db_count = sum(
            1 for r in participant_rows
            if r.get("has_past_participant_db_record") == "Yes"
            and r.get("is_current_member") != "Yes"
        )
        past_db_history_count = sum(
            1 for r in participant_rows
            if r.get("has_past_participant_db_record") == "Yes"
        )
        message_only_count = sum(1 for r in participant_rows if r.get("message_only_participant") == "Yes")
        owner_rows = [r for r in participant_rows if r.get("is_device_owner") == "Yes"]
        owner_name = owner_rows[0].get("display_name", "") if owner_rows else ""

        # ---- Device owner: derive friendly role / can-send / is-current
        # flags so the header tells the investigator at a glance whether
        # the owner is in the group and whether they can post. ----
        owner_phone_disp = ""
        if owner_rows:
            ph = (owner_rows[0].get("phone_number") or "").strip()
            if not ph:
                ph = (owner_rows[0].get("lid_masked_phone") or "").strip()
            if ph and not ph.startswith("+"):
                ph = "+" + ph
            owner_phone_disp = ph

        owner_status_raw = (owner_info.get("status") or "").lower()
        owner_role_label = {
            "creator":        "Creator (super admin)",
            "admin":          "Admin",
            "member":         "Member",
            "current":        "Member",
            "former":         "No longer in group",
            "former_left":    "No longer in group (left voluntarily)",
            "former_removed": "No longer in group (removed by admin)",
            "unknown":        "Unknown",
        }.get(owner_status_raw, owner_status_raw.title() or "Unknown")

        if owner_status_raw in ("creator", "admin", "member", "current"):
            owner_is_current_lbl = "Yes"
        elif owner_status_raw in ("former", "former_left", "former_removed"):
            owner_is_current_lbl = "No"
        else:
            owner_is_current_lbl = "Unknown"

        announcement_only = bool(conv.get("announcement_group"))
        if announcement_only and owner_status_raw in ("creator", "admin"):
            owner_can_send_lbl = "Yes (admin in announcement group)"
        elif announcement_only and owner_status_raw in ("member", "current"):
            owner_can_send_lbl = "No (announcement group — only admins can send)"
        elif announcement_only and owner_status_raw in ("former", "former_left", "former_removed"):
            owner_can_send_lbl = "No (no longer in group)"
        elif announcement_only:
            owner_can_send_lbl = "Unknown (announcement group)"
        elif owner_status_raw in ("creator", "admin", "member", "current"):
            owner_can_send_lbl = "Yes"
        elif owner_status_raw in ("former", "former_left", "former_removed"):
            owner_can_send_lbl = "No (no longer in group)"
        else:
            owner_can_send_lbl = "Unknown"

        def _owner_ts(ts):
            if not ts:
                return ""
            try:
                return _fmt_ts(int(ts), "forensic_tz")
            except Exception:
                return ""
        owner_joined_lbl = _owner_ts(owner_info.get("join_ts"))
        owner_left_lbl   = _owner_ts(owner_info.get("leave_ts"))

        SRC_LABELS = {
            "participation_status": "msgstore.db conversation.participation_status (authoritative)",
            "system_event":         "msgstore.db system_event records",
            "inferred":              "inferred from sent messages (no explicit record)",
            "none":                  "no data",
        }
        owner_source_lbl = SRC_LABELS.get(
            owner_info.get("source", "none"),
            str(owner_info.get("source", "")),
        )

        def val(key: str):
            return conv.get(key, "")

        rows = [
            ("Export Type", "Group participants workbook"),
            ("Export Generated", _fmt_ts(int(datetime.now().timestamp() * 1000), "forensic_tz")),
            ("Conversation ID", str(conv_id)),
            ("Source Chat ID", str(val("source_chat_id") or "")),
            ("Group Name", str(val("display_name") or self._group_name or "")),
            ("Group Subject", str(val("subject") or "")),
            ("Group JID", str(val("jid_raw_string") or "")),
            ("Chat Type", str(val("chat_type") or "")),
            ("Group Description", str(val("description") or "")),
            ("Created Timestamp", str(val("created_timestamp") or "")),
            ("Created Time", self._csv_ts(val("created_timestamp"))),
            ("First Message Time", self._csv_ts(val("first_message_ts"))),
            ("Last Message Time", self._csv_ts(val("last_message_ts"))),
            ("Creator / Super Admin", creator_name),
            ("Creator JID Raw", str(val("creator_jid_raw") or "")),
            ("Participant Count In Conversation", str(val("participant_count") or "")),
            ("Participant Rows Exported", f"{len(participant_rows):,}"),
            ("Current Members Exported", f"{current_count:,}"),
            ("Past Participants From DB", f"{past_db_count:,}"),
            ("Contacts With Any DB Past-Participant History", f"{past_db_history_count:,}"),
            ("Message-Only Participants", f"{message_only_count:,}"),
            ("Conversation Message Count", str(val("message_count") or "")),
            ("Conversation Media Count", str(val("media_count") or "")),
            ("Group Type", str(val("group_type") or "")),
            ("Community Parent Conversation ID", str(val("community_parent_id") or "")),
            ("Addressing Mode", str(val("addressing_mode") or "")),
            ("Ephemeral Duration Seconds", str(val("ephemeral_duration") or "")),
            ("Archived", self._yn(val("is_archived"))),
            ("Pinned", self._yn(val("is_pinned"))),
            ("Muted", self._yn(val("is_muted"))),
            ("Locked", self._yn(val("is_locked"))),
            ("Participation Status Raw", str(val("participation_status") or "")),
            # Group-level permissions, phrased for humans (raw flags from
            # wa.db are still implicit in these labels)
            ("Send Messages", "Only admins (announcement group)" if val("announcement_group") else "Anyone"),
            ("Edit Group Info", "Only admins" if val("restrict_mode") else "Anyone"),
            ("Add Members", "Only admins" if val("member_add_mode") else "Anyone"),
            ("Join Approval", "Admin approval required" if val("require_membership_approval") else "Open / invite-only"),
            # Device owner facts - the investigator wants to see at a
            # glance "is the case-phone owner currently in this group,
            # and can they post here"
            ("Device Owner Contact ID", str(owner_cid or "")),
            ("Device Owner Phone", owner_phone_disp),
            ("Device Owner Name", owner_name),
            ("Device Owner Is Current Member", owner_is_current_lbl),
            ("Device Owner Role In Group", owner_role_label),
            ("Device Owner Can Send Messages", owner_can_send_lbl),
            ("Device Owner Joined", owner_joined_lbl),
            ("Device Owner Left", owner_left_lbl),
            ("Device Owner Status Source", owner_source_lbl),
            ("Device Owner Status Detail", str(owner_info.get("detail", ""))),
        ]
        return rows

    def _write_participants_xlsx_openpyxl(self, file_path: str, columns: list[tuple[str, str]],
                                          rows: list[dict],
                                          metadata_rows: list[tuple[str, str]]) -> None:
        """Write the participant workbook using openpyxl.

        Forensic-grade output:
          • DP image is anchored ONE-cell with editAs='oneCell' so it
            moves and sizes with the cell when the user reorders rows
            via the auto-filter (default 'twoCell' anchor lets the
            picture float and end up on the wrong row).
          • LID-only rows get amber row tint + the masked phone
            ("+91∙∙∙∙∙∙∙∙58") shown in the Phone column so the column
            never lies; the actual LID lives in lid_jid alongside.
          • Auto-filter is set on the Table only - having BOTH
            ws.auto_filter.ref AND a Table caused Excel to show
            duplicate filter dropdowns or drop one of them.
          • Numeric cells get number formats and timestamp_ms cells
            display in seconds + as a UTC datetime.
        """
        from io import BytesIO
        from datetime import datetime as _dt, timezone as _tz

        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
        from openpyxl.utils.units import pixels_to_EMU
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = Workbook()
        ws = wb.active
        ws.title = "Participants"
        meta_ws = wb.create_sheet("Group Metadata")

        title_fill = PatternFill("solid", fgColor="00897B")
        header_fill = PatternFill("solid", fgColor="E0F2F1")
        header_font = Font(bold=True, color="263238")
        title_font = Font(size=14, bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D7DEE2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Row tints for forensic at-a-glance state
        lid_only_fill = PatternFill("solid", fgColor="FFF4E0")     # amber
        saved_fill    = PatternFill("solid", fgColor="E8F5E9")     # mint

        ws.cell(row=1, column=1, value=f"{self._group_name or 'Group'} participants")
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).fill = title_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

        for col_idx, (_, label) in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        numeric_keys = {
            "conversation_id", "contact_id", "linked_device_count",
            "join_timestamp_ms", "left_timestamp_ms",
            "past_last_seen_timestamp_ms", "first_message_id",
            "first_source_msg_id", "first_message_timestamp_ms",
            "last_message_id", "last_source_msg_id",
            "last_message_timestamp_ms", "total_messages",
            "sent_from_device_owner_count", "received_from_participant_count",
            "chat_messages_excluding_system", "text_count", "image_count",
            "video_count", "audio_count", "document_count", "sticker_count",
            "gif_count", "location_count", "contact_card_count", "poll_count",
            "call_count", "system_count", "other_count", "media_count",
            "link_count", "forwarded_count", "starred_count", "edited_count",
            "revoked_count", "ephemeral_count", "reaction_given_count",
            "reaction_received_count", "mention_made_count",
            "mention_received_count", "membership_record_count",
            "platform_confidence",
        }
        ts_ms_keys = {k for k in numeric_keys if k.endswith("_timestamp_ms")}

        # Width of the DP column in pixels - matches our 32x32 image
        # plus a tiny breathing margin.  The image is anchored to fit
        # inside this cell exactly.
        DP_PX = 36
        ROW_PX = 36

        for row_idx, row in enumerate(rows, 3):
            ws.row_dimensions[row_idx].height = ROW_PX * 0.75  # px -> pt
            # Row-level tint based on resolution / saved state.  Saved
            # wins because it's a stronger signal (you trust the name).
            is_resolved = row.get("is_phone_resolved") == "Yes"
            is_saved = row.get("is_saved_in_address_book") == "Yes"
            row_tint = (saved_fill if is_saved
                        else (None if is_resolved else lid_only_fill))

            for col_idx, (key, _) in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if key == "dp":
                    value = ""
                elif key == "phone_number":
                    # Phone column NEVER lies - if the contact's phone
                    # is genuinely unresolved, surface WhatsApp's masked
                    # hint here ("+91∙∙∙∙∙∙∙∙58") and italicise it so
                    # the investigator sees at a glance this isn't the
                    # full number.  The actual LID is in lid_jid.
                    real_phone = (row.get("phone_number") or "").strip()
                    if real_phone:
                        value = ("+" + real_phone) if not real_phone.startswith("+") else real_phone
                    else:
                        masked = (row.get("lid_masked_phone") or "").strip()
                        value = masked or ""
                else:
                    value = row.get(key, "")

                if key in numeric_keys and value not in ("", None):
                    try:
                        value = float(value)
                        if value.is_integer():
                            value = int(value)
                    except (TypeError, ValueError):
                        pass
                cell.value = value

                # Numeric formats so 1773166063000 shows as "1,773,166,063,000"
                # not as a date oddly.  ts_ms_keys still display as int
                # but row already has a paired "*_time" string column.
                if key in numeric_keys and isinstance(value, (int, float)):
                    cell.number_format = "#,##0"

                cell.alignment = Alignment(
                    horizontal="right" if key in numeric_keys else "left",
                    vertical="center",
                    wrap_text=key.endswith("_text") or key in {
                        "logic_added_reason",
                        "non_roster_sender_note",
                    },
                )
                cell.border = border

                # Italic + amber for the masked-phone case so it reads
                # as "this is a hint, not the real number".
                if key == "phone_number" and not row.get("phone_number"):
                    if (row.get("lid_masked_phone") or "").strip():
                        cell.font = Font(italic=True, color="9B6E00")

                if row_tint and key != "dp":
                    cell.fill = row_tint

            avatar_blob = row.get("_avatar_blob")
            if avatar_blob:
                try:
                    image = XLImage(BytesIO(bytes(avatar_blob)))
                    image.width = 32
                    image.height = 32
                    # TwoCellAnchor with ``editAs='oneCell'`` =
                    # "move with cells, don't resize".  The image
                    # is bound to the FROM cell so when a row is
                    # hidden by an autofilter the image hides
                    # with it.  ``OneCellAnchor`` has no editAs
                    # attribute in OOXML, so the image floats and
                    # filtering would leave pictures stranded in
                    # rows that no longer correspond.
                    from_marker = AnchorMarker(col=0, row=row_idx - 1,
                                                colOff=pixels_to_EMU(2),
                                                rowOff=pixels_to_EMU(2))
                    to_marker = AnchorMarker(col=0, row=row_idx - 1,
                                              colOff=pixels_to_EMU(34),
                                              rowOff=pixels_to_EMU(34))
                    image.anchor = TwoCellAnchor(
                        editAs="oneCell",
                        _from=from_marker,
                        to=to_marker,
                    )
                    ws.add_image(image)
                except Exception:
                    pass

        ws.freeze_panes = "B3"
        last_row = max(2, len(rows) + 2)
        last_col = len(columns)

        # Use a styled Table - this gives Excel:
        #   • a single auto-filter dropdown row (no duplicates)
        #   • alternating row stripes
        #   • a sortable named range "GroupParticipants"
        # We do NOT also set ws.auto_filter.ref because that double-binds
        # the filter and Excel sometimes drops one of the two on save.
        if rows:
            table = Table(displayName="GroupParticipants",
                          ref=f"A2:{get_column_letter(last_col)}{last_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,  # row stripes fight our row tints
                showColumnStripes=False,
            )
            ws.add_table(table)
        else:
            ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_row}"

        widths = {
            1: 6,    # DP image column - exactly fits 32px image
            2: 28,   # Name
            3: 24,   # Group-specific name
            4: 22,   # Phone (formatted, or masked when LID-only)
            5: 14,   # Role
            6: 20,   # Label
            7: 12, 8: 10, 9: 10, 10: 10, 11: 10, 12: 10,
            13: 22, 14: 22, 15: 22, 16: 22,
        }
        for idx in range(1, last_col + 1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(idx, 18)

        meta_ws.cell(row=1, column=1, value="Group Metadata")
        meta_ws.cell(row=1, column=1).font = title_font
        meta_ws.cell(row=1, column=1).fill = title_fill
        meta_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        meta_ws.cell(row=2, column=1, value="Field")
        meta_ws.cell(row=2, column=2, value="Value")
        for c in (meta_ws.cell(row=2, column=1), meta_ws.cell(row=2, column=2)):
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        for row_idx, (field, value) in enumerate(metadata_rows, 3):
            meta_ws.cell(row=row_idx, column=1, value=field)
            meta_ws.cell(row=row_idx, column=2, value=value)
            for c in (meta_ws.cell(row=row_idx, column=1), meta_ws.cell(row=row_idx, column=2)):
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.border = border
        meta_ws.column_dimensions["A"].width = 34
        meta_ws.column_dimensions["B"].width = 86
        meta_ws.freeze_panes = "A3"
        meta_ws.auto_filter.ref = f"A2:B{max(2, len(metadata_rows) + 2)}"

        wb.save(file_path)

    def _write_participants_xlsx(self, file_path: str, columns: list[tuple[str, str]],
                                 rows: list[dict], metadata_rows: list[tuple[str, str]]) -> None:
        """Write a styled XLSX workbook with embedded DP images.

        Uses openpyxl for Microsoft Excel compatibility. The hand-written
        fallback below is kept only as an emergency path for environments
        that cannot install the spreadsheet dependency.
        """
        try:
            self._write_participants_xlsx_openpyxl(file_path, columns, rows, metadata_rows)
            return
        except ModuleNotFoundError as e:
            raise RuntimeError(
                "Participants XLSX export requires openpyxl and Pillow. "
                "Install requirements.txt, then retry."
            ) from e

        import re
        import zipfile
        from xml.sax.saxutils import escape

        numeric_keys = {
            "conversation_id", "contact_id", "linked_device_count",
            "join_timestamp_ms", "left_timestamp_ms",
            "past_last_seen_timestamp_ms", "first_message_id",
            "first_source_msg_id", "first_message_timestamp_ms",
            "last_message_id", "last_source_msg_id",
            "last_message_timestamp_ms", "total_messages",
            "sent_from_device_owner_count", "received_from_participant_count",
            "chat_messages_excluding_system", "text_count", "image_count",
            "video_count", "audio_count", "document_count", "sticker_count",
            "gif_count", "location_count", "contact_card_count", "poll_count",
            "call_count", "system_count", "other_count", "media_count",
            "link_count", "forwarded_count", "starred_count", "edited_count",
            "revoked_count", "ephemeral_count", "reaction_given_count",
            "reaction_received_count", "mention_made_count",
            "mention_received_count", "membership_record_count",
            "platform_confidence",
        }

        def col_name(index: int) -> str:
            name = ""
            while index:
                index, rem = divmod(index - 1, 26)
                name = chr(65 + rem) + name
            return name

        def xml_text(value) -> str:
            if value is None or isinstance(value, (bytes, bytearray)):
                return ""
            text = str(value)
            text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)
            return escape(text, {"'": "&apos;", '"': "&quot;"})

        shared_strings: list[str] = []
        shared_string_ids: dict[str, int] = {}

        def shared_string_index(value) -> int:
            text = "" if value is None else str(value)
            text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)
            if text not in shared_string_ids:
                shared_string_ids[text] = len(shared_strings)
                shared_strings.append(text)
            return shared_string_ids[text]

        def cell_xml(row_num: int, col_num: int, key: str, value, style: int = 3) -> str:
            ref = f"{col_name(col_num)}{row_num}"
            if value is None or value == "" or isinstance(value, (bytes, bytearray)):
                return f'<c r="{ref}" s="{style}"/>'
            if key in numeric_keys:
                try:
                    num = float(value)
                    if num.is_integer():
                        num = int(num)
                    return f'<c r="{ref}" s="{style}"><v>{num}</v></c>'
                except (TypeError, ValueError):
                    pass
            idx = shared_string_index(value)
            return f'<c r="{ref}" s="{style}" t="s"><v>{idx}</v></c>'

        def image_payload(blob) -> tuple[bytes, str] | None:
            if not blob:
                return None
            data = bytes(blob)
            if data.startswith(b"\xff\xd8"):
                return data, "jpeg"
            if data.startswith(b"\x89PNG\r\n\x1a\n"):
                return data, "png"
            try:
                from PySide6.QtCore import QByteArray, QBuffer, QIODevice
                from PySide6.QtGui import QImage

                img = QImage()
                if not img.loadFromData(data):
                    return None
                arr = QByteArray()
                buf = QBuffer(arr)
                buf.open(QIODevice.WriteOnly)
                img.save(buf, "PNG")
                buf.close()
                return bytes(arr), "png"
            except Exception:
                return None

        last_col = len(columns)
        last_row = max(2, len(rows) + 2)
        last_cell = f"{col_name(last_col)}{last_row}"
        title = f"{self._group_name or 'Group'} participants"

        col_widths = {
            1: 8, 2: 28, 3: 24, 4: 18, 5: 14, 6: 20,
            7: 12, 8: 10, 9: 10, 10: 10, 11: 10, 12: 10,
            13: 22, 14: 22, 15: 22, 16: 22,
        }
        cols_xml = ["<cols>"]
        for idx in range(1, last_col + 1):
            width = col_widths.get(idx, 18)
            cols_xml.append(f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>')
        cols_xml.append("</cols>")

        sheet_rows = []
        sheet_rows.append(
            f'<row r="1" ht="24" customHeight="1">'
            f'{cell_xml(1, 1, "title", title, 1)}</row>'
        )
        header_cells = [
            cell_xml(2, idx, "header", label, 2)
            for idx, (_, label) in enumerate(columns, 1)
        ]
        sheet_rows.append(f'<row r="2" ht="28" customHeight="1">{"".join(header_cells)}</row>')

        images: list[dict] = []
        image_exts: set[str] = set()
        for row_offset, data_row in enumerate(rows, 3):
            cells = []
            for col_idx, (key, _) in enumerate(columns, 1):
                if key == "dp":
                    cells.append(cell_xml(row_offset, col_idx, key, "", 3))
                else:
                    style = 4 if key in numeric_keys else 3
                    cells.append(cell_xml(row_offset, col_idx, key, data_row.get(key, ""), style))
            sheet_rows.append(
                f'<row r="{row_offset}" ht="34" customHeight="1">{"".join(cells)}</row>'
            )

            payload = image_payload(data_row.get("_avatar_blob"))
            if payload:
                img_bytes, ext = payload
                image_index = len(images) + 1
                image_exts.add(ext)
                images.append({
                    "index": image_index,
                    "row": row_offset - 1,  # zero-based for drawing anchors
                    "ext": ext,
                    "bytes": img_bytes,
                })

        drawing_tag = '<drawing r:id="rId1"/>' if images else ""
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<dimension ref="A1:{last_cell}"/>'
            '<sheetViews><sheetView workbookViewId="0">'
            '<pane xSplit="1" ySplit="2" topLeftCell="B3" activePane="bottomRight" state="frozen"/>'
            '<selection pane="bottomRight" activeCell="B3" sqref="B3"/>'
            '</sheetView></sheetViews>'
            '<sheetFormatPr defaultRowHeight="15"/>'
            f'{"".join(cols_xml)}'
            f'<sheetData>{"".join(sheet_rows)}</sheetData>'
            f'<mergeCells count="1"><mergeCell ref="A1:{col_name(last_col)}1"/></mergeCells>'
            f'<autoFilter ref="A2:{last_cell}"/>'
            f'{drawing_tag}'
            '</worksheet>'
        )

        metadata_sheet_rows = [
            f'<row r="1" ht="24" customHeight="1">{cell_xml(1, 1, "title", "Group Metadata", 1)}</row>',
            f'<row r="2" ht="24" customHeight="1">{cell_xml(2, 1, "header", "Field", 2)}{cell_xml(2, 2, "header", "Value", 2)}</row>',
        ]
        for meta_idx, (field, value) in enumerate(metadata_rows, 3):
            metadata_sheet_rows.append(
                f'<row r="{meta_idx}" ht="22" customHeight="1">'
                f'{cell_xml(meta_idx, 1, "metadata_field", field, 3)}'
                f'{cell_xml(meta_idx, 2, "metadata_value", value, 3)}'
                '</row>'
            )
        metadata_last_row = max(2, len(metadata_rows) + 2)
        metadata_sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<dimension ref="A1:B{metadata_last_row}"/>'
            '<sheetViews><sheetView workbookViewId="0">'
            '<pane ySplit="2" topLeftCell="A3" activePane="bottomLeft" state="frozen"/>'
            '<selection pane="bottomLeft" activeCell="A3" sqref="A3"/>'
            '</sheetView></sheetViews>'
            '<sheetFormatPr defaultRowHeight="15"/>'
            '<cols><col min="1" max="1" width="34" customWidth="1"/>'
            '<col min="2" max="2" width="86" customWidth="1"/></cols>'
            f'<sheetData>{"".join(metadata_sheet_rows)}</sheetData>'
            '<mergeCells count="1"><mergeCell ref="A1:B1"/></mergeCells>'
            f'<autoFilter ref="A2:B{metadata_last_row}"/>'
            '</worksheet>'
        )

        drawing_xml = ""
        drawing_rels = ""
        if images:
            anchors = []
            rels = []
            for img in images:
                rid = f"rId{img['index']}"
                rels.append(
                    f'<Relationship Id="{rid}" '
                    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
                    f'Target="../media/image{img["index"]}.{img["ext"]}"/>'
                )
                anchors.append(f'''
                <xdr:oneCellAnchor>
                  <xdr:from><xdr:col>0</xdr:col><xdr:colOff>95250</xdr:colOff><xdr:row>{img["row"]}</xdr:row><xdr:rowOff>95250</xdr:rowOff></xdr:from>
                  <xdr:ext cx="304800" cy="304800"/>
                  <xdr:pic>
                    <xdr:nvPicPr>
                      <xdr:cNvPr id="{img["index"]}" name="DP {img["index"]}" descr="Participant display picture"/>
                      <xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>
                    </xdr:nvPicPr>
                    <xdr:blipFill>
                      <a:blip r:embed="{rid}" cstate="print"/>
                      <a:stretch><a:fillRect/></a:stretch>
                    </xdr:blipFill>
                    <xdr:spPr>
                      <a:xfrm><a:off x="0" y="0"/><a:ext cx="304800" cy="304800"/></a:xfrm>
                      <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
                    </xdr:spPr>
                  </xdr:pic>
                  <xdr:clientData/>
                </xdr:oneCellAnchor>''')
            drawing_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f'{"".join(anchors)}</xdr:wsDr>'
            )
            drawing_rels = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                f'{"".join(rels)}</Relationships>'
            )

        image_defaults = ""
        if "jpeg" in image_exts:
            image_defaults += '<Default Extension="jpeg" ContentType="image/jpeg"/>'
        if "png" in image_exts:
            image_defaults += '<Default Extension="png" ContentType="image/png"/>'
        drawing_override = (
            '<Override PartName="/xl/drawings/drawing1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
            if images else ""
        )
        content_types = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            f'{image_defaults}'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
            '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
            '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
            f'{drawing_override}</Types>'
        )

        sheet_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'
            '</Relationships>'
            if images else None
        )

        styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
          <fonts count="3">
            <font><sz val="11"/><name val="Calibri"/></font>
            <font><b/><sz val="14"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>
            <font><b/><sz val="11"/><color rgb="FF263238"/><name val="Calibri"/></font>
          </fonts>
          <fills count="4">
            <fill><patternFill patternType="none"/></fill>
            <fill><patternFill patternType="gray125"/></fill>
            <fill><patternFill patternType="solid"><fgColor rgb="FF00897B"/></patternFill></fill>
            <fill><patternFill patternType="solid"><fgColor rgb="FFE0F2F1"/></patternFill></fill>
          </fills>
          <borders count="2">
            <border><left/><right/><top/><bottom/><diagonal/></border>
            <border><left style="thin"><color rgb="FFD7DEE2"/></left><right style="thin"><color rgb="FFD7DEE2"/></right><top style="thin"><color rgb="FFD7DEE2"/></top><bottom style="thin"><color rgb="FFD7DEE2"/></bottom><diagonal/></border>
          </borders>
          <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
          <cellXfs count="5">
            <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
            <xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1"><alignment vertical="center"/></xf>
            <xf numFmtId="0" fontId="2" fillId="3" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
            <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"><alignment vertical="center" wrapText="1"/></xf>
            <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"><alignment horizontal="right" vertical="center"/></xf>
          </cellXfs>
          <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
        </styleSheet>'''

        rels_root = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
          <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
          <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
          <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
        </Relationships>'''
        workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
          <sheets><sheet name="Participants" sheetId="1" r:id="rId1"/><sheet name="Group Metadata" sheetId="2" r:id="rId3"/></sheets>
        </workbook>'''
        workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
          <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
          <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
          <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>
          <Relationship Id="rId4" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
        </Relationships>'''
        app_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><Application>WAInsight</Application></Properties>'''
        core_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><dc:creator>WAInsight</dc:creator><dc:title>Group Participants Export</dc:title></cp:coreProperties>'''

        si_parts = []
        for text in shared_strings:
            escaped = xml_text(text)
            space = ' xml:space="preserve"' if text.strip() != text or "\n" in text else ""
            si_parts.append(f'<si><t{space}>{escaped}</t></si>')
        shared_strings_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{len(shared_strings)}" uniqueCount="{len(shared_strings)}">'
            f'{"".join(si_parts)}</sst>'
        )

        with zipfile.ZipFile(file_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("[Content_Types].xml", content_types)
            zf.writestr("_rels/.rels", rels_root)
            zf.writestr("docProps/app.xml", app_xml)
            zf.writestr("docProps/core.xml", core_xml)
            zf.writestr("xl/workbook.xml", workbook_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
            zf.writestr("xl/styles.xml", styles_xml)
            zf.writestr("xl/sharedStrings.xml", shared_strings_xml)
            zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
            zf.writestr("xl/worksheets/sheet2.xml", metadata_sheet_xml)
            if sheet_rels:
                zf.writestr("xl/worksheets/_rels/sheet1.xml.rels", sheet_rels)
                zf.writestr("xl/drawings/drawing1.xml", drawing_xml)
                zf.writestr("xl/drawings/_rels/drawing1.xml.rels", drawing_rels)
                for img in images:
                    zf.writestr(f"xl/media/image{img['index']}.{img['ext']}", img["bytes"])

    def _build_participants_export_rows(self, db, conv_id: int) -> tuple[list[str], list[dict]]:
        """Build one forensic export row per contact seen in this group."""
        headers = [
            "conversation_id", "group_name", "group_jid",
            "participant_source", "participant_origin_category",
            "contact_id", "display_name",
            "group_specific_name",
            "resolved_name", "address_book_name", "whatsapp_name",
            "phone_number", "phone_jid", "lid_jid", "lid_masked_phone",
            "is_phone_resolved", "is_saved_in_address_book", "is_business",
            "is_meta_verified", "platform_estimate", "platform_confidence",
            "linked_device_count", "status_text", "business_name",
            "company", "title",
            "is_device_owner",
            "has_group_member_record", "is_current_member",
            "has_past_participant_db_record", "db_past_participant_currently_former",
            "found_by_messages", "is_non_roster_sender", "non_roster_sender_type",
            "non_roster_sender_note", "logic_added_reason",
            "message_only_participant", "role", "label",
            "join_timestamp_ms", "join_time", "join_method_code",
            "join_method_label", "added_by",
            "left_timestamp_ms", "left_time", "left_reason",
            "past_state", "past_state_label",
            "past_last_seen_timestamp_ms", "past_last_seen_time",
            "first_message_id", "first_source_msg_id",
            "first_message_timestamp_ms", "first_message_time",
            "first_message_type", "first_message_text",
            "last_message_id", "last_source_msg_id",
            "last_message_timestamp_ms", "last_message_time",
            "last_message_type", "last_message_text",
            "total_messages", "sent_from_device_owner_count",
            "received_from_participant_count", "chat_messages_excluding_system",
            "text_count", "image_count", "video_count", "audio_count",
            "document_count", "sticker_count", "gif_count", "location_count",
            "contact_card_count", "poll_count", "call_count", "system_count",
            "other_count", "media_count", "link_count", "forwarded_count",
            "starred_count", "edited_count", "revoked_count", "ephemeral_count",
            "reaction_given_count", "reaction_received_count",
            "mention_made_count", "mention_received_count",
            "membership_record_count", "membership_record_ids",
            "all_join_times", "all_left_times", "all_sources",
        ]

        group = db.fetchone(
            "SELECT display_name, jid_raw_string FROM conversation WHERE id = ?",
            (conv_id,),
        )
        group_name = (group["display_name"] if group else None) or self._group_name or ""
        group_jid = (group["jid_raw_string"] if group else None) or ""

        owner_cid = self._resolve_device_owner_cid(db)
        owner_info = self._detect_owner_membership(db, conv_id, owner_cid)

        records: dict[int, dict] = {}
        contact_ids: set[int] = set()

        def rec_for(contact_id: int) -> dict:
            if contact_id not in records:
                records[contact_id] = {h: "" for h in headers}
                records[contact_id].update({
                    "conversation_id": conv_id,
                    "group_name": group_name,
                    "group_jid": group_jid,
                    "contact_id": contact_id,
                    "is_device_owner": self._yn(owner_cid and contact_id == owner_cid),
                    "device_owner_group_status": owner_info.get("status", ""),
                    "has_group_member_record": "No",
                    "is_current_member": "No",
                    "has_past_participant_db_record": "No",
                    "found_by_messages": "No",
                    "message_only_participant": "No",
                    "is_non_roster_sender": "No",
                    "is_phone_resolved": "No",
                    "is_saved_in_address_book": "No",
                    "is_business": "No",
                    "is_meta_verified": "No",
                    "_sources": set(),
                    "_membership_ids": [],
                    "_join_times": [],
                    "_left_times": [],
                    "_join_ts": [],
                    "_left_ts": [],
                    "_roles": [],
                    "_labels": [],
                    "_join_methods": [],
                    "_added_by": [],
                    "_past_states": [],
                    "_past_seen_ts": [],
                })
            contact_ids.add(contact_id)
            return records[contact_id]

        added_by_map = self._group_added_by_map(db, conv_id)

        for row in db.fetchall("""
            SELECT gm.id, gm.contact_id, gm.role, gm.label,
                   gm.join_timestamp, gm.join_method, gm.is_current,
                   gm.left_timestamp, gm.left_reason
            FROM group_member gm
            WHERE gm.conversation_id = ?
            ORDER BY gm.is_current DESC, gm.join_timestamp ASC
        """, (conv_id,)):
            d = dict(row)
            cid = d.get("contact_id")
            if not cid:
                continue
            rec = rec_for(cid)
            rec["has_group_member_record"] = "Yes"
            rec["_sources"].add("group_member_current" if d.get("is_current") else "group_member_past")
            if d.get("is_current"):
                rec["is_current_member"] = "Yes"
            rec["_membership_ids"].append(str(d.get("id") or ""))
            if d.get("role"):
                rec["_roles"].append(str(d["role"]))
            if d.get("label"):
                rec["_labels"].append(str(d["label"]))
            if d.get("join_timestamp"):
                rec["_join_ts"].append(d["join_timestamp"])
                rec["_join_times"].append(self._csv_ts(d["join_timestamp"]))
            if d.get("left_timestamp"):
                rec["_left_ts"].append(d["left_timestamp"])
                rec["_left_times"].append(self._csv_ts(d["left_timestamp"]))
            if d.get("left_reason"):
                rec["left_reason"] = self._merge_text(rec["left_reason"], d["left_reason"])
            if d.get("join_method") is not None:
                method_label = self._format_join_method(d.get("join_method"))
                rec["_join_methods"].append(f"{d.get('join_method')}:{method_label}")
            if added_by_map.get(cid):
                rec["_added_by"].append(added_by_map[cid])

        if self._table_exists(db, "group_past_participant"):
            for row in db.fetchall("""
                SELECT gpp.contact_id, gpp.state, gpp.last_seen_ts
                FROM group_past_participant gpp
                WHERE gpp.conversation_id = ?
                ORDER BY gpp.last_seen_ts DESC
            """, (conv_id,)):
                d = dict(row)
                cid = d.get("contact_id")
                if not cid:
                    continue
                rec = rec_for(cid)
                rec["has_past_participant_db_record"] = "Yes"
                rec["_sources"].add("group_past_participant_db")
                if d.get("state") is not None:
                    rec["_past_states"].append(str(d["state"]))
                if d.get("last_seen_ts"):
                    rec["_past_seen_ts"].append(d["last_seen_ts"])

        # Owner-content messages in groups land in msgstore.db with
        # sender_jid_row_id = 0 (WhatsApp implicit owner identity).  Our
        # ingestion leaves analysis.message.sender_id NULL for those, so
        # the per-sender counts query needs an owner-fallback else the
        # owner row reports 0 text/0 image/0 voice while the real number
        # is dozens.  Pass owner_cid so the helpers can attribute
        # from_me=1 + sender_id IS NULL messages to the owner.
        message_stats = self._participant_message_stats(db, conv_id, owner_cid)
        for cid, stats in message_stats.items():
            rec = rec_for(cid)
            rec["found_by_messages"] = "Yes"
            rec["_sources"].add("messages")
            for key, val in stats.items():
                if key in rec:
                    rec[key] = val

        endpoint_rows = self._participant_message_endpoints(db, conv_id, owner_cid)
        for cid, endpoints in endpoint_rows.items():
            rec = rec_for(cid)
            for key, val in endpoints.items():
                rec[key] = val

        # IMPORTANT: link / reaction / mention counts attach to
        # EXISTING records only.  ``rec_for(cid)`` silently
        # creates a record for any contact_id that hasn't already
        # been added by the group_member / group_past_participant
        # / messages loops above, which would produce empty
        # "ghost rows" in the export for contacts that were only
        # ever *mentioned* or *reacted-to* but never actually
        # participated in the group.
        #
        # The rule: a contact counts as a participant only if
        # they have a group_member row, a group_past_participant
        # row, OR sent at least one message in this group.
        # Mentions / reactions *received* are not participation —
        # they are just content.
        link_counts = self._participant_link_counts(db, conv_id, owner_cid)
        for cid, cnt in link_counts.items():
            if cid in records:
                records[cid]["link_count"] = cnt

        reaction_counts = self._participant_reaction_counts(db, conv_id)
        for cid, counts in reaction_counts.items():
            if cid in records:
                records[cid]["reaction_given_count"] = counts.get("given", 0)
                records[cid]["reaction_received_count"] = counts.get("received", 0)

        mention_counts = self._participant_mention_counts(db, conv_id)
        for cid, counts in mention_counts.items():
            if cid in records:
                records[cid]["mention_made_count"] = counts.get("made", 0)
                records[cid]["mention_received_count"] = counts.get("received", 0)

        contact_rows = self._contacts_for_export(db, sorted(contact_ids))
        for cid, c in contact_rows.items():
            rec = rec_for(cid)
            display_name = self._export_display_name(c)
            rec.update({
                "display_name": display_name,
                "_avatar_blob": c.get("avatar_blob"),
                "resolved_name": c.get("resolved_name") or "",
                "address_book_name": c.get("display_name") or "",
                "whatsapp_name": c.get("wa_name") or "",
                "phone_number": c.get("phone_number") or "",
                "phone_jid": c.get("phone_jid") or "",
                "lid_jid": c.get("lid_jid") or "",
                "lid_masked_phone": c.get("lid_masked_phone") or "",
                "is_phone_resolved": self._yn(bool(c.get("phone_number") or c.get("phone_jid"))),
                "is_saved_in_address_book": self._yn(c.get("is_saved")),
                "is_business": self._yn(c.get("is_business")),
                "is_meta_verified": self._yn(c.get("is_meta_verified")),
                "platform_estimate": c.get("platform_estimate") or "",
                "platform_confidence": c.get("platform_confidence") or "",
                "linked_device_count": c.get("linked_device_count") or 0,
                "status_text": c.get("status_text") or "",
                "business_name": c.get("business_name") or "",
                "company": c.get("company") or "",
                "title": c.get("title") or "",
            })

        for rec in records.values():
            rec["membership_record_count"] = len([x for x in rec["_membership_ids"] if x])
            rec["membership_record_ids"] = "; ".join([x for x in rec["_membership_ids"] if x])
            rec["all_join_times"] = "; ".join(dict.fromkeys(rec["_join_times"]))
            rec["all_left_times"] = "; ".join(dict.fromkeys(rec["_left_times"]))
            rec["all_sources"] = "; ".join(sorted(rec["_sources"]))
            rec["role"] = self._best_role(rec["_roles"])
            rec["label"] = "; ".join(dict.fromkeys(rec["_labels"]))
            rec["group_specific_name"] = rec["label"]
            rec["join_timestamp_ms"] = min(rec["_join_ts"]) if rec["_join_ts"] else ""
            rec["join_time"] = self._csv_ts(rec["join_timestamp_ms"])
            rec["join_method_code"] = "; ".join(dict.fromkeys(
                jm.split(":", 1)[0] for jm in rec["_join_methods"] if jm
            ))
            rec["join_method_label"] = "; ".join(dict.fromkeys(
                jm.split(":", 1)[1] for jm in rec["_join_methods"] if ":" in jm
            ))
            rec["added_by"] = "; ".join(dict.fromkeys(rec["_added_by"]))
            rec["left_timestamp_ms"] = max(rec["_left_ts"]) if rec["_left_ts"] else ""
            rec["left_time"] = self._csv_ts(rec["left_timestamp_ms"])
            rec["past_state"] = "; ".join(dict.fromkeys(rec["_past_states"]))
            rec["past_state_label"] = "; ".join(
                dict.fromkeys(self._past_state_label(s) for s in rec["_past_states"])
            )
            rec["past_last_seen_timestamp_ms"] = max(rec["_past_seen_ts"]) if rec["_past_seen_ts"] else ""
            rec["past_last_seen_time"] = self._csv_ts(rec["past_last_seen_timestamp_ms"])
            rec["participant_source"] = self._participant_source_label(rec)
            rec["participant_origin_category"] = self._participant_origin_category(rec)
            rec["db_past_participant_currently_former"] = self._yn(
                rec["has_past_participant_db_record"] == "Yes"
                and rec["is_current_member"] != "Yes"
            )
            rec["message_only_participant"] = self._yn(
                rec["found_by_messages"] == "Yes"
                and rec["has_group_member_record"] == "No"
                and rec["has_past_participant_db_record"] == "No"
            )
            # CRITICAL EXCEPTION: the device owner is NEVER in
            # group_member — WhatsApp's group_participants table
            # only tracks OTHER participants from the owner's
            # perspective.  Owner membership lives in
            # ``conversation.participation_status``, which
            # ``_detect_owner_membership()`` already evaluated
            # above.  Without this override the owner would be
            # falsely flagged as a non-roster message sender.
            if rec.get("is_device_owner") == "Yes":
                # Owner: not a non-roster sender, ever.  Membership flows
                # from participation_status (the source we already
                # captured in device_owner_status_*).
                rec["message_only_participant"] = "No"
                rec["is_non_roster_sender"] = "No"
                rec["non_roster_sender_type"] = ""
                rec["non_roster_sender_note"] = ""
                owner_status = (rec.get("device_owner_group_status") or "").lower()
                if owner_status == "current":
                    rec["is_current_member"] = "Yes"
                    rec["has_group_member_record"] = "No (owner: implicit by WhatsApp design)"
                    rec["logic_added_reason"] = (
                        "Device owner - implicit member of every group they belong to. "
                        "WhatsApp's group_participants table excludes the owner by design; "
                        "owner membership comes from conversation.participation_status."
                    )
                elif owner_status in ("past", "left", "removed"):
                    rec["logic_added_reason"] = (
                        "Device owner left this group. participation_status indicates "
                        f"'{owner_status}'; messages remain from the period they were in."
                    )
                else:
                    rec["logic_added_reason"] = (
                        "Device owner of the case. WhatsApp's group_participants table "
                        "never lists the owner; status comes from participation_status."
                    )
            elif rec["message_only_participant"] == "Yes":
                rec["is_non_roster_sender"] = "Yes"
                rec["non_roster_sender_type"] = self._non_roster_sender_type(rec)
                rec["non_roster_sender_note"] = (
                    "This identity sent messages in the group but is absent from "
                    "group_member and group_past_participant tables. It can be a "
                    "bot/service identity such as Meta AI, not necessarily a past member."
                )
                rec["logic_added_reason"] = "Sender has messages in this group but no group_member or group_past_participant row"
            elif rec["has_past_participant_db_record"] == "Yes" and rec["is_current_member"] == "Yes":
                rec["logic_added_reason"] = "DB has past-participant history, but current member row takes precedence"
            rec["media_count"] = sum(
                int(rec.get(k) or 0) for k in (
                    "image_count", "video_count", "audio_count", "document_count",
                    "sticker_count", "gif_count",
                )
            )
            rec["other_count"] = max(
                0,
                int(rec.get("total_messages") or 0)
                - sum(int(rec.get(k) or 0) for k in (
                    "text_count", "image_count", "video_count", "audio_count",
                    "document_count", "sticker_count", "gif_count", "location_count",
                    "contact_card_count", "poll_count", "call_count", "system_count",
                )),
            )
            for private_key in list(rec.keys()):
                if private_key.startswith("_") and private_key != "_avatar_blob":
                    del rec[private_key]

        # Owner-priority rule: if the device owner is currently a member
        # of this group, the group_member table is the AUTHORITATIVE
        # source of truth (WhatsApp keeps it accurate while the owner is
        # in the group).  If the owner has LEFT, group_member loses
        # rows over time and message-presence becomes the most reliable
        # signal.  Tag every record with which authority applies so the
        # downstream report reader knows what to trust.
        owner_currently_member = (owner_info.get("status") or "").lower() == "current"
        authority = ("group_member_table" if owner_currently_member
                     else "messages_flow_fallback")
        for rec in records.values():
            rec["membership_authority"] = authority
            # When owner has left, prefer message-flow membership for
            # display: the participant_source label is rewritten so
            # historical message senders are surfaced as "active in
            # messages" even if their group_member row dropped off.
            if not owner_currently_member and rec.get("found_by_messages") == "Yes":
                if rec.get("has_group_member_record") == "No":
                    rec["logic_added_reason"] = (
                        rec.get("logic_added_reason") or
                        "Owner is NOT a current member - membership inferred "
                        "from message senders (group_member table is no longer "
                        "authoritative once the owner leaves)."
                    )

        # FINAL PRUNE: drop "ghost rows" - records that exist only
        # because a contact was mentioned / reacted-to / linked-to but
        # never actually participated.  A real participant has at least
        # ONE of: group_member row, past_participant row, sent a message,
        # OR is the device owner (always shown).
        def _is_real_participant(rec: dict) -> bool:
            return (
                rec.get("is_device_owner") == "Yes"
                or rec.get("has_group_member_record") == "Yes"
                or rec.get("has_past_participant_db_record") == "Yes"
                or rec.get("found_by_messages") == "Yes"
            )
        records = {cid: r for cid, r in records.items() if _is_real_participant(r)}

        # Add the new column to headers so it shows up in the export.
        if "membership_authority" not in headers:
            headers.append("membership_authority")

        rows = sorted(
            records.values(),
            key=lambda r: (
                0 if r["is_device_owner"] == "Yes" else 1,
                0 if r["is_current_member"] == "Yes" else 1,
                -int(r.get("total_messages") or 0),
                str(r.get("display_name") or "").lower(),
            ),
        )
        return headers, rows

    @staticmethod
    def _yn(value) -> str:
        return "Yes" if bool(value) else "No"

    @staticmethod
    def _merge_text(existing, value) -> str:
        if not value:
            return existing or ""
        if not existing:
            return str(value)
        parts = [p.strip() for p in str(existing).split(";") if p.strip()]
        if str(value) not in parts:
            parts.append(str(value))
        return "; ".join(parts)

    @staticmethod
    def _best_role(roles: list[str]) -> str:
        if not roles:
            return ""
        priority = {"superadmin": 0, "admin": 1, "member": 2}
        return sorted(dict.fromkeys(roles), key=lambda r: priority.get(str(r).lower(), 9))[0]

    @staticmethod
    def _table_exists(db, table_name: str) -> bool:
        try:
            row = db.fetchone(
                "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
                (table_name,),
            )
            return bool(row)
        except Exception:
            return False

    @staticmethod
    def _csv_ts(ts_ms) -> str:
        if not ts_ms:
            return ""
        try:
            return _fmt_ts(int(ts_ms), "forensic_tz")
        except Exception:
            return ""

    @staticmethod
    def _export_display_name(c: dict) -> str:
        if c.get("is_saved") and c.get("display_name"):
            return c["display_name"]
        return (
            c.get("business_name")
            or c.get("resolved_name")
            or (f"~{c.get('wa_name')}" if c.get("wa_name") else "")
            or c.get("phone_number")
            or c.get("phone_jid")
            or c.get("lid_jid")
            or f"cid#{c.get('id')}"
        )

    @staticmethod
    def _past_state_label(state) -> str:
        if state in ("", None):
            return ""
        try:
            state_int = int(state)
        except (TypeError, ValueError):
            return str(state)
        labels = {
            -2: "message-only sender",
            -1: "left member record",
            0: "current/member",
        }
        return labels.get(state_int, f"past participant state {state_int}")

    @staticmethod
    def _participant_source_label(rec: dict) -> str:
        # Device owner gets a label that names the actual evidence
        # (participation_status), not "inferred from messages".
        if rec.get("is_device_owner") == "Yes":
            owner_status = (rec.get("device_owner_group_status") or "").lower()
            if owner_status == "current":
                return "device owner - current member (via participation_status)"
            if owner_status in ("past", "left", "removed"):
                return f"device owner - {owner_status} (via participation_status)"
            return "device owner"
        if rec["is_current_member"] == "Yes":
            if rec["has_past_participant_db_record"] == "Yes":
                return "current member (DB also has past-participant history)"
            return "current member"
        if rec["has_past_participant_db_record"] == "Yes":
            return "past participant from DB"
        if rec["has_group_member_record"] == "Yes":
            return "past participant from group_member"
        if rec["found_by_messages"] == "Yes":
            return "participant inferred from messages"
        return "unknown"

    @staticmethod
    def _participant_origin_category(rec: dict) -> str:
        # Owner is in its own bucket - they're never in group_member,
        # never "inferred from messages".  Their authority is
        # conversation.participation_status.
        if rec.get("is_device_owner") == "Yes":
            owner_status = (rec.get("device_owner_group_status") or "").lower()
            if owner_status == "current":
                return "DEVICE_OWNER_CURRENT"
            if owner_status in ("past", "left", "removed"):
                return f"DEVICE_OWNER_{owner_status.upper()}"
            return "DEVICE_OWNER"
        if rec["is_current_member"] == "Yes":
            return "CURRENT"
        if rec["has_past_participant_db_record"] == "Yes":
            return "PAST_DB"
        if rec["has_group_member_record"] == "Yes":
            return "PAST_GROUP_MEMBER"
        if rec["found_by_messages"] == "Yes":
            return "ADDED_BY_MESSAGE_LOGIC"
        return "UNKNOWN"

    @staticmethod
    def _non_roster_sender_type(rec: dict) -> str:
        haystack = " ".join(str(rec.get(k) or "") for k in (
            "display_name", "resolved_name", "address_book_name",
            "whatsapp_name", "phone_jid", "lid_jid",
        )).lower()
        if "meta ai" in haystack or "meta.ai" in haystack:
            return "META_AI_OR_BOT"
        if "bot" in haystack or "ai" in haystack:
            return "POSSIBLE_BOT_OR_SERVICE"
        return "NON_ROSTER_MESSAGE_SENDER"

    def _group_added_by_map(self, db, conv_id: int) -> dict[int, str]:
        try:
            rows = db.fetchall("""
                SELECT se.target_id,
                       COALESCE(ac.resolved_name, ac.wa_name, ac.phone_number, ac.phone_jid) AS actor_name
                FROM system_event se
                LEFT JOIN contact ac ON ac.id = se.actor_id
                WHERE se.conversation_id = ?
                  AND se.target_id IS NOT NULL
                  AND se.actor_id IS NOT NULL
                  AND se.actor_id != se.target_id
                  AND se.event_label IN ('participant_joined', 'participant_added', 'you_were_added')
                ORDER BY se.timestamp ASC
            """, (conv_id,))
        except Exception:
            return {}
        out: dict[int, str] = {}
        for r in rows:
            if r["target_id"] and r["actor_name"]:
                out[r["target_id"]] = r["actor_name"]
        return out

    def _participant_message_stats(self, db, conv_id: int,
                                    owner_cid: int | None = None) -> dict[int, dict]:
        # WhatsApp groups: owner-content messages have from_me=1 and
        # sender_jid_row_id=0 in msgstore.db (the owner is implicit).
        # Our ingestion leaves analysis.message.sender_id NULL for those.
        # Without the owner-fallback below, the owner row would report
        # zero text/image/voice/etc. even though the real number is the
        # entire from_me=1 stream.  CASE expr attributes those rows to
        # owner_cid so they group correctly.
        rows = db.fetchall("""
            SELECT CASE WHEN m.from_me = 1 AND m.sender_id IS NULL THEN ?
                        ELSE m.sender_id END AS effective_sender_id,
                   COUNT(*) AS total_messages,
                   SUM(CASE WHEN m.from_me = 1 THEN 1 ELSE 0 END) AS sent_from_device_owner_count,
                   SUM(CASE WHEN m.from_me = 0 THEN 1 ELSE 0 END) AS received_from_participant_count,
                   SUM(CASE WHEN m.message_type != 7 THEN 1 ELSE 0 END) AS chat_messages_excluding_system,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) = 'text' OR m.message_type = 0 THEN 1 ELSE 0 END) AS text_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) = 'image' OR m.message_type = 1 THEN 1 ELSE 0 END) AS image_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) IN ('audio','voice','voice_note') OR m.message_type = 2 THEN 1 ELSE 0 END) AS audio_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) = 'video' OR m.message_type = 3 THEN 1 ELSE 0 END) AS video_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) = 'document' OR m.message_type = 9 THEN 1 ELSE 0 END) AS document_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) = 'sticker' THEN 1 ELSE 0 END) AS sticker_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) = 'gif' THEN 1 ELSE 0 END) AS gif_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) IN ('location','live_location') THEN 1 ELSE 0 END) AS location_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) IN ('contact_card','vcard') THEN 1 ELSE 0 END) AS contact_card_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) = 'poll' THEN 1 ELSE 0 END) AS poll_count,
                   SUM(CASE WHEN LOWER(COALESCE(m.type_label,'')) = 'call' THEN 1 ELSE 0 END) AS call_count,
                   SUM(CASE WHEN m.message_type = 7 OR LOWER(COALESCE(m.type_label,'')) LIKE '%system%' THEN 1 ELSE 0 END) AS system_count,
                   SUM(CASE WHEN m.is_forwarded = 1 THEN 1 ELSE 0 END) AS forwarded_count,
                   SUM(CASE WHEN m.is_starred = 1 THEN 1 ELSE 0 END) AS starred_count,
                   SUM(CASE WHEN m.is_edited = 1 THEN 1 ELSE 0 END) AS edited_count,
                   SUM(CASE WHEN m.is_revoked = 1 THEN 1 ELSE 0 END) AS revoked_count,
                   SUM(CASE WHEN m.is_ephemeral = 1 THEN 1 ELSE 0 END) AS ephemeral_count
            FROM message m
            WHERE m.conversation_id = ?
              AND (m.sender_id IS NOT NULL OR m.from_me = 1)
            GROUP BY effective_sender_id
        """, (owner_cid, conv_id))
        return {
            r["effective_sender_id"]: {k: (r[k] or 0) for k in r.keys() if k != "effective_sender_id"}
            for r in rows if r["effective_sender_id"]
        }

    def _participant_message_endpoints(self, db, conv_id: int,
                                        owner_cid: int | None = None) -> dict[int, dict]:
        # CASE attributes from_me=1 + sender_id IS NULL rows to owner so
        # the owner's first/last message timestamps cover their actual
        # content (not just the rare system event that resolved to them).
        try:
            rows = db.fetchall("""
                SELECT effective_sender_id AS sender_id, id, source_msg_id,
                       timestamp, type_label,
                       COALESCE(text_content, '') AS text_content,
                       first_rank, last_rank
                FROM (
                    SELECT
                        CASE WHEN m.from_me = 1 AND m.sender_id IS NULL THEN ?
                             ELSE m.sender_id END AS effective_sender_id,
                        m.id, m.source_msg_id, m.timestamp,
                        m.type_label, m.text_content,
                        ROW_NUMBER() OVER (
                            PARTITION BY CASE WHEN m.from_me = 1 AND m.sender_id IS NULL THEN ?
                                              ELSE m.sender_id END
                            ORDER BY m.timestamp ASC, m.id ASC
                        ) AS first_rank,
                        ROW_NUMBER() OVER (
                            PARTITION BY CASE WHEN m.from_me = 1 AND m.sender_id IS NULL THEN ?
                                              ELSE m.sender_id END
                            ORDER BY m.timestamp DESC, m.id DESC
                        ) AS last_rank
                    FROM message m
                    WHERE m.conversation_id = ?
                      AND (m.sender_id IS NOT NULL OR m.from_me = 1)
                )
                WHERE first_rank = 1 OR last_rank = 1
            """, (owner_cid, owner_cid, owner_cid, conv_id))
        except Exception:
            return {}
        out: dict[int, dict] = {}
        for r in rows:
            cid = r["sender_id"]
            if not cid:
                continue
            out.setdefault(cid, {})
            prefixes = []
            if r["first_rank"] == 1:
                prefixes.append("first")
            if r["last_rank"] == 1:
                prefixes.append("last")
            for prefix in prefixes:
                out[cid][f"{prefix}_message_id"] = r["id"] or ""
                out[cid][f"{prefix}_source_msg_id"] = r["source_msg_id"] or ""
                out[cid][f"{prefix}_message_timestamp_ms"] = r["timestamp"] or ""
                out[cid][f"{prefix}_message_time"] = self._csv_ts(r["timestamp"])
                out[cid][f"{prefix}_message_type"] = r["type_label"] or ""
                out[cid][f"{prefix}_message_text"] = r["text_content"] or ""
        return out

    def _participant_link_counts(self, db, conv_id: int,
                                  owner_cid: int | None = None) -> dict[int, int]:
        if not self._table_exists(db, "message_link_detail"):
            return {}
        # Same owner-fallback story: owner-sent links live on messages
        # whose sender_id is NULL (from_me=1 in groups).
        rows = db.fetchall("""
            SELECT CASE WHEN m.from_me = 1 AND m.sender_id IS NULL THEN ?
                        ELSE m.sender_id END AS effective_sender_id,
                   COUNT(*) AS cnt
            FROM message_link_detail mld
            JOIN message m ON m.id = mld.message_id
            WHERE m.conversation_id = ?
              AND (m.sender_id IS NOT NULL OR m.from_me = 1)
            GROUP BY effective_sender_id
        """, (owner_cid, conv_id))
        return {r["effective_sender_id"]: r["cnt"] or 0 for r in rows if r["effective_sender_id"]}

    def _participant_reaction_counts(self, db, conv_id: int) -> dict[int, dict]:
        if not self._table_exists(db, "reaction"):
            return {}
        out: dict[int, dict] = {}
        for r in db.fetchall("""
            SELECT reactor_id AS contact_id, COUNT(*) AS cnt
            FROM reaction
            WHERE conversation_id = ? AND reactor_id IS NOT NULL
            GROUP BY reactor_id
        """, (conv_id,)):
            out.setdefault(r["contact_id"], {})["given"] = r["cnt"] or 0
        for r in db.fetchall("""
            SELECT m.sender_id AS contact_id, COUNT(*) AS cnt
            FROM reaction rx
            JOIN message m ON m.id = rx.message_id
            WHERE rx.conversation_id = ? AND m.sender_id IS NOT NULL
            GROUP BY m.sender_id
        """, (conv_id,)):
            out.setdefault(r["contact_id"], {})["received"] = r["cnt"] or 0
        return out

    def _participant_mention_counts(self, db, conv_id: int) -> dict[int, dict]:
        if not self._table_exists(db, "mention"):
            return {}
        out: dict[int, dict] = {}
        for r in db.fetchall("""
            SELECT m.sender_id AS contact_id, COUNT(*) AS cnt
            FROM mention mn
            JOIN message m ON m.id = mn.message_id
            WHERE m.conversation_id = ? AND m.sender_id IS NOT NULL
            GROUP BY m.sender_id
        """, (conv_id,)):
            out.setdefault(r["contact_id"], {})["made"] = r["cnt"] or 0
        for r in db.fetchall("""
            SELECT mn.mentioned_id AS contact_id, COUNT(*) AS cnt
            FROM mention mn
            JOIN message m ON m.id = mn.message_id
            WHERE m.conversation_id = ? AND mn.mentioned_id IS NOT NULL
            GROUP BY mn.mentioned_id
        """, (conv_id,)):
            out.setdefault(r["contact_id"], {})["received"] = r["cnt"] or 0
        return out

    def _contacts_for_export(self, db, contact_ids: list[int]) -> dict[int, dict]:
        if not contact_ids:
            return {}
        out: dict[int, dict] = {}
        for start in range(0, len(contact_ids), 800):
            chunk = contact_ids[start:start + 800]
            placeholders = ",".join("?" for _ in chunk)
            rows = db.fetchall(f"""
                SELECT id, phone_jid, lid_jid, phone_number, display_name,
                       wa_name, given_name, family_name, nickname,
                       lid_display_name, lid_username, resolved_name,
                       company, title, status_text, is_saved, is_business,
                       business_name, is_meta_verified, lid_masked_phone,
                       platform_estimate, platform_confidence,
                       linked_device_count, avatar_blob
                FROM contact
                WHERE id IN ({placeholders})
            """, tuple(chunk))
            for r in rows:
                out[r["id"]] = dict(r)
        return out

    def _export_chat_html(self) -> None:
        """Export JUST this conversation as a V2 viewer bundle (standalone ZIP)."""
        if not self._conv_id:
            return

        import importlib.util
        from pathlib import Path as _P
        from PySide6.QtWidgets import QFileDialog, QMessageBox, QApplication

        # Locate the viewer bundle exporter module (same trick as export dialog)
        here = _P(__file__).resolve()
        exporter_cls = None
        for parent in here.parents:
            candidate = parent / "backend" / "app" / "export" / "viewer_bundle_exporter.py"
            if candidate.is_file():
                spec = importlib.util.spec_from_file_location(
                    "wainsight_viewer_bundle_exporter", str(candidate))
                mod = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(mod)
                exporter_cls = mod.ViewerBundleExporter
                break
        if not exporter_cls:
            QMessageBox.warning(self, "Export failed",
                                "Could not locate viewer_bundle_exporter.py")
            return

        db = Database.get()
        db_path = db.path
        default_dir = str(db_path.parent / "exports")
        _P(default_dir).mkdir(parents=True, exist_ok=True)

        output_dir = QFileDialog.getExistingDirectory(
            self, "Select output folder for chat HTML export", default_dir,
        )
        if not output_dir:
            return

        # Pull case metadata for forensic provenance
        case_info = {"analysis_db": str(db_path)}
        try:
            for k, v in db.fetchall(
                "SELECT key, value FROM case_metadata WHERE key IN "
                "('case_id','examiner','notes','analysis_db_sha256',"
                " 'source_msgstore_sha256','source_msgstore_path')"
            ):
                case_info[k] = v
        except Exception:
            pass

        self._export_btn.setText("Exporting...")
        self._export_btn.setEnabled(False)
        QApplication.processEvents()

        worker = exporter_cls(
            conversation_ids=[self._conv_id],
            db_path=str(db_path),
            output_dir=output_dir,
            include_media=True,
            make_zip=True,
            title=self._group_name or f"Conv {self._conv_id}",
            case_info=case_info,
        )

        def _on_finished(path: str, err: str):
            self._export_btn.setText("\u21E9 Export Chat HTML")
            self._export_btn.setEnabled(True)
            if err or not path:
                QMessageBox.warning(self, "Export failed", err or "Unknown error")
                return
            QMessageBox.information(
                self, "Export complete",
                f"Exported '{self._group_name}' to:\n\n{path}\n\n"
                "Unzip and double-click index.html to browse."
            )
            from PySide6.QtCore import QUrl
            from PySide6.QtGui import QDesktopServices
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(_P(path).parent)))

        worker.finished.connect(_on_finished)
        # Hold a strong ref so the QThread doesn't get GC'd while running
        self._chat_export_worker = worker
        worker.start()

    def _generate_group_report(self) -> None:
        """Open the customisable group-report dialog and run the
        chosen format (HTML / PDF) to the selected save location.

        Flow:
          1. Show :class:`GroupReportDialog` so the analyst can pick:
             format, date range, sections to include, top-N cap, and
             save location.
          2. Generate the HTML via ``backend.app.reports.group_report``.
          3. If PDF was requested, render the HTML through
             :class:`QWebEngineView`'s ``printToPdf`` and discard the
             intermediate HTML file.
          4. Open the resulting file in the default OS viewer.
        """
        if not self._conv_id:
            return

        import webbrowser
        import tempfile
        from pathlib import Path

        from app.views.dialogs.group_report_dialog import GroupReportDialog
        from app.views.pages._report_loader import load_group_report
        generate_group_report = load_group_report()

        db = Database.get()
        db_path = db.path
        default_dir = db_path.parent / "reports"
        try:
            default_dir.mkdir(exist_ok=True)
        except Exception:
            default_dir = db_path.parent

        group_name = self._group_name or f"Group_{self._conv_id}"
        dlg = GroupReportDialog(self, group_name, default_dir, self._conv_id)
        if dlg.exec() != dlg.DialogCode.Accepted or not dlg.is_ok:
            return

        out_path: Path = dlg.output_path
        out_format = dlg.output_format  # "html" | "pdf"
        date_from = dlg.date_from_ms
        date_to = dlg.date_to_ms
        sections = dlg.sections
        top_n = dlg.top_n

        try:
            self._report_btn.setText("Generating...")
            self._report_btn.setEnabled(False)
            from PySide6.QtWidgets import QApplication, QMessageBox
            QApplication.processEvents()

            # Always generate HTML first.  For PDF we run the HTML
            # through QWebEngineView.printToPdf() so the PDF output
            # captures the same styles, charts and embedded BLOBs.
            if out_format == "pdf":
                tmp_html = Path(tempfile.gettempdir()) / (out_path.stem + ".html")
            else:
                tmp_html = out_path

            generate_group_report(
                analysis_db_path=str(db_path),
                conversation_id=self._conv_id,
                output_path=str(tmp_html),
                date_from_ms=date_from,
                date_to_ms=date_to,
                sections=sections,
                top_n=top_n,
            )

            if out_format == "pdf":
                self._render_html_to_pdf(tmp_html, out_path)
            else:
                # HTML path is already the final output
                pass

            self._report_btn.setText("\U0001F4CB Report")
            self._report_btn.setEnabled(True)

            # Open in default OS viewer.
            try:
                webbrowser.open(out_path.as_uri())
            except Exception:
                pass

        except Exception as e:
            self._report_btn.setText("\U0001F4CB Report")
            self._report_btn.setEnabled(True)
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(
                self, "Report Generation Failed",
                f"Failed to generate group report:\n\n{e}",
            )

    def _render_html_to_pdf(self, html_path, pdf_path) -> None:
        """Render an HTML file to PDF via QWebEngineView's printToPdf.

        Runs synchronously by spinning a local QEventLoop until the
        page emits ``loadFinished`` and ``pdfPrintingFinished``.  This
        is the standard Qt pattern for "render to PDF without showing
        the view"; QPrinter/QPainter cannot render HTML directly.
        """
        from pathlib import Path
        from PySide6.QtCore import QEventLoop, QMarginsF, QUrl, QTimer
        from PySide6.QtGui import QPageLayout, QPageSize
        from PySide6.QtWebEngineWidgets import QWebEngineView

        html_path = Path(html_path)
        pdf_path = Path(pdf_path)

        view = QWebEngineView()
        # Wider off-screen viewport so the report's wide tables (current
        # members, mention network, top forwarders) lay out at the
        # right column widths before we hand off to the PDF renderer.
        view.resize(1400, 1800)

        loop = QEventLoop()

        def _on_load(ok: bool) -> None:
            if not ok:
                loop.quit()
                return
            # Give the page a moment to lay out fully before printing.
            QTimer.singleShot(400, _do_print)

        def _do_print() -> None:
            # Landscape A4 + tight margins so the 13-column Members
            # table fits without truncation.  Portrait clipped the
            # rightmost columns ("First Msg", "Last Msg") on big
            # groups; landscape gives ≈297 mm of usable width which
            # fits all columns plus the avatar comfortably.
            layout = QPageLayout(
                QPageSize(QPageSize.A4),
                QPageLayout.Landscape,
                QMarginsF(8, 10, 8, 10),
            )
            view.page().printToPdf(str(pdf_path), layout)

        def _on_pdf_done(path: str, ok: bool) -> None:
            loop.quit()

        view.loadFinished.connect(_on_load)
        view.page().pdfPrintingFinished.connect(_on_pdf_done)
        view.load(QUrl.fromLocalFile(str(html_path.resolve())))

        # Fail-safe so the loop can't hang forever if the page never
        # signals — caps at 30s for very large reports.
        QTimer.singleShot(30000, loop.quit)
        loop.exec()
        view.deleteLater()

    # ------------------------------------------------------------------ #
    # Group Edit History Dialog
    # ------------------------------------------------------------------ #

    def _show_edit_history(self) -> None:
        """Open dialog showing the complete timeline of group metadata changes."""
        if not self._conv_id:
            return

        db = Database.get()
        rows = db.fetchall(
            """
            SELECT gmc.id, gmc.change_type, gmc.old_value, gmc.new_value,
                   gmc.old_photo, gmc.new_photo,
                   gmc.changed_by_id, gmc.message_id, gmc.source_msg_id,
                   gmc.action_type, gmc.timestamp,
                   COALESCE(c.resolved_name, c.wa_name, c.display_name) AS changer_name,
                   c.phone_number AS changer_phone,
                   c.phone_jid AS changer_jid,
                   c.lid_jid AS changer_lid
            FROM group_metadata_change gmc
            LEFT JOIN contact c ON c.id = gmc.changed_by_id
            WHERE gmc.conversation_id = ?
            ORDER BY gmc.timestamp DESC
            """,
            (self._conv_id,),
        )

        if not rows:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.information(
                self, "Group Edit History",
                "No group metadata changes found for this group.\n\n"
                "This could mean:\n"
                "• The group name/description/icon was never changed\n"
                "• System events haven't been ingested yet\n"
                "• The msgstore.db doesn't contain change records for this group",
            )
            return

        # Detect device owner contact ID for badge display
        owner_cid = self._resolve_device_owner_cid(db)

        # Convert sqlite3.Row to dicts and set is_device_owner flag properly
        row_dicts = []
        for r in rows:
            d = dict(r)
            d["is_device_owner"] = (d.get("changed_by_id") == owner_cid) if owner_cid else False
            row_dicts.append(d)

        dlg = GroupEditHistoryDialog(
            row_dicts, self._group_name, self._conv_id, parent=self,
        )
        dlg.go_to_message.connect(self._on_go_to_message)
        dlg.exec()

    def _on_go_to_message(self, msg_id: int) -> None:
        """Navigate to a specific message in the chat viewer."""
        if not self._conv_id:
            return
        # Find the main window and navigate to chat viewer
        from app.views.main_window import MainWindow
        main = self.window()
        if isinstance(main, MainWindow):
            main._switch_to_conversation(self._conv_id, msg_id)

    @staticmethod
    def _resolve_device_owner_cid(db) -> int | None:
        """Resolve the device owner's contact_id using multiple fallback methods."""
        try:
            # Method 1: direct contact_id
            row = db.fetchone("SELECT value FROM case_metadata WHERE key = 'device_owner_contact_id'")
            if row and row[0]:
                return int(row[0])

            # Method 2: by JID
            row = db.fetchone("SELECT value FROM case_metadata WHERE key = 'device_owner_jid'")
            if row and row[0]:
                cr = db.fetchone("SELECT id FROM contact WHERE phone_jid = ?", (row[0],))
                if cr:
                    return cr["id"]

            # Method 3: by phone
            row = db.fetchone("SELECT value FROM case_metadata WHERE key = 'device_owner_phone'")
            if row and row[0]:
                cr = db.fetchone("SELECT id FROM contact WHERE phone_number = ?", (row[0],))
                if cr:
                    return cr["id"]

            # Method 4: from_me heuristic
            row = db.fetchone("""
                SELECT m.sender_id FROM message m
                WHERE m.from_me = 1 AND m.sender_id IS NOT NULL
                  AND m.message_type != 7
                LIMIT 1
            """)
            if row and row["sender_id"]:
                return row["sender_id"]
        except Exception:
            pass
        return None

    @staticmethod
    def _detect_owner_membership(db, conv_id: int, owner_cid: int | None) -> dict:
        """Determine whether the device owner is currently part of a group.

        Returns a dict with keys:
            status:  'creator' | 'admin' | 'member' | 'current' |
                     'former_left' | 'former_removed' | 'former' |
                     'unknown'
            detail:  human-readable explanation
            join_ts: timestamp when the owner joined (ms), or None
            leave_ts: timestamp of the latest leave/removal event (ms), or None
            source:  'participation_status' | 'system_event' | 'inferred' | 'none'

        PRIMARY source — ``msgstore.db chat.participation_status``.
        This single integer is authoritative:
            NULL = channel / broadcast / some individual chats
            0    = personal chat (owner is the account holder)
            1    = owner NO LONGER a member of the group / community
            2    = owner IS a regular member (not admin)
            3    = owner IS an appointed admin
            4    = owner IS the creator AND admin of the group/community

        FALLBACK — system_event records (when participation_status is NULL
        or the analysis.db was ingested before we started capturing the
        column). Labels we care about:
            you_were_added / participant_added (target=owner)     → joined
            participant_joined_via_link (actor=owner)             → joined
            community_or_group_created  (actor=owner)             → created
            you_were_removed                                      → removed
            you_left / participant_left (actor=owner)             → left

        LAST-RESORT — if owner has sent any non-system from_me=1 message in
        this conversation, infer CURRENT with a clear "inferred" source.
        """
        if not conv_id:
            return {"status": "unknown", "detail": "no conversation id",
                    "join_ts": None, "leave_ts": None, "source": "none"}

        # ---- PRIMARY: participation_status from msgstore.db.chat ----
        try:
            row = db.fetchone(
                "SELECT participation_status FROM conversation WHERE id = ?",
                (conv_id,),
            )
            ps = row[0] if row else None
        except Exception:
            ps = None

        if ps is not None:
            PS_MAP = {
                0: ("current",        "individual chat (owner is the account holder)"),
                1: ("former",         "owner no longer a member of this group"),
                2: ("member",         "owner is a regular member (not an admin)"),
                3: ("admin",          "owner is an appointed admin"),
                4: ("creator",        "owner is the creator and admin"),
            }
            status, detail = PS_MAP.get(ps, (
                "unknown", f"unrecognised participation_status value: {ps}"
            ))
            return {
                "status": status, "detail": detail,
                "join_ts": None, "leave_ts": None,
                "source": "participation_status",
            }

        # participation_status is NULL — fall back to system events. This
        # requires owner_cid; if we can't resolve the owner, mark unknown.
        if not owner_cid:
            return {"status": "unknown",
                    "detail": "participation_status is NULL and owner contact could not be resolved",
                    "join_ts": None, "leave_ts": None, "source": "none"}

        # Event labels the owner-flow cares about. Labels with a "you_"
        # prefix are WhatsApp's self-referential form (the action affected
        # the device owner) — these do NOT need actor/target validation
        # because the label itself identifies the owner. Labels without
        # that prefix are generic participant events; we only count them
        # when actor_id or target_id explicitly equals the owner's
        # contact_id.
        YOU_JOIN_LABELS   = {"you_were_added"}
        YOU_LEAVE_LABELS  = {"you_were_removed", "you_left"}
        SELF_JOIN_LABELS  = {"participant_joined_via_link",
                             "community_or_group_created"}
        GENERIC_JOIN_LBL  = {"participant_added"}
        GENERIC_LEAVE_LBL = {"participant_left"}

        # Fetch owner-relevant events: either the label is self-referential
        # ("you_*"), or actor/target explicitly points at the owner.
        try:
            events = db.fetchall(
                """
                SELECT event_label, actor_id, target_id, timestamp
                FROM system_event
                WHERE conversation_id = ?
                  AND event_label IS NOT NULL
                  AND (actor_id = ? OR target_id = ?
                       OR event_label LIKE 'you!_%' ESCAPE '!')
                ORDER BY timestamp ASC
                """,
                (conv_id, owner_cid, owner_cid),
            )
        except Exception:
            events = []

        join_ts = None
        leave_ts = None
        last_action = None  # 'join' | 'leave' | 'creator'
        detail = ""

        for e in events:
            label = e["event_label"]
            ts = e["timestamp"]
            actor = e["actor_id"]
            target = e["target_id"]

            # Self-referential owner events — the label says "you".
            if label in YOU_JOIN_LABELS:
                is_owner_join = True
            elif label in YOU_LEAVE_LABELS:
                is_owner_join = False
            # Generic events — need explicit actor/target match.
            elif label in SELF_JOIN_LABELS and actor == owner_cid:
                is_owner_join = True
            elif label in GENERIC_JOIN_LBL and target == owner_cid:
                is_owner_join = True
            else:
                is_owner_join = False

            if label in YOU_LEAVE_LABELS:
                is_owner_leave = True
            elif label in GENERIC_LEAVE_LBL and actor == owner_cid:
                is_owner_leave = True
            else:
                is_owner_leave = False

            if is_owner_join:
                if join_ts is None or (ts and ts < join_ts):
                    join_ts = ts
                if label == "community_or_group_created":
                    last_action = "creator"
                    detail = "created the group"
                else:
                    last_action = "join"
                    detail = {
                        "you_were_added": "added by an admin",
                        "participant_joined_via_link": "joined via invite link",
                    }.get(label, "added to the group")
            elif is_owner_leave:
                leave_ts = ts
                last_action = "leave"
                detail = (
                    "removed by an admin" if label == "you_were_removed"
                    else "left voluntarily"
                )

        # Resolve status from the last event
        if last_action == "creator":
            return {"status": "creator", "detail": detail,
                    "join_ts": join_ts, "leave_ts": None,
                    "source": "system_event"}
        if last_action == "join":
            return {"status": "current", "detail": detail,
                    "join_ts": join_ts, "leave_ts": None,
                    "source": "system_event"}
        if last_action == "leave":
            # Determine whether it was a removal or voluntary leave
            if "removed" in (detail or ""):
                return {"status": "former_removed", "detail": detail,
                        "join_ts": join_ts, "leave_ts": leave_ts,
                        "source": "system_event"}
            return {"status": "former_left", "detail": detail,
                    "join_ts": join_ts, "leave_ts": leave_ts,
                    "source": "system_event"}

        # No explicit owner events → fall back to message activity
        try:
            has_msgs = db.fetchone(
                "SELECT COUNT(*) FROM message "
                "WHERE conversation_id = ? AND from_me = 1 "
                "AND message_type != 7",
                (conv_id,),
            )
            msg_count = has_msgs[0] if has_msgs else 0
        except Exception:
            msg_count = 0

        if msg_count > 0:
            return {"status": "current",
                    "detail": f"inferred from {msg_count} message(s) sent by owner",
                    "join_ts": None, "leave_ts": None,
                    "source": "inferred"}

        return {"status": "unknown",
                "detail": "no owner-related events and no sent messages",
                "join_ts": None, "leave_ts": None,
                "source": "none"}

    @staticmethod
    def _format_join_method(method, actor_name: str = "", target_name: str = "") -> str:
        """Convert raw join_method codes into human-readable labels.
        join_method is INTEGER: 0=member, 1=added by admin, 3=invite link."""
        if method is None:
            return ""
        # Integer codes from WhatsApp source DB
        int_map = {
            0: "Member",
            1: "Added by Admin",
            3: "Invite Link",
            5: "Via Community",
        }
        if isinstance(method, int):
            return int_map.get(method, f"Method {method}")
        # Fallback for string values
        method_str = str(method).strip()
        if not method_str:
            return ""
        str_map = {
            "invite": "Invite Link",
            "invite_link": "Invite Link",
            "added": "Added by Admin",
            "added_by": "Added by Admin",
            "created": "Group Creator",
            "community": "Via Community",
            "linked_group": "Linked Group",
        }
        return str_map.get(method_str.lower(), method_str.replace("_", " ").title())

    @staticmethod
    def _format_timestamp(ts_ms) -> str:
        """Format a millisecond timestamp for display (timezone-aware)."""
        if not ts_ms:
            return ""
        try:
            return _fmt_ts(ts_ms)
        except (ValueError, OSError):
            return ""


# ====================================================================== #
# Group Edit History Dialog
# ====================================================================== #

# Human-readable labels for change types
_CHANGE_TYPE_LABELS: dict[str, str] = {
    "subject":                "Group Name Changed",
    "description":            "Description Changed",
    "icon":                   "Profile Picture Changed",
    "admin_only_edit_on":     "Admin-Only Edit: ON",
    "admin_only_edit_off":    "Admin-Only Edit: OFF",
    "admin_only_send_on":     "Admin-Only Send: ON",
    "admin_only_send_off":    "Admin-Only Send: OFF",
    "disappearing":           "Disappearing Messages Changed",
    "invite_link_reset":      "Invite Link Reset",
    "approval_mode":          "Approval Mode Changed",
    "membership_approval":    "Membership Approval Changed",
}

_CHANGE_TYPE_ICONS: dict[str, str] = {
    "subject":                "\u270E",   # ✎
    "description":            "\U0001F4DD",  # 📝
    "icon":                   "\U0001F4F7",  # 📷
    "admin_only_edit_on":     "\U0001F512",  # 🔒
    "admin_only_edit_off":    "\U0001F513",  # 🔓
    "admin_only_send_on":     "\U0001F4E2",  # 📢
    "admin_only_send_off":    "\U0001F4AC",  # 💬
    "disappearing":           "\u23F1",   # ⏱
    "invite_link_reset":      "\U0001F517",  # 🔗
    "approval_mode":          "\u2611",   # ☑
    "membership_approval":    "\u2611",   # ☑
}

_CHANGE_TYPE_COLORS: dict[str, str] = {
    "subject":                "#1565c0",
    "description":            "#00897b",
    "icon":                   "#6a1b9a",
    "admin_only_edit_on":     "#e65100",
    "admin_only_edit_off":    "#2e7d32",
    "admin_only_send_on":     "#e65100",
    "admin_only_send_off":    "#2e7d32",
    "disappearing":           "#ef6c00",
    "invite_link_reset":      "#c62828",
    "approval_mode":          "#4527a0",
    "membership_approval":    "#4527a0",
}


class GroupEditHistoryDialog(QDialog):
    """Rich dialog showing the complete timeline of group metadata changes.

    Shows:
    - Subject (name) changes with old → new values
    - Description changes with full text
    - Profile picture (DP) changes with before/after thumbnails
    - Settings changes with human-readable descriptions
    - Who made each change and when
    - "Go To Message" button to navigate to the system event in chat
    """

    go_to_message = Signal(int)  # Emits message_id

    def __init__(self, rows: list, group_name: str, conv_id: int,
                 parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Group Edit History — {group_name}")
        self.resize(900, 700)
        self.setMinimumSize(700, 400)

        tm = ThemeManager.get()
        _lt = tm.is_light
        c_bg = "#fafafa" if _lt else "#0b141a"
        c_text = "#111b21" if _lt else "#e9edef"
        c_text2 = "#667781" if _lt else "#78909c"
        c_border = "#e0e3e7" if _lt else "rgba(128,128,128,0.12)"
        c_card = "#ffffff" if _lt else "rgba(128,128,128,0.06)"
        c_accent = "#00897b" if _lt else "#00bcd4"
        c_scroll = "rgba(0,0,0,0.12)" if _lt else "rgba(128,128,128,0.18)"

        self.setStyleSheet(f"""
            QDialog {{
                background: {c_bg};
            }}
            QScrollBar:vertical {{
                background: transparent; width: 8px; border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {c_scroll}; border-radius: 4px; min-height: 30px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # ---- Header ----
        header = QFrame()
        header.setFixedHeight(56)
        header.setStyleSheet(f"""
            QFrame {{
                background: {"#f0f2f5" if _lt else "rgba(128,128,128,0.06)"};
                border-bottom: 1px solid {c_border};
            }}
        """)
        hl = QHBoxLayout(header)
        hl.setContentsMargins(20, 0, 20, 0)

        title = QLabel(f"\u270E Group Edit History — {group_name}")
        tf = QFont()
        tf.setPointSize(13)
        tf.setBold(True)
        title.setFont(tf)
        title.setStyleSheet(f"color: {c_text};")
        hl.addWidget(title, 1)

        # Summary stats
        n_subject = sum(1 for r in rows if r["change_type"] == "subject")
        n_desc = sum(1 for r in rows if r["change_type"] == "description")
        n_icon = sum(1 for r in rows if r["change_type"] == "icon")
        n_settings = len(rows) - n_subject - n_desc - n_icon

        stats_parts = []
        if n_subject:
            stats_parts.append(f"{n_subject} name")
        if n_desc:
            stats_parts.append(f"{n_desc} description")
        if n_icon:
            stats_parts.append(f"{n_icon} DP")
        if n_settings:
            stats_parts.append(f"{n_settings} settings")

        stats_lbl = QLabel(f"{len(rows)} changes: {', '.join(stats_parts)}")
        stats_lbl.setStyleSheet(f"color: {c_text2}; font-size: 11px;")
        hl.addWidget(stats_lbl)

        layout.addWidget(header)

        # ---- Filter bar ----
        filter_bar = QFrame()
        filter_bar.setFixedHeight(44)
        filter_bar.setStyleSheet(f"""
            QFrame {{
                background: {c_bg};
                border-bottom: 1px solid {c_border};
            }}
        """)
        fl = QHBoxLayout(filter_bar)
        fl.setContentsMargins(16, 6, 16, 6)
        fl.setSpacing(6)

        self._filter_btns: dict[str, QPushButton] = {}
        filters = [
            ("all", f"All ({len(rows)})"),
            ("subject", f"\u270E Name ({n_subject})"),
            ("description", f"\U0001F4DD Desc ({n_desc})"),
            ("icon", f"\U0001F4F7 DP ({n_icon})"),
            ("settings", f"\u2699 Settings ({n_settings})"),
        ]
        for fid, flabel in filters:
            btn = QPushButton(flabel)
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setFixedHeight(28)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: transparent;
                    border: 1px solid {c_border};
                    border-radius: 4px;
                    color: {c_text2};
                    font-size: 10px; font-weight: bold;
                    padding: 2px 10px;
                }}
                QPushButton:checked {{
                    background: {c_accent};
                    color: #ffffff;
                    border-color: {c_accent};
                }}
                QPushButton:hover:!checked {{
                    background: rgba(128,128,128,0.08);
                }}
            """)
            btn.clicked.connect(lambda checked, f=fid: self._apply_filter(f))
            if fid == "all":
                btn.setChecked(True)
            fl.addWidget(btn)
            self._filter_btns[fid] = btn

        fl.addStretch()
        layout.addWidget(filter_bar)

        # ---- Scrollable content ----
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        self._content = QWidget()
        self._content_layout = QVBoxLayout(self._content)
        self._content_layout.setContentsMargins(16, 12, 16, 16)
        self._content_layout.setSpacing(8)

        # Store rows and card refs
        self._rows = rows
        self._cards: list[tuple[str, QFrame]] = []  # (change_type, card_widget)
        self._c_card = c_card
        self._c_border = c_border
        self._c_text = c_text
        self._c_text2 = c_text2
        self._c_accent = c_accent
        self._is_light = _lt

        self._build_cards()

        self._content_layout.addStretch()
        scroll.setWidget(self._content)
        layout.addWidget(scroll, 1)

    def _build_cards(self) -> None:
        """Build timeline cards for each metadata change."""
        for row in self._rows:
            card = self._make_change_card(row)
            self._content_layout.addWidget(card)
            self._cards.append((row["change_type"], card))

    def _make_change_card(self, row: dict) -> QFrame:
        """Create a single change card widget."""
        change_type = row["change_type"]
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: {self._c_card};
                border: 1px solid {self._c_border};
                border-radius: 8px;
            }}
        """)

        vl = QVBoxLayout(card)
        vl.setContentsMargins(16, 12, 16, 12)
        vl.setSpacing(8)

        # ---- Top row: icon + label + timestamp + go-to button ----
        top = QHBoxLayout()
        top.setSpacing(10)

        icon = _CHANGE_TYPE_ICONS.get(change_type, "\u2699")
        color = _CHANGE_TYPE_COLORS.get(change_type, "#607d8b")
        label = _CHANGE_TYPE_LABELS.get(change_type, change_type.replace("_", " ").title())

        badge = QLabel(f'<span style="color:{color}; font-size:16px;">{icon}</span>'
                       f'&nbsp;&nbsp;<b style="color:{color};">{label}</b>')
        badge.setTextFormat(Qt.RichText)
        top.addWidget(badge)

        top.addStretch()

        # Timestamp
        ts = row["timestamp"]
        ts_str = _fmt_ts(ts) if ts else ""
        ts_lbl = QLabel(ts_str)
        ts_lbl.setStyleSheet(f"color: {self._c_text2}; font-size: 10px;")
        top.addWidget(ts_lbl)

        # Go to message button
        msg_id = row["message_id"]
        if msg_id:
            goto_btn = QPushButton("Go To \u25B6")
            goto_btn.setFixedSize(70, 24)
            goto_btn.setCursor(Qt.PointingHandCursor)
            goto_btn.setToolTip("Navigate to this system event in the chat viewer")
            goto_btn.setStyleSheet(f"""
                QPushButton {{
                    background: transparent;
                    border: 1px solid {self._c_accent};
                    border-radius: 4px;
                    color: {self._c_accent};
                    font-size: 10px; font-weight: bold;
                }}
                QPushButton:hover {{
                    background: {"rgba(0,137,123,0.1)" if self._is_light else "rgba(0,188,212,0.1)"};
                }}
            """)
            goto_btn.clicked.connect(lambda _, mid=msg_id: self._emit_goto(mid))
            top.addWidget(goto_btn)

        vl.addLayout(top)

        # ---- Changed by (full identity) ----
        changer = row.get("changer_name") or ""
        changer_phone = row.get("changer_phone") or ""
        changer_jid = row.get("changer_jid") or ""
        changer_lid = row.get("changer_lid") or ""
        is_owner = row.get("is_device_owner")

        # Build identity parts
        identity_parts = []
        if is_owner:
            identity_parts.append('<span style="color:#e65100; font-weight:bold;">You (Device Owner)</span>')
        if changer:
            identity_parts.append(f"<b>{_escape_html(changer)}</b>")
        if changer_phone:
            identity_parts.append(f"+{changer_phone}" if not changer_phone.startswith("+") else changer_phone)
        if changer_jid and "@" in changer_jid:
            identity_parts.append(f'<span style="color:{"#90a4ae" if self._is_light else "#546e7a"}; font-size:10px;">'
                                  f'JID: {_escape_html(changer_jid)}</span>')
        if changer_lid and "@" in changer_lid:
            identity_parts.append(f'<span style="color:{"#90a4ae" if self._is_light else "#546e7a"}; font-size:10px;">'
                                  f'LID: {_escape_html(changer_lid)}</span>')
        if not identity_parts:
            identity_parts.append("Unknown")

        by_lbl = QLabel(f"Changed by: {' &nbsp;·&nbsp; '.join(identity_parts)}")
        by_lbl.setTextFormat(Qt.RichText)
        by_lbl.setWordWrap(True)
        by_lbl.setStyleSheet(f"color: {self._c_text2}; font-size: 11px;")
        vl.addWidget(by_lbl)

        # ---- Change-specific content ----
        if change_type == "subject":
            self._add_subject_change(vl, row)
        elif change_type == "description":
            self._add_description_change(vl, row)
        elif change_type == "icon":
            self._add_icon_change(vl, row)
        else:
            self._add_settings_change(vl, row)

        # ---- Forensic footer ----
        source_id = row.get("source_msg_id")
        action_type = row.get("action_type")
        forensic_parts = []
        if source_id:
            forensic_parts.append(f"msgstore.message._id: {source_id}")
        if action_type is not None:
            forensic_parts.append(f"action_type: {action_type}")
        if forensic_parts:
            forensic_lbl = QLabel(" · ".join(forensic_parts))
            forensic_lbl.setStyleSheet(
                f"color: {'#b0bec5' if self._is_light else '#546e7a'}; "
                f"font-size: 9px; font-style: italic;"
            )
            vl.addWidget(forensic_lbl)

        return card

    def _add_subject_change(self, layout: QVBoxLayout, row: dict) -> None:
        """Add subject (name) change visualization."""
        old_val = row.get("old_value") or ""
        new_val = row.get("new_value") or ""

        change_frame = QFrame()
        change_frame.setStyleSheet(f"""
            QFrame {{
                background: {"#f3f8ff" if self._is_light else "rgba(21,101,192,0.08)"};
                border: 1px solid {"#bbdefb" if self._is_light else "rgba(21,101,192,0.2)"};
                border-radius: 6px;
            }}
        """)
        cl = QHBoxLayout(change_frame)
        cl.setContentsMargins(12, 8, 12, 8)
        cl.setSpacing(12)

        if old_val:
            old_lbl = QLabel(f'<span style="color:#c62828; text-decoration: line-through;">'
                             f'{_escape_html(old_val)}</span>')
            old_lbl.setTextFormat(Qt.RichText)
            old_lbl.setWordWrap(True)
            cl.addWidget(old_lbl, 1)

            arrow = QLabel("\u2192")  # →
            arrow.setStyleSheet(f"color: {self._c_text2}; font-size: 16px; font-weight: bold;")
            cl.addWidget(arrow)

        if new_val:
            new_lbl = QLabel(f'<b style="color:#1565c0;">{_escape_html(new_val)}</b>')
            new_lbl.setTextFormat(Qt.RichText)
            new_lbl.setWordWrap(True)
            cl.addWidget(new_lbl, 1)

        layout.addWidget(change_frame)

    def _add_description_change(self, layout: QVBoxLayout, row: dict) -> None:
        """Add description change visualization."""
        new_val = row.get("new_value") or ""
        if not new_val:
            lbl = QLabel("<i>Description cleared</i>")
            lbl.setTextFormat(Qt.RichText)
            lbl.setStyleSheet(f"color: {self._c_text2}; font-size: 11px;")
            layout.addWidget(lbl)
            return

        desc_frame = QFrame()
        desc_frame.setStyleSheet(f"""
            QFrame {{
                background: {"#e8f5e9" if self._is_light else "rgba(0,137,123,0.08)"};
                border: 1px solid {"#c8e6c9" if self._is_light else "rgba(0,137,123,0.2)"};
                border-radius: 6px;
            }}
        """)
        dl = QVBoxLayout(desc_frame)
        dl.setContentsMargins(12, 8, 12, 8)

        header = QLabel("New Description:")
        header.setStyleSheet(f"color: {self._c_text2}; font-size: 10px; font-weight: bold;")
        dl.addWidget(header)

        # Truncate long descriptions for display
        display_text = new_val[:500] + ("..." if len(new_val) > 500 else "")
        desc_lbl = QLabel(_escape_html(display_text))
        desc_lbl.setTextFormat(Qt.RichText)
        desc_lbl.setWordWrap(True)
        desc_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        desc_lbl.setStyleSheet(f"color: {self._c_text}; font-size: 11px;")
        dl.addWidget(desc_lbl)

        layout.addWidget(desc_frame)

    def _add_icon_change(self, layout: QVBoxLayout, row: dict) -> None:
        """Add DP/icon change visualization with before/after thumbnails."""
        old_photo = row.get("old_photo")
        new_photo = row.get("new_photo")
        new_val = row.get("new_value") or ""

        photos_frame = QFrame()
        photos_frame.setStyleSheet(f"""
            QFrame {{
                background: {"#f3e5f5" if self._is_light else "rgba(106,27,154,0.08)"};
                border: 1px solid {"#e1bee7" if self._is_light else "rgba(106,27,154,0.2)"};
                border-radius: 6px;
            }}
        """)
        pl = QHBoxLayout(photos_frame)
        pl.setContentsMargins(12, 10, 12, 10)
        pl.setSpacing(16)

        # Old photo
        if old_photo and len(old_photo) > 100:
            old_col = QVBoxLayout()
            old_col.setSpacing(4)
            old_header = QLabel("Previous DP")
            old_header.setStyleSheet(f"color: {self._c_text2}; font-size: 10px; font-weight: bold;")
            old_header.setAlignment(Qt.AlignCenter)
            old_col.addWidget(old_header)

            old_img = QLabel()
            old_img.setFixedSize(96, 96)
            old_img.setAlignment(Qt.AlignCenter)
            pxm = QPixmap()
            pxm.loadFromData(old_photo)
            if not pxm.isNull():
                scaled = pxm.scaled(96, 96, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                old_img.setPixmap(scaled)
            else:
                old_img.setText("?")
            old_img.setStyleSheet(
                f"border: 2px solid {'#ef9a9a' if self._is_light else '#ef5350'}; "
                f"border-radius: 48px; background: {'#fafafa' if self._is_light else '#1a1a2e'};"
            )
            old_col.addWidget(old_img, 0, Qt.AlignCenter)
            pl.addLayout(old_col)

            # Arrow
            arrow = QLabel("\u2192")  # →
            arrow.setStyleSheet(f"color: {self._c_text2}; font-size: 20px; font-weight: bold;")
            arrow.setAlignment(Qt.AlignCenter)
            pl.addWidget(arrow)

        # New photo
        if new_photo and len(new_photo) > 100:
            new_col = QVBoxLayout()
            new_col.setSpacing(4)
            new_header = QLabel("New DP")
            new_header.setStyleSheet(f"color: {self._c_text2}; font-size: 10px; font-weight: bold;")
            new_header.setAlignment(Qt.AlignCenter)
            new_col.addWidget(new_header)

            new_img = QLabel()
            new_img.setFixedSize(96, 96)
            new_img.setAlignment(Qt.AlignCenter)
            pxm = QPixmap()
            pxm.loadFromData(new_photo)
            if not pxm.isNull():
                scaled = pxm.scaled(96, 96, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                new_img.setPixmap(scaled)
            else:
                new_img.setText("?")
            new_img.setStyleSheet(
                f"border: 2px solid {'#a5d6a7' if self._is_light else '#66bb6a'}; "
                f"border-radius: 48px; background: {'#fafafa' if self._is_light else '#1a1a2e'};"
            )
            new_col.addWidget(new_img, 0, Qt.AlignCenter)
            pl.addLayout(new_col)
        elif new_val == "removed":
            removed_lbl = QLabel("\U0001F6AB DP Removed")
            removed_lbl.setStyleSheet(
                f"color: #c62828; font-size: 12px; font-weight: bold;"
            )
            removed_lbl.setAlignment(Qt.AlignCenter)
            pl.addWidget(removed_lbl)

        if not old_photo and not new_photo and new_val != "removed":
            no_data = QLabel("Photo data not available (BLOBs may not be stored)")
            no_data.setStyleSheet(f"color: {self._c_text2}; font-size: 11px; font-style: italic;")
            pl.addWidget(no_data)

        pl.addStretch()
        layout.addWidget(photos_frame)

    def _add_settings_change(self, layout: QVBoxLayout, row: dict) -> None:
        """Add settings change visualization."""
        change_type = row["change_type"]
        old_val = row.get("old_value") or ""
        new_val = row.get("new_value") or ""

        settings_map = {
            "admin_only_edit_on":  "Only admins can edit group settings",
            "admin_only_edit_off": "All members can edit group settings",
            "admin_only_send_on":  "Only admins can send messages",
            "admin_only_send_off": "All members can send messages",
            "invite_link_reset":   "Group invite link was reset",
            "approval_mode":       "New member approval mode changed",
            "membership_approval": "Membership approval settings changed",
            "disappearing":        "Disappearing messages duration changed",
        }

        desc = settings_map.get(change_type, change_type.replace("_", " ").title())
        if change_type == "disappearing" and new_val:
            desc = f"Disappearing messages: {new_val}"

        setting_frame = QFrame()
        setting_frame.setStyleSheet(f"""
            QFrame {{
                background: {"#fff3e0" if self._is_light else "rgba(239,108,0,0.08)"};
                border: 1px solid {"#ffe0b2" if self._is_light else "rgba(239,108,0,0.2)"};
                border-radius: 6px;
            }}
        """)
        sl = QVBoxLayout(setting_frame)
        sl.setContentsMargins(12, 8, 12, 8)

        desc_lbl = QLabel(desc)
        desc_lbl.setStyleSheet(f"color: {self._c_text}; font-size: 11px;")
        desc_lbl.setWordWrap(True)
        sl.addWidget(desc_lbl)

        if old_val and old_val not in ("invite_link", "regular"):
            old_lbl = QLabel(f"Previous: {old_val}")
            old_lbl.setStyleSheet(f"color: {self._c_text2}; font-size: 10px;")
            sl.addWidget(old_lbl)

        layout.addWidget(setting_frame)

    def _emit_goto(self, msg_id: int) -> None:
        """Emit go_to_message signal and close dialog."""
        self.go_to_message.emit(msg_id)
        self.accept()

    def _apply_filter(self, filter_id: str) -> None:
        """Filter displayed cards by change type."""
        # Update button states
        for fid, btn in self._filter_btns.items():
            btn.setChecked(fid == filter_id)

        settings_types = {
            "admin_only_edit_on", "admin_only_edit_off",
            "admin_only_send_on", "admin_only_send_off",
            "disappearing", "invite_link_reset",
            "approval_mode", "membership_approval",
        }

        for change_type, card in self._cards:
            if filter_id == "all":
                card.setVisible(True)
            elif filter_id == "settings":
                card.setVisible(change_type in settings_types)
            else:
                card.setVisible(change_type == filter_id)


# ====================================================================== #
# Member Group Profile Dialog
# ====================================================================== #

class MemberGroupProfileDialog(QDialog):
    """Full profile of a contact scoped to a specific group.

    Shows: identity (name, phone, JID, LID), role, join info,
    message stats in this group, hourly activity, top mentions given/received,
    media breakdown, reactions, and recent messages.
    """

    go_to_message = Signal(int)

    def __init__(
        self,
        contact_id: int,
        conversation_id: int,
        group_name: str,
        member_name: str,
        parent=None,
    ):
        super().__init__(parent)
        self._tm = ThemeManager.get()
        _lt = self._tm.is_light
        self._is_light = _lt

        # Theme colors
        self._c_bg = "#ffffff" if _lt else "#0b141a"
        self._c_bg2 = "#f5f6f6" if _lt else "#111b21"
        self._c_bg3 = "#edf0f2" if _lt else "#1a2630"
        self._c_text = "#111b21" if _lt else "#e9edef"
        self._c_text2 = "#667781" if _lt else "#8696a0"
        self._c_text3 = "#8696a0" if _lt else "#667781"
        self._c_accent = "#00897b" if _lt else "#00bcd4"
        self._c_accent_bg = "rgba(0,137,123,0.08)" if _lt else "rgba(0,188,212,0.08)"
        self._c_accent_border = "rgba(0,137,123,0.2)" if _lt else "rgba(0,188,212,0.2)"
        self._c_border = "#e0e3e7" if _lt else "rgba(128,128,128,0.12)"
        self._c_border2 = "#f0f2f5" if _lt else "rgba(128,128,128,0.08)"

        self.setWindowTitle(f"Member Profile — {member_name} in {group_name}")
        self.setMinimumSize(820, 650)
        self.resize(900, 720)

        self.setStyleSheet(f"QDialog {{ background: {self._c_bg}; }}")

        db = Database.get()

        # ---- Fetch all data ----
        contact = db.fetchone("""
            SELECT id, resolved_name, wa_name, display_name, phone_number,
                   phone_jid, lid_jid, lid_display_name, status_text,
                   is_business, business_name, avatar_blob, platform_estimate
            FROM contact WHERE id = ?
        """, (contact_id,))

        if not contact:
            layout = QVBoxLayout(self)
            layout.addWidget(QLabel("Contact not found."))
            return

        contact = dict(contact)

        # Group membership
        membership = db.fetchone("""
            SELECT gm.role, gm.label, gm.join_timestamp, gm.join_method
            FROM group_member gm WHERE gm.conversation_id = ? AND gm.contact_id = ?
        """, (conversation_id, contact_id))
        membership = dict(membership) if membership else {}

        # Message stats in this group
        msg_stats = db.fetchone("""
            SELECT COUNT(*) AS total,
                   SUM(CASE WHEN m.message_type = 0 THEN 1 ELSE 0 END) AS text_msgs,
                   SUM(CASE WHEN m.message_type = 1 THEN 1 ELSE 0 END) AS images,
                   SUM(CASE WHEN m.message_type = 3 THEN 1 ELSE 0 END) AS videos,
                   SUM(CASE WHEN m.message_type = 2 THEN 1 ELSE 0 END) AS audio,
                   SUM(CASE WHEN m.message_type = 9 THEN 1 ELSE 0 END) AS documents,
                   SUM(CASE WHEN m.message_type = 20 THEN 1 ELSE 0 END) AS stickers,
                   SUM(CASE WHEN m.message_type = 13 THEN 1 ELSE 0 END) AS gifs,
                   SUM(CASE WHEN m.message_type IN (5,16) THEN 1 ELSE 0 END) AS locations,
                   SUM(CASE WHEN m.message_type IN (42,43) THEN 1 ELSE 0 END) AS view_once,
                   SUM(CASE WHEN m.is_forwarded = 1 THEN 1 ELSE 0 END) AS forwards,
                   SUM(CASE WHEN m.is_revoked = 1 THEN 1 ELSE 0 END) AS deletes,
                   MIN(m.timestamp) AS first_msg_ts,
                   MAX(m.timestamp) AS last_msg_ts
            FROM message m
            WHERE m.conversation_id = ? AND m.sender_id = ? AND m.message_type != 7
        """, (conversation_id, contact_id))
        msg_stats = dict(msg_stats) if msg_stats else {}

        # Count edits from edit_version table (separate from message)
        try:
            _edit_cnt = db.scalar("""
                SELECT COUNT(DISTINCT ev.message_id) FROM edit_version ev
                JOIN message m ON m.id = ev.message_id
                WHERE m.conversation_id = ? AND m.sender_id = ?
            """, (conversation_id, contact_id))
            msg_stats["edits"] = _edit_cnt or 0
        except Exception:
            msg_stats["edits"] = 0

        # Hourly activity
        hourly = db.fetchall("""
            SELECT CAST(strftime('%H', m.timestamp/1000, 'unixepoch') AS INTEGER) AS hour,
                   COUNT(*) AS cnt
            FROM message m
            WHERE m.conversation_id = ? AND m.sender_id = ? AND m.message_type != 7
            GROUP BY hour ORDER BY hour
        """, (conversation_id, contact_id))
        hourly_map = {r["hour"]: r["cnt"] for r in hourly}

        # Links
        link_count = db.scalar("""
            SELECT COUNT(*) FROM message_link_detail mld
            JOIN message m ON m.id = mld.message_id
            WHERE m.conversation_id = ? AND m.sender_id = ?
        """, (conversation_id, contact_id)) or 0

        link_domains = db.fetchall("""
            SELECT mld.domain, COUNT(*) AS cnt
            FROM message_link_detail mld
            JOIN message m ON m.id = mld.message_id
            WHERE m.conversation_id = ? AND m.sender_id = ?
            GROUP BY mld.domain ORDER BY cnt DESC LIMIT 8
        """, (conversation_id, contact_id))

        # Mentions given (who this contact mentions in this group)
        mentions_given = db.fetchall("""
            SELECT mn.mentioned_id,
                   COALESCE(c.resolved_name, c.wa_name, c.phone_number) AS mentioned_name,
                   c.phone_jid AS mentioned_jid,
                   COUNT(*) AS cnt
            FROM mention mn
            JOIN message m ON m.id = mn.message_id
            LEFT JOIN contact c ON c.id = mn.mentioned_id
            WHERE m.sender_id = ? AND m.conversation_id = ?
              AND mn.mentioned_id IS NOT NULL
            GROUP BY mn.mentioned_id ORDER BY cnt DESC LIMIT 10
        """, (contact_id, conversation_id))

        # Mentions received (who mentions this contact in this group)
        mentions_received = db.fetchall("""
            SELECT m.sender_id,
                   COALESCE(c.resolved_name, c.wa_name, c.phone_number) AS mentioner_name,
                   c.phone_jid AS mentioner_jid,
                   COUNT(*) AS cnt
            FROM mention mn
            JOIN message m ON m.id = mn.message_id
            LEFT JOIN contact c ON c.id = m.sender_id
            WHERE mn.mentioned_id = ? AND m.conversation_id = ?
            GROUP BY m.sender_id ORDER BY cnt DESC LIMIT 10
        """, (contact_id, conversation_id))

        # Reactions given
        reactions_given = db.fetchall("""
            SELECT r.emoji, COUNT(*) AS cnt
            FROM reaction r
            JOIN message m ON m.id = r.message_id
            WHERE r.reactor_id = ? AND m.conversation_id = ?
            GROUP BY r.emoji ORDER BY cnt DESC LIMIT 8
        """, (contact_id, conversation_id))

        # Reactions received
        reactions_received = db.fetchall("""
            SELECT r.emoji, COUNT(*) AS cnt
            FROM reaction r
            JOIN message m ON m.id = r.message_id
            WHERE m.sender_id = ? AND r.reactor_id != ? AND m.conversation_id = ?
            GROUP BY r.emoji ORDER BY cnt DESC LIMIT 8
        """, (contact_id, contact_id, conversation_id))

        # Group metadata changes by this contact
        edit_count = 0
        try:
            edit_count = db.scalar("""
                SELECT COUNT(*) FROM group_metadata_change
                WHERE conversation_id = ? AND changed_by_id = ?
            """, (conversation_id, contact_id)) or 0
        except Exception:
            pass

        # ---- Build UI ----
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Title bar
        title_bar = QFrame()
        title_bar.setFixedHeight(56)
        title_bar.setStyleSheet(f"""
            QFrame {{ background: {self._c_bg2}; border-bottom: 1px solid {self._c_border}; }}
        """)
        tbl = QHBoxLayout(title_bar)
        tbl.setContentsMargins(16, 0, 16, 0)
        tbl.setSpacing(12)

        # Avatar in title
        c_name = contact.get("resolved_name") or contact.get("wa_name") or contact.get("phone_number") or "?"
        avatar_lbl = QLabel()
        avatar_lbl.setFixedSize(40, 40)
        avatar_lbl.setAlignment(Qt.AlignCenter)
        avatar_blob = contact.get("avatar_blob")
        if avatar_blob and len(avatar_blob) > 100:
            pxm = QPixmap()
            pxm.loadFromData(avatar_blob)
            if not pxm.isNull():
                scaled = pxm.scaled(40, 40, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
                avatar_lbl.setPixmap(scaled)
                avatar_lbl.setStyleSheet("border-radius: 20px;")
            else:
                avatar_lbl.setText(c_name[0].upper())
                avatar_lbl.setStyleSheet("background: #00897b; color: white; border-radius: 20px; font-weight: bold; font-size: 16px;")
        else:
            avatar_lbl.setText(c_name[0].upper())
            avatar_lbl.setStyleSheet("background: #00897b; color: white; border-radius: 20px; font-weight: bold; font-size: 16px;")
        tbl.addWidget(avatar_lbl)

        title_lbl = QLabel(f"\U0001F464 {_escape_html(c_name)} — in {_escape_html(group_name)}")
        title_lbl.setStyleSheet(f"color: {self._c_text}; font-size: 14px; font-weight: bold;")
        title_lbl.setTextFormat(Qt.RichText)
        tbl.addWidget(title_lbl, 1)

        # Report button in title
        report_btn = QPushButton("\U0001F4CB Report")
        report_btn.setFixedHeight(28)
        report_btn.setCursor(Qt.PointingHandCursor)
        report_btn.setStyleSheet(f"""
            QPushButton {{ background: {self._c_accent_bg}; border: 1px solid {self._c_accent_border};
                           border-radius: 5px; color: {self._c_accent}; font-size: 10px;
                           font-weight: bold; padding: 3px 10px; }}
            QPushButton:hover {{ background: {"rgba(0,137,123,0.15)" if _lt else "rgba(0,188,212,0.15)"}; }}
        """)
        report_btn.clicked.connect(
            lambda: self._gen_report(contact_id, member_name, conversation_id, group_name)
        )
        tbl.addWidget(report_btn)

        main_layout.addWidget(title_bar)

        # Scrollable content
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet(f"QScrollArea {{ background: {self._c_bg}; border: none; }}")

        content = QWidget()
        content.setStyleSheet(f"background: {self._c_bg};")
        cl = QVBoxLayout(content)
        cl.setContentsMargins(20, 16, 20, 20)
        cl.setSpacing(12)

        # ---- Identity Card ----
        id_frame = self._make_section("Contact Identity")
        id_layout = id_frame.layout()

        identity_pairs = [
            ("Name", contact.get("resolved_name") or contact.get("display_name") or "—"),
            ("WhatsApp Name", contact.get("wa_name") or "—"),
            ("Phone", (f"+{contact['phone_number']}" if contact.get("phone_number") and not str(contact["phone_number"]).startswith("+") else contact.get("phone_number")) or "—"),
            ("Phone JID (msgstore.db)", contact.get("phone_jid") or "—"),
            ("LID JID (msgstore.db)", contact.get("lid_jid") or "—"),
            ("Platform", contact.get("platform_estimate") or "—"),
        ]
        if contact.get("is_business"):
            identity_pairs.append(("Business", contact.get("business_name") or "Yes"))
        if contact.get("status_text"):
            identity_pairs.append(("Status", contact["status_text"]))

        for label, value in identity_pairs:
            row_w = QHBoxLayout()
            row_w.setSpacing(8)
            lbl = QLabel(label)
            lbl.setFixedWidth(160)
            lbl.setStyleSheet(f"color: {self._c_text3}; font-size: 11px; font-weight: 600;")
            row_w.addWidget(lbl)

            is_jid = "JID" in label
            val_text = f"<code style='background:{self._c_bg3};padding:1px 4px;border-radius:2px;font-size:10px;color:{self._c_accent};'>{_escape_html(str(value))}</code>" if is_jid and value != "—" else _escape_html(str(value))
            val_lbl = QLabel(val_text)
            val_lbl.setTextFormat(Qt.RichText)
            val_lbl.setStyleSheet(f"color: {self._c_text}; font-size: 11px;")
            val_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
            row_w.addWidget(val_lbl, 1)
            id_layout.addLayout(row_w)

        cl.addWidget(id_frame)

        # ---- Group Membership ----
        mem_frame = self._make_section("Group Membership")
        mem_layout = mem_frame.layout()

        role = (membership.get("role") or "member").lower()
        role_display = {"superadmin": "Super Admin (Creator)", "admin": "Admin"}.get(role, "Member")
        role_color = {"superadmin": "#f4511e", "admin": "#ffb300"}.get(role, self._c_text2)

        mem_pairs = [
            ("Role", f"<span style='color:{role_color};font-weight:bold;'>{role_display}</span>"),
        ]
        if membership.get("label"):
            mem_pairs.append(("Label", membership["label"]))
        if membership.get("join_timestamp"):
            try:
                jt = _fmt_ts(membership["join_timestamp"], "minute")
                mem_pairs.append(("Joined", jt))
            except (ValueError, OSError):
                pass
        jm = membership.get("join_method")
        if jm is not None:
            jm_map = {0: "Member", 1: "Added by Admin", 3: "Invite Link", 5: "Via Community"}
            mem_pairs.append(("Join Method", jm_map.get(jm, f"Method {jm}")))
        if edit_count:
            mem_pairs.append(("Group Changes Made", f"{edit_count} (name/DP/description/settings)"))

        for label, value in mem_pairs:
            row_w = QHBoxLayout()
            row_w.setSpacing(8)
            lbl = QLabel(label)
            lbl.setFixedWidth(160)
            lbl.setStyleSheet(f"color: {self._c_text3}; font-size: 11px; font-weight: 600;")
            row_w.addWidget(lbl)
            val_lbl = QLabel(str(value))
            val_lbl.setTextFormat(Qt.RichText)
            val_lbl.setStyleSheet(f"color: {self._c_text}; font-size: 11px;")
            row_w.addWidget(val_lbl, 1)
            mem_layout.addLayout(row_w)

        cl.addWidget(mem_frame)

        # ---- Message Stats ----
        stats_frame = self._make_section("Activity in This Group")
        sl = stats_frame.layout()

        total = msg_stats.get("total") or 0
        stats_grid = QHBoxLayout()
        stats_grid.setSpacing(8)

        stat_items = [
            ("Messages", str(total), self._c_accent),
            ("Text", str(msg_stats.get("text_msgs") or 0), "#42a5f5"),
            ("Images", str(msg_stats.get("images") or 0), "#66bb6a"),
            ("Videos", str(msg_stats.get("videos") or 0), "#ef5350"),
            ("Audio", str(msg_stats.get("audio") or 0), "#ffa726"),
            ("Docs", str(msg_stats.get("documents") or 0), "#ab47bc"),
            ("Links", str(link_count), "#29b6f6"),
            ("Stickers", str(msg_stats.get("stickers") or 0), "#78909c"),
            ("Forwards", str(msg_stats.get("forwards") or 0), "#8d6e63"),
            ("Edits", str(msg_stats.get("edits") or 0), "#ffa726"),
            ("Deletes", str(msg_stats.get("deletes") or 0), "#ef5350"),
        ]

        for stat_label, stat_val, stat_color in stat_items:
            if stat_val == "0" and stat_label not in ("Messages", "Text"):
                continue
            card = QFrame()
            card.setFixedSize(90, 48)
            card.setStyleSheet(f"""
                QFrame {{ background: {self._c_bg3}; border: 1px solid {self._c_border2};
                          border-radius: 6px; }}
            """)
            card_l = QVBoxLayout(card)
            card_l.setContentsMargins(6, 4, 6, 4)
            card_l.setSpacing(1)
            v_lbl = QLabel(stat_val)
            v_lbl.setAlignment(Qt.AlignCenter)
            v_lbl.setStyleSheet(f"color: {stat_color}; font-size: 13px; font-weight: bold;")
            card_l.addWidget(v_lbl)
            s_lbl = QLabel(stat_label)
            s_lbl.setAlignment(Qt.AlignCenter)
            s_lbl.setStyleSheet(f"color: {self._c_text3}; font-size: 8px;")
            card_l.addWidget(s_lbl)
            stats_grid.addWidget(card)

        stats_grid.addStretch()
        sl.addLayout(stats_grid)

        # Time range
        first_ts = msg_stats.get("first_msg_ts")
        last_ts = msg_stats.get("last_msg_ts")
        if first_ts and last_ts:
            try:
                first_str = _fmt_ts(first_ts, "minute")
                last_str = _fmt_ts(last_ts, "minute")
                range_lbl = QLabel(f"First message: {first_str}  •  Last message: {last_str}")
                range_lbl.setStyleSheet(f"color: {self._c_text3}; font-size: 10px; margin-top: 6px;")
                sl.addWidget(range_lbl)
            except (ValueError, OSError):
                pass

        cl.addWidget(stats_frame)

        # ---- Hourly Activity Chart ----
        if hourly_map:
            hour_frame = self._make_section("Hourly Activity Pattern")
            hl = hour_frame.layout()

            chart_widget = QWidget()
            chart_widget.setFixedHeight(90)
            chart_layout = QHBoxLayout(chart_widget)
            chart_layout.setContentsMargins(0, 0, 0, 0)
            chart_layout.setSpacing(2)

            max_h = max(hourly_map.values()) if hourly_map else 1
            for h in range(24):
                cnt = hourly_map.get(h, 0)
                pct = int((cnt / max_h * 100)) if max_h else 0

                bar_w = QWidget()
                bar_l = QVBoxLayout(bar_w)
                bar_l.setContentsMargins(0, 0, 0, 0)
                bar_l.setSpacing(1)

                # Spacer to push bar down
                spacer = QWidget()
                spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                bar_l.addWidget(spacer)

                bar = QFrame()
                bar_height = max(2, int(60 * pct / 100))
                bar.setFixedHeight(bar_height)
                bar.setStyleSheet(f"""
                    QFrame {{
                        background: qlineargradient(y1:0, y2:1, stop:0 {self._c_accent}, stop:1 #00897b);
                        border-radius: 2px;
                    }}
                """)
                bar.setToolTip(f"{h:02d}:00 — {cnt} messages")
                bar_l.addWidget(bar)

                hour_lbl = QLabel(f"{h:02d}")
                hour_lbl.setAlignment(Qt.AlignCenter)
                hour_lbl.setStyleSheet(f"color: {self._c_text3}; font-size: 8px;")
                bar_l.addWidget(hour_lbl)

                chart_layout.addWidget(bar_w)

            hl.addWidget(chart_widget)
            cl.addWidget(hour_frame)

        # ---- Mentions Section ----
        if mentions_given or mentions_received:
            mention_frame = self._make_section("Mention Relationships")
            ml = mention_frame.layout()

            mention_grid = QHBoxLayout()
            mention_grid.setSpacing(16)

            # Mentions given
            given_w = QWidget()
            given_l = QVBoxLayout(given_w)
            given_l.setContentsMargins(0, 0, 0, 0)
            given_l.setSpacing(4)
            given_hdr = QLabel(f"<b>Mentions Others</b> <span style='color:{self._c_text3};'>({len(mentions_given)})</span>")
            given_hdr.setTextFormat(Qt.RichText)
            given_hdr.setStyleSheet(f"color: {self._c_text}; font-size: 11px;")
            given_l.addWidget(given_hdr)
            for mg in mentions_given:
                name = mg["mentioned_name"] or "Unknown"
                jid = mg["mentioned_jid"] or ""
                cnt = mg["cnt"]
                row_html = f"<b>{_escape_html(name)}</b>"
                if jid:
                    row_html += f" <span style='color:{self._c_accent};font-size:9px;'>{_escape_html(jid)}</span>"
                row_html += f" <span style='color:{self._c_text3};'>×{cnt}</span>"
                r_lbl = QLabel(row_html)
                r_lbl.setTextFormat(Qt.RichText)
                r_lbl.setStyleSheet(f"color: {self._c_text}; font-size: 10px;")
                r_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
                given_l.addWidget(r_lbl)
            given_l.addStretch()
            mention_grid.addWidget(given_w, 1)

            # Mentions received
            recv_w = QWidget()
            recv_l = QVBoxLayout(recv_w)
            recv_l.setContentsMargins(0, 0, 0, 0)
            recv_l.setSpacing(4)
            recv_hdr = QLabel(f"<b>Mentioned By</b> <span style='color:{self._c_text3};'>({len(mentions_received)})</span>")
            recv_hdr.setTextFormat(Qt.RichText)
            recv_hdr.setStyleSheet(f"color: {self._c_text}; font-size: 11px;")
            recv_l.addWidget(recv_hdr)
            for mr in mentions_received:
                name = mr["mentioner_name"] or "Unknown"
                jid = mr["mentioner_jid"] or ""
                cnt = mr["cnt"]
                row_html = f"<b>{_escape_html(name)}</b>"
                if jid:
                    row_html += f" <span style='color:{self._c_accent};font-size:9px;'>{_escape_html(jid)}</span>"
                row_html += f" <span style='color:{self._c_text3};'>×{cnt}</span>"
                r_lbl = QLabel(row_html)
                r_lbl.setTextFormat(Qt.RichText)
                r_lbl.setStyleSheet(f"color: {self._c_text}; font-size: 10px;")
                r_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
                recv_l.addWidget(r_lbl)
            recv_l.addStretch()
            mention_grid.addWidget(recv_w, 1)

            ml.addLayout(mention_grid)
            cl.addWidget(mention_frame)

        # ---- Link Domains ----
        if link_domains:
            link_frame = self._make_section(f"Top Link Domains ({link_count} total links)")
            ll = link_frame.layout()
            for ld in link_domains:
                domain = ld["domain"] or "?"
                cnt = ld["cnt"]
                d_lbl = QLabel(
                    f"<code style='background:{self._c_bg3};padding:1px 4px;border-radius:2px;"
                    f"font-size:10px;color:{self._c_accent};'>{_escape_html(domain)}</code>"
                    f" <span style='color:{self._c_text3};'>×{cnt}</span>"
                )
                d_lbl.setTextFormat(Qt.RichText)
                d_lbl.setStyleSheet(f"color: {self._c_text}; font-size: 11px;")
                ll.addWidget(d_lbl)
            cl.addWidget(link_frame)

        # ---- Reactions ----
        if reactions_given or reactions_received:
            react_frame = self._make_section("Reactions")
            rl = react_frame.layout()

            react_grid = QHBoxLayout()
            react_grid.setSpacing(16)

            if reactions_given:
                given_html = " ".join(
                    f"<span style='background:{self._c_bg3};padding:3px 8px;border-radius:12px;"
                    f"font-size:14px;'>{r['emoji']} <small style='color:{self._c_text3};'>×{r['cnt']}</small></span>"
                    for r in reactions_given
                )
                g_lbl = QLabel(f"<b style='font-size:10px;color:{self._c_text2};'>Given:</b> {given_html}")
                g_lbl.setTextFormat(Qt.RichText)
                g_lbl.setWordWrap(True)
                react_grid.addWidget(g_lbl, 1)

            if reactions_received:
                recv_html = " ".join(
                    f"<span style='background:{self._c_bg3};padding:3px 8px;border-radius:12px;"
                    f"font-size:14px;'>{r['emoji']} <small style='color:{self._c_text3};'>×{r['cnt']}</small></span>"
                    for r in reactions_received
                )
                r_lbl = QLabel(f"<b style='font-size:10px;color:{self._c_text2};'>Received:</b> {recv_html}")
                r_lbl.setTextFormat(Qt.RichText)
                r_lbl.setWordWrap(True)
                react_grid.addWidget(r_lbl, 1)

            rl.addLayout(react_grid)
            cl.addWidget(react_frame)

        cl.addStretch()
        scroll.setWidget(content)
        main_layout.addWidget(scroll, 1)

    def _make_section(self, title: str) -> QFrame:
        """Create a styled section frame with a title."""
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background: {self._c_bg2};
                border: 1px solid {self._c_border};
                border-radius: 8px;
            }}
        """)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(6)

        hdr = QLabel(title)
        hdr.setStyleSheet(f"""
            color: {self._c_text}; font-size: 12px; font-weight: bold;
            padding-bottom: 4px; border-bottom: 1px solid {self._c_border2};
        """)
        layout.addWidget(hdr)
        return frame

    def _gen_report(self, contact_id: int, member_name: str,
                    conversation_id: int, group_name: str) -> None:
        """Generate and open a group-scoped contact report."""
        import webbrowser
        from app.views.pages._report_loader import load_contact_report
        generate_contact_report = load_contact_report()

        db = Database.get()
        db_path = db.path

        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in member_name)[:40]
        safe_group = "".join(c if c.isalnum() or c in " _-" else "_" for c in group_name)[:30]

        output_dir = db_path.parent / "reports"
        output_dir.mkdir(exist_ok=True)
        output_file = output_dir / f"contact_report_{safe_name}_in_{safe_group}_{contact_id}.html"

        try:
            result_path = generate_contact_report(
                analysis_db_path=str(db_path),
                contact_id=contact_id,
                output_path=str(output_file),
                group_conversation_id=conversation_id,
            )
            webbrowser.open(result_path.as_uri())
        except Exception as e:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(
                self, "Report Failed",
                f"Failed to generate report:\n\n{e}",
            )


def _escape_html(text: str) -> str:
    """Minimal HTML escape for safe display in QLabel rich text."""
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br>")
    )
